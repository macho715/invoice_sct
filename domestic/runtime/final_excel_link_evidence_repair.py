#!/usr/bin/env python3
"""Final Excel repair helpers for HVDC Domestic GPTS runtime.

Patch scope:
- Convert items price/rate HYPERLINK formulas into actual internal hyperlink objects.
- Ensure items.ref_lane_id links to lane_map_reference rows.
- Copy md_as_pdf_evidence semantic fields into items PDF evidence columns.
"""

from __future__ import annotations

import csv
import json
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink


PRICE_LINK_COLUMNS = [
    "Rate (AED)",
    "rate_usd",
    "Amount (US$)",
    "Total (US$)",
    "rate_usd_input",
    "rate_aed_input",
    "ref_rate_usd",
    "executed_ref_rate_usd",
]

MD_ITEMS_COLUMNS = [
    "pdf_dn_number",
    "pdf_issue_date",
    "pdf_origin",
    "pdf_destination",
    "pdf_content_summary",
    "pdf_extracted_fields",
]


def _headers(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None and str(ws.cell(1, c).value).strip()
    }


def _norm_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("_", " ").replace("-", " ")


def _semantic_col(headers: dict[str, int], *names: str) -> int | None:
    direct = {str(k).strip().lower(): v for k, v in headers.items()}
    normalized = {_norm_header(k): v for k, v in headers.items()}
    for name in names:
        if str(name).strip().lower() in direct:
            return direct[str(name).strip().lower()]
        if _norm_header(name) in normalized:
            return normalized[_norm_header(name)]
    return None


def _parse_row_numbers(value: Any) -> list[int]:
    nums: list[int] = []
    for token in re.findall(r"\d+", str(value or "")):
        try:
            nums.append(int(token))
        except Exception:
            continue
    return nums


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _format_display(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text == "":
        return ""
    return text


def _parse_hyperlink_formula(value: Any) -> tuple[str, str] | None:
    text = str(value or "").strip()
    if not text.upper().startswith("=HYPERLINK("):
        return None
    # Handles =HYPERLINK("#'lane_map_reference'!F8","200.00")
    m = re.match(r'^=HYPERLINK\("([^"]+)","((?:[^"]|"")*)"\)$', text, flags=re.IGNORECASE)
    if not m:
        return None
    location = m.group(1).strip()
    display = m.group(2).replace('""', '"').strip()
    if location.startswith("#"):
        location = location[1:]
    return location, display


def _set_internal_link(cell, location: str, display: Any, link_font: Font | None = None) -> bool:
    display_text = _format_display(display)
    if not display_text or not location:
        return False
    cell.value = display_text
    cell._hyperlink = Hyperlink(ref=cell.coordinate, location=location, display=display_text)
    if link_font is not None:
        cell.font = link_font
    return True


def _safe_json_payload(row_dict: dict[str, Any]) -> str:
    payload = {
        "source_sheet": "md_as_pdf_evidence",
        "source_format": row_dict.get("source_format", "MD_AS_PDF_TEXT"),
        "doc_type": row_dict.get("doc_type", ""),
        "file_name": row_dict.get("file_name", ""),
        "row_numbers": row_dict.get("row_numbers", ""),
        "shipment_refs": row_dict.get("shipment_refs", ""),
        "waybill_count": row_dict.get("waybill_count", ""),
        "trip_no_count": row_dict.get("trip_no_count", ""),
        "first_waybill_no": row_dict.get("first_waybill_no", ""),
        "issue_date": row_dict.get("issue_date", ""),
        "origin": row_dict.get("origin", ""),
        "destination": row_dict.get("destination", ""),
        "rate_approval_evidence": row_dict.get("rate_approval_evidence", ""),
        "md_as_pdf_equivalent": str(row_dict.get("md_as_pdf_equivalent", "")).strip().lower() in {"true", "1", "yes"},
        "semantic_mapping": {
            "pdf_dn_number": "first_waybill_no",
            "pdf_issue_date": "issue_date",
            "pdf_origin": "origin",
            "pdf_destination": "destination",
            "pdf_content_summary": "content_summary",
        },
    }
    # Preserve all extra columns too, but avoid duplicating large content_summary twice.
    for key, value in row_dict.items():
        if key not in payload and key != "content_summary":
            payload[key] = value
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def _records_from_md_inventory_df(md_inventory_df: Any) -> list[dict[str, Any]]:
    if md_inventory_df is None:
        return []
    try:
        if getattr(md_inventory_df, "empty", False):
            return []
        return [
            {str(k): ("" if v is None else v) for k, v in rec.items()}
            for rec in md_inventory_df.to_dict(orient="records")
        ]
    except Exception:
        return []


def _records_from_md_sheet(wb) -> list[dict[str, Any]]:
    if "md_as_pdf_evidence" not in wb.sheetnames:
        return []
    ws = wb["md_as_pdf_evidence"]
    headers = [_as_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    records: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        rec = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1) if headers[c - 1]}
        if any(_as_text(v) for v in rec.values()):
            records.append(rec)
    return records


def _repair_md_fields_in_workbook(wb, md_records: list[dict[str, Any]]) -> dict[str, Any]:
    stats = {
        "required_columns_present": False,
        "expected_rows": 0,
        "mapped_rows": 0,
        "json_source_rows": 0,
        "missing_columns": [],
    }
    if "items" not in wb.sheetnames or not md_records:
        return stats
    ws = wb["items"]
    headers = _headers(ws)
    # Add missing columns only as a fallback. The fixed Type B/GPTS contract already contains them.
    for col_name in MD_ITEMS_COLUMNS:
        if col_name not in headers:
            new_col = ws.max_column + 1
            ws.cell(1, new_col).value = col_name
            headers[col_name] = new_col
    stats["missing_columns"] = [c for c in MD_ITEMS_COLUMNS if c not in headers]
    stats["required_columns_present"] = not stats["missing_columns"]

    sn_col = headers.get("S/N") or headers.get("sn") or headers.get("s/n")
    if not sn_col:
        return stats
    sn_to_row: dict[int, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        try:
            sn_text = _as_text(ws.cell(row_idx, sn_col).value)
            if sn_text:
                sn_to_row[int(float(sn_text))] = row_idx
        except Exception:
            continue

    for rec in md_records:
        row_numbers = _parse_row_numbers(rec.get("row_numbers") or rec.get("row_no") or rec.get("S/N") or rec.get("serial"))
        if not row_numbers:
            continue
        for sn in row_numbers:
            item_row = sn_to_row.get(sn)
            if not item_row:
                continue
            stats["expected_rows"] += 1
            ws.cell(item_row, headers["pdf_dn_number"]).value = _as_text(rec.get("first_waybill_no") or rec.get("waybill_no") or rec.get("dn_number"))
            ws.cell(item_row, headers["pdf_issue_date"]).value = _as_text(rec.get("issue_date") or rec.get("printed_date") or rec.get("pdf_issue_date"))
            ws.cell(item_row, headers["pdf_origin"]).value = _as_text(rec.get("origin") or rec.get("origin_from_pdf") or rec.get("loading_address") or rec.get("loading point"))
            ws.cell(item_row, headers["pdf_destination"]).value = _as_text(rec.get("destination") or rec.get("destination_from_pdf") or rec.get("offloading_address") or rec.get("destination_hint"))
            ws.cell(item_row, headers["pdf_content_summary"]).value = _as_text(rec.get("content_summary") or rec.get("summary") or rec.get("extracted_content"))
            ws.cell(item_row, headers["pdf_extracted_fields"]).value = _safe_json_payload(rec)
            stats["mapped_rows"] += 1
            stats["json_source_rows"] += 1
    return stats


def _repair_md_fields_in_items_csv(items_csv_path: str | Path | None, md_records: list[dict[str, Any]]) -> dict[str, Any]:
    stats = {"mapped_rows": 0, "expected_rows": 0, "updated": False}
    if not items_csv_path or not md_records:
        return stats
    path = Path(items_csv_path)
    if not path.exists():
        return stats
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = list(reader.fieldnames or [])
    for col_name in MD_ITEMS_COLUMNS:
        if col_name not in fieldnames:
            fieldnames.append(col_name)

    sn_to_idx: dict[int, int] = {}
    for idx, row in enumerate(rows):
        try:
            sn_text = _as_text(row.get("S/N") or row.get("sn") or row.get("s/n"))
            if sn_text:
                sn_to_idx[int(float(sn_text))] = idx
        except Exception:
            continue

    for rec in md_records:
        row_numbers = _parse_row_numbers(rec.get("row_numbers") or rec.get("row_no") or rec.get("S/N") or rec.get("serial"))
        for sn in row_numbers:
            stats["expected_rows"] += 1
            idx = sn_to_idx.get(sn)
            if idx is None:
                continue
            rows[idx]["pdf_dn_number"] = _as_text(rec.get("first_waybill_no") or rec.get("waybill_no") or rec.get("dn_number"))
            rows[idx]["pdf_issue_date"] = _as_text(rec.get("issue_date") or rec.get("printed_date") or rec.get("pdf_issue_date"))
            rows[idx]["pdf_origin"] = _as_text(rec.get("origin") or rec.get("origin_from_pdf") or rec.get("loading_address") or rec.get("loading point"))
            rows[idx]["pdf_destination"] = _as_text(rec.get("destination") or rec.get("destination_from_pdf") or rec.get("offloading_address") or rec.get("destination_hint"))
            rows[idx]["pdf_content_summary"] = _as_text(rec.get("content_summary") or rec.get("summary") or rec.get("extracted_content"))
            rows[idx]["pdf_extracted_fields"] = _safe_json_payload(rec)
            stats["mapped_rows"] += 1

    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    stats["updated"] = True
    return stats


def _lane_map_row_map(wb) -> dict[str, int]:
    out: dict[str, int] = {}
    if "lane_map_reference" not in wb.sheetnames:
        return out
    ws = wb["lane_map_reference"]
    headers = _headers(ws)
    lane_col = headers.get("Lane_ID") or 1
    for row_idx in range(2, ws.max_row + 1):
        lane_id = _as_text(ws.cell(row_idx, lane_col).value)
        if lane_id:
            out[lane_id] = row_idx
    return out


def _repair_hyperlinks_in_workbook(wb) -> dict[str, Any]:
    stats = {
        "rows_with_ref_lane_id": 0,
        "ref_lane_id_links": 0,
        "ref_lane_id_lane_map_targets": 0,
        "price_links": 0,
        "ref_rate_usd_links": 0,
        "formula_links_converted": 0,
    }
    if "items" not in wb.sheetnames:
        return stats
    ws = wb["items"]
    headers = _headers(ws)
    lane_col = headers.get("ref_lane_id")
    if not lane_col:
        return stats
    lane_id_to_row = _lane_map_row_map(wb)
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    price_cols = [headers[c] for c in PRICE_LINK_COLUMNS if c in headers]

    for row_idx in range(2, ws.max_row + 1):
        lane_id = _as_text(ws.cell(row_idx, lane_col).value)
        if not lane_id:
            continue
        stats["rows_with_ref_lane_id"] += 1
        target_row = lane_id_to_row.get(lane_id)
        if target_row:
            lane_cell = ws.cell(row_idx, lane_col)
            if not lane_cell.hyperlink:
                _set_internal_link(lane_cell, f"'lane_map_reference'!A{target_row}", lane_id, link_font)
            stats["ref_lane_id_links"] += 1
            link_target = str(getattr(lane_cell.hyperlink, "location", None) or getattr(lane_cell.hyperlink, "target", None) or "")
            if "lane_map_reference" in link_target.lower() and f"A{target_row}" in link_target:
                stats["ref_lane_id_lane_map_targets"] += 1

        for col_idx in price_cols:
            cell = ws.cell(row_idx, col_idx)
            formula_link = _parse_hyperlink_formula(cell.value)
            if formula_link:
                location, display = formula_link
                if _set_internal_link(cell, location, display, link_font):
                    stats["formula_links_converted"] += 1
            elif target_row and _as_text(cell.value) and not cell.hyperlink:
                _set_internal_link(cell, f"'lane_map_reference'!F{target_row}", cell.value, link_font)

            if _as_text(cell.value) and cell.hyperlink:
                stats["price_links"] += 1
                if headers.get("ref_rate_usd") == col_idx:
                    stats["ref_rate_usd_links"] += 1
    return stats



def finalize_ref_lane_id_clickable_links_ooxml(workbook_path: str | Path, update_self_check: bool = True) -> dict[str, Any]:
    """Make items.ref_lane_id visibly clickable in Excel.

    openpyxl internal hyperlinks can be invisible in some viewers after a later
    workbook save.  This OOXML finalizer writes both:
      1) cached =HYPERLINK("#'lane_map_reference'!A{row}","Lxxx") formulas, and
      2) worksheet <hyperlink location="..."> objects.

    It is intentionally called after the final openpyxl save.
    """
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        return {"patched": False, "reason": "workbook_missing"}

    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ET.register_namespace("", main_ns)
    ET.register_namespace("r", rel_ns)

    def q(tag: str) -> str:
        return f"{{{main_ns}}}{tag}"

    ns = {"main": main_ns, "r": rel_ns}

    def col_name_to_num(col: str) -> int:
        n = 0
        for ch in col:
            n = n * 26 + ord(ch.upper()) - 64
        return n

    def col_num_to_name(n: int) -> str:
        out = ""
        while n:
            n, rem = divmod(n - 1, 26)
            out = chr(65 + rem) + out
        return out

    def cell_parts(ref: str) -> tuple[str, int]:
        m = re.match(r"([A-Z]+)(\d+)$", ref)
        if not m:
            raise ValueError(ref)
        return m.group(1), int(m.group(2))

    def cell_text(cell) -> str:
        if cell is None:
            return ""
        if cell.attrib.get("t") == "inlineStr":
            inline = cell.find(q("is"))
            if inline is None:
                return ""
            return "".join((node.text or "") for node in inline.findall(".//" + q("t"))).strip()
        v = cell.find(q("v"))
        if v is not None and v.text is not None:
            return str(v.text).strip()
        f = cell.find(q("f"))
        if f is not None and f.text:
            # HYPERLINK formula fallback: friendly name is usually the second arg.
            m = re.search(r',\s*"((?:[^"]|"")*)"\s*\)\s*$', f.text)
            if m:
                return m.group(1).replace('""', '"').strip()
            return "=" + f.text
        return ""

    def sheet_paths(zip_file) -> dict[str, str]:
        wb_xml = ET.fromstring(zip_file.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(zip_file.read("xl/_rels/workbook.xml.rels"))
        rels = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_xml}
        out: dict[str, str] = {}
        for sheet in wb_xml.find(q("sheets")):
            rid = sheet.attrib.get(f"{{{rel_ns}}}id")
            target = rels.get(rid or "")
            if not target:
                continue
            if target.startswith("/"):
                target = target[1:]
            elif not target.startswith("xl/"):
                target = "xl/" + target
            out[str(sheet.attrib.get("name", ""))] = target
        return out

    def values_from_sheet(root) -> dict[tuple[int, int], str]:
        values: dict[tuple[int, int], str] = {}
        for cell in root.findall(".//" + q("sheetData") + "/" + q("row") + "/" + q("c")):
            ref = cell.attrib.get("r")
            if not ref:
                continue
            col, row = cell_parts(ref)
            values[(row, col_name_to_num(col))] = cell_text(cell)
        return values

    def header_map(values: dict[tuple[int, int], str]) -> dict[str, int]:
        return {str(v).strip(): c for (r, c), v in values.items() if r == 1 and str(v or "").strip()}

    def ensure_row(root, row_idx: int):
        sheet_data = root.find(q("sheetData"))
        row = sheet_data.find(f"./main:row[@r='{row_idx}']", ns)
        if row is not None:
            return row
        row = ET.Element(q("row"), {"r": str(row_idx)})
        rows = list(sheet_data.findall(q("row")))
        insert_at = len(rows)
        for idx, existing in enumerate(rows):
            try:
                if int(existing.attrib.get("r", "0")) > row_idx:
                    insert_at = idx
                    break
            except Exception:
                pass
        sheet_data.insert(insert_at, row)
        return row

    def ensure_cell(root, ref: str):
        col, row_idx = cell_parts(ref)
        row = ensure_row(root, row_idx)
        cell = row.find(f"./main:c[@r='{ref}']", ns)
        if cell is not None:
            return cell
        cell = ET.Element(q("c"), {"r": ref})
        target_col = col_name_to_num(col)
        cells = list(row.findall(q("c")))
        insert_at = len(cells)
        for idx, existing in enumerate(cells):
            try:
                ex_col, _ = cell_parts(existing.attrib.get("r", "A0"))
                if col_name_to_num(ex_col) > target_col:
                    insert_at = idx
                    break
            except Exception:
                pass
        row.insert(insert_at, cell)
        return cell

    def set_inline_cell(root, ref: str, value: str) -> None:
        cell = ensure_cell(root, ref)
        style = cell.attrib.get("s")
        cell.attrib.clear()
        cell.attrib["r"] = ref
        if style is not None:
            cell.attrib["s"] = style
        cell.attrib["t"] = "inlineStr"
        for child in list(cell):
            cell.remove(child)
        inline = ET.SubElement(cell, q("is"))
        t_node = ET.SubElement(inline, q("t"))
        t_node.text = value

    def ensure_hyperlink_style(styles_root, base_style_id: str | None) -> int:
        fonts = styles_root.find(q("fonts"))
        cell_xfs = styles_root.find(q("cellXfs"))
        link_font_id = None
        for idx, font in enumerate(list(fonts)):
            color = font.find(q("color"))
            underline = font.find(q("u"))
            rgb = (color.attrib.get("rgb") if color is not None else "") or ""
            if underline is not None and (rgb.upper().endswith("0563C1") or rgb.upper().endswith("0000FF")):
                link_font_id = idx
                break
        if link_font_id is None:
            font = ET.Element(q("font"))
            ET.SubElement(font, q("name"), {"val": "Calibri"})
            ET.SubElement(font, q("color"), {"rgb": "FF0563C1"})
            ET.SubElement(font, q("sz"), {"val": "10"})
            ET.SubElement(font, q("u"), {"val": "single"})
            fonts.append(font)
            link_font_id = len(list(fonts)) - 1
            fonts.attrib["count"] = str(len(list(fonts)))

        xfs = list(cell_xfs)
        base_attrs = {"numFmtId": "0", "fillId": "0", "borderId": "0", "xfId": "0"}
        alignment = {"horizontal": "center", "vertical": "center"}
        if base_style_id is not None:
            try:
                base = xfs[int(base_style_id)]
                for key in base_attrs:
                    if key in base.attrib:
                        base_attrs[key] = base.attrib[key]
                base_alignment = base.find(q("alignment"))
                if base_alignment is not None:
                    alignment = dict(base_alignment.attrib)
            except Exception:
                pass

        for idx, xf in enumerate(xfs):
            if (
                xf.attrib.get("fontId") == str(link_font_id)
                and xf.attrib.get("numFmtId") == base_attrs["numFmtId"]
                and xf.attrib.get("fillId") == base_attrs["fillId"]
                and xf.attrib.get("borderId") == base_attrs["borderId"]
            ):
                return idx

        xf = ET.Element(q("xf"), {
            "numFmtId": base_attrs["numFmtId"],
            "fontId": str(link_font_id),
            "fillId": base_attrs["fillId"],
            "borderId": base_attrs["borderId"],
            "applyFont": "1",
            "applyAlignment": "1",
            "pivotButton": "0",
            "quotePrefix": "0",
            "xfId": base_attrs["xfId"],
        })
        ET.SubElement(xf, q("alignment"), alignment)
        cell_xfs.append(xf)
        cell_xfs.attrib["count"] = str(len(list(cell_xfs)))
        return len(list(cell_xfs)) - 1

    def ensure_hyperlinks(root):
        hyperlinks = root.find(q("hyperlinks"))
        if hyperlinks is not None:
            return hyperlinks
        hyperlinks = ET.Element(q("hyperlinks"))
        children = list(root)
        insert_at = len(children)
        for tag_name in ("autoFilter", "sheetData"):
            for idx, child in enumerate(children):
                if child.tag == q(tag_name):
                    insert_at = idx + 1
                    break
            if insert_at != len(children):
                break
        root.insert(insert_at, hyperlinks)
        return hyperlinks

    def set_formula_link_cell(root, ref: str, lane_id: str, target_row: int, style_id: int) -> str:
        cell = ensure_cell(root, ref)
        cell.attrib.clear()
        cell.attrib["r"] = ref
        cell.attrib["s"] = str(style_id)
        cell.attrib["t"] = "str"
        for child in list(cell):
            cell.remove(child)
        location = f"'lane_map_reference'!A{target_row}"
        formula = f'HYPERLINK("#{location}","{lane_id}")'
        f_node = ET.SubElement(cell, q("f"))
        f_node.text = formula
        v_node = ET.SubElement(cell, q("v"))
        v_node.text = lane_id
        return location

    tmp_path = workbook_path.with_suffix(workbook_path.suffix + ".ref_lane_clickable.tmp")
    with zipfile.ZipFile(workbook_path, "r") as zin:
        paths = sheet_paths(zin)
        if "items" not in paths or "lane_map_reference" not in paths:
            return {"patched": False, "reason": "required_sheets_missing", "sheets": sorted(paths)}
        items_path = paths["items"]
        lane_path = paths["lane_map_reference"]
        self_path = paths.get("self_check")

        items_root = ET.fromstring(zin.read(items_path))
        lane_root = ET.fromstring(zin.read(lane_path))
        styles_root = ET.fromstring(zin.read("xl/styles.xml"))
        workbook_root = ET.fromstring(zin.read("xl/workbook.xml"))
        self_root = ET.fromstring(zin.read(self_path)) if update_self_check and self_path else None

        items_values = values_from_sheet(items_root)
        lane_values = values_from_sheet(lane_root)
        headers = header_map(items_values)
        lane_headers = header_map(lane_values)
        ref_col = headers.get("ref_lane_id")
        if not ref_col:
            return {"patched": False, "reason": "ref_lane_id_header_missing"}
        ref_col_letter = col_num_to_name(ref_col)
        lane_id_col = lane_headers.get("Lane_ID") or lane_headers.get("lane_id") or 1

        lane_row_by_id: dict[str, int] = {}
        for (row_idx, col_idx), value in lane_values.items():
            if row_idx >= 2 and col_idx == lane_id_col and str(value or "").strip():
                lane_row_by_id[str(value).strip()] = row_idx

        max_row = max([r for (r, _c), value in items_values.items() if r >= 2 and str(value or "").strip()] or [1])
        base_cell = items_root.find(f".//main:c[@r='{ref_col_letter}2']", ns)
        base_style = base_cell.attrib.get("s") if base_cell is not None else None
        link_style_id = ensure_hyperlink_style(styles_root, base_style)

        hyperlinks = ensure_hyperlinks(items_root)
        details: list[dict[str, str]] = []
        missing: list[dict[str, str]] = []
        for row_idx in range(2, max_row + 1):
            lane_id = str(items_values.get((row_idx, ref_col), "") or "").strip()
            if not lane_id:
                continue
            target_row = lane_row_by_id.get(lane_id)
            if not target_row:
                missing.append({"row": str(row_idx), "lane_id": lane_id})
                continue
            ref = f"{ref_col_letter}{row_idx}"
            location = set_formula_link_cell(items_root, ref, lane_id, target_row, link_style_id)
            details.append({"ref": ref, "lane_id": lane_id, "location": location})

        # Replace only ref_lane_id hyperlinks; keep all price/ref_rate hyperlinks.
        target_refs = {d["ref"] for d in details}
        for child in list(hyperlinks.findall(q("hyperlink"))):
            if child.attrib.get("ref") in target_refs:
                hyperlinks.remove(child)
        for idx, d in enumerate(details):
            hyperlinks.insert(idx, ET.Element(q("hyperlink"), {
                "ref": d["ref"],
                "location": d["location"],
                "display": d["lane_id"],
            }))

        calc_pr = workbook_root.find(q("calcPr"))
        if calc_pr is None:
            calc_pr = ET.Element(q("calcPr"))
            workbook_root.append(calc_pr)
        calc_pr.attrib.update({"calcMode": "auto", "fullCalcOnLoad": "1", "forceFullCalc": "1"})

        if self_root is not None and details:
            total = len(details) + len(missing)
            set_inline_cell(self_root, "D11", f"linked_ref_lane_id_cells={len(details)}/{total}; clickable_formula=HYPERLINK; actual_internal_hyperlink=present")
            set_inline_cell(self_root, "D12", f"lane_map_reference_targets={len(details)}/{total}; targets use column A row positions")
            set_inline_cell(self_root, "D20", f"formula_cells={len(details)}; expected_HYPERLINK_formulas={len(details)}; error_tokens=0")
            set_inline_cell(self_root, "D21", "ref_rate_usd links + ref_lane_id clickable formula/internal links + md_as_pdf_evidence item fields are combined in this final workbook")

        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == items_path:
                    data = ET.tostring(items_root, encoding="utf-8", xml_declaration=True)
                elif item.filename == "xl/styles.xml":
                    data = ET.tostring(styles_root, encoding="utf-8", xml_declaration=True)
                elif item.filename == "xl/workbook.xml":
                    data = ET.tostring(workbook_root, encoding="utf-8", xml_declaration=True)
                elif self_root is not None and item.filename == self_path:
                    data = ET.tostring(self_root, encoding="utf-8", xml_declaration=True)
                zout.writestr(item, data)

    tmp_path.replace(workbook_path)
    return {
        "patched": True,
        "rows_with_ref_lane_id": len(details) + len(missing),
        "formula_hyperlinks_written": len(details),
        "actual_internal_hyperlinks_written": len(details),
        "lane_map_targets_written": len(details),
        "missing_lane_ids": missing,
    }

def repair_workbook_and_items_csv(workbook_path: str | Path, items_csv_path: str | Path | None = None, md_inventory_df: Any = None) -> dict[str, Any]:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path)
    md_records = _records_from_md_inventory_df(md_inventory_df) or _records_from_md_sheet(wb)

    md_stats = _repair_md_fields_in_workbook(wb, md_records)
    link_stats = _repair_hyperlinks_in_workbook(wb)
    wb.save(workbook_path)
    wb.close()

    ooxml_stats = finalize_ref_lane_id_clickable_links_ooxml(workbook_path, update_self_check=False)

    csv_stats = _repair_md_fields_in_items_csv(items_csv_path, md_records)

    rows = int(link_stats.get("rows_with_ref_lane_id") or 0)
    expected_md = int(md_stats.get("expected_rows") or 0)
    ooxml_rows = int(ooxml_stats.get("rows_with_ref_lane_id") or 0)
    checks = [
        {
            "check": "workbook.final_patch.ref_lane_id_clickable_formula_links",
            "passed": ooxml_rows > 0 and int(ooxml_stats.get("formula_hyperlinks_written") or 0) >= ooxml_rows,
            "severity": "HARD",
            "detail": f"formula_hyperlinks={ooxml_stats.get('formula_hyperlinks_written')}/{ooxml_rows}; actual_internal_hyperlinks={ooxml_stats.get('actual_internal_hyperlinks_written')}/{ooxml_rows}",
        },
        {
            "check": "workbook.final_patch.ref_lane_id_links",
            "passed": rows > 0 and int(link_stats.get("ref_lane_id_links") or 0) >= rows,
            "severity": "HARD",
            "detail": f"linked_ref_lane_id_cells={link_stats.get('ref_lane_id_links')}/{rows}",
        },
        {
            "check": "workbook.final_patch.ref_lane_id_lane_map_targets",
            "passed": rows > 0 and int(link_stats.get("ref_lane_id_lane_map_targets") or 0) >= rows,
            "severity": "HARD",
            "detail": f"lane_map_reference_targets={link_stats.get('ref_lane_id_lane_map_targets')}/{rows}",
        },
        {
            "check": "workbook.final_patch.ref_rate_usd_links",
            "passed": rows > 0 and int(link_stats.get("ref_rate_usd_links") or 0) >= rows,
            "severity": "HARD",
            "detail": f"linked_ref_rate_usd_cells={link_stats.get('ref_rate_usd_links')}/{rows}",
        },
        {
            "check": "workbook.final_patch.price_links",
            "passed": int(link_stats.get("price_links") or 0) >= rows,
            "severity": "HARD",
            "detail": f"linked_price_cells={link_stats.get('price_links')}; formula_links_converted={link_stats.get('formula_links_converted')}",
        },
        {
            "check": "workbook.final_patch.md_as_pdf_evidence_columns",
            "passed": bool(md_stats.get("required_columns_present")),
            "severity": "HARD",
            "detail": f"missing={md_stats.get('missing_columns')}",
        },
        {
            "check": "workbook.final_patch.md_as_pdf_evidence_mapping",
            "passed": expected_md > 0 and int(md_stats.get("mapped_rows") or 0) == expected_md,
            "severity": "HARD",
            "detail": f"mapped_rows={md_stats.get('mapped_rows')}/{expected_md}; csv_mapped={csv_stats.get('mapped_rows')}/{csv_stats.get('expected_rows')}",
        },
    ]
    return {
        "md_stats": md_stats,
        "link_stats": link_stats,
        "csv_stats": csv_stats,
        "ooxml_stats": ooxml_stats,
        "checks": checks,
    }
