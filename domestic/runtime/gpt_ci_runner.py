#!/usr/bin/env python3
"""ChatGPT Code Interpreter runner for the HVDC Domestic invoice audit.

Autopatch 2026-06-07:
- Accepts uploaded Excel/CSV, text/Markdown invoice tables, and optional input URLs.
- Automatically stages uploaded PDF/MD POD files as supporting documents.
- Runs a self-verification loop before the workbook is returned to the user.
- Final all-patches repair: ref_rate_usd links, ref_lane_id lane_map links, and md_as_pdf_evidence fields in items.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import shutil
import subprocess
import sys
import urllib.parse
import urllib.request
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

try:
    from md_as_pdf_utils import (
        ROW_KEY_PREFIX,
        build_md_as_pdf_inventory,
        extract_hvdc_refs as _md_extract_hvdc_refs,
        extract_row_numbers_from_filename as _md_extract_row_numbers_from_filename,
        extract_waybill_numbers as _md_extract_waybill_numbers,
        extract_trip_numbers as _md_extract_trip_numbers,
        extract_rate_approval as _md_extract_rate_approval,
        read_text as _md_read_text,
    )
except Exception:
    ROW_KEY_PREFIX = "__ROW_SN__:"
    build_md_as_pdf_inventory = None
    _md_extract_hvdc_refs = None
    _md_extract_row_numbers_from_filename = None
    _md_extract_waybill_numbers = None
    _md_extract_trip_numbers = None
    _md_extract_rate_approval = None
    _md_read_text = None


KNOWN_SUPPORT_FILENAMES = {
    "approvedlanemap_enhanced.xlsx",
    "approvedlanemap_enhanced.json",
    "domestic_with_distances.xlsx",
    "domestic_rate_ledger.json",
    "domestic_rate_ledger.md",
    "site_alias_map.csv",
    "items.csv",
    "domestic_audit_proof_v2.json",
    "domestic_audit_run_log.txt",
    "gpts_run_summary.json",
    "gpts_run_summary.md",
    "final_md_pod_reconciliation_20260607.csv",
    "final_md_pod_reconciliation_20260607.md",
}

INVOICE_EXTENSIONS = {".xlsx", ".xls", ".csv", ".tsv", ".txt", ".md", ".markdown"}
SPREADSHEET_EXTENSIONS = {".xlsx", ".xls", ".csv", ".tsv"}
TEXT_TABLE_EXTENSIONS = {".txt", ".md", ".markdown"}
SUPPORTING_DOC_EXTENSIONS = {".pdf", ".md", ".markdown"}

INVOICE_KEYWORDS = [
    "s/n",
    "shipment",
    "shipment reference",
    "place of loading",
    "place of delivery",
    "loading",
    "delivery",
    "vehicle",
    "rate",
    "amount",
    "total",
    "trips",
    "# trips",
]
REQUIRED_INVOICE_CONCEPTS = {
    "shipment": ["shipment", "reference", "hvdc"],
    "lane": ["place of loading", "place of delivery", "origin", "destination", "loading", "delivery"],
    "vehicle": ["vehicle", "truck"],
    "rate": ["rate", "amount", "total"],
    "qty": ["# trips", "trips", "qty", "quantity"],
}
URL_RE = re.compile(r"https?://[^\s<>\"']+", re.IGNORECASE)


# CI host noise that can appear on stderr before user code runs.
# This is emitted by the ChatGPT spreadsheet artifact warmup layer, not by the
# Domestic audit code. It must not convert a complete audit run into "failed".
KNOWN_BENIGN_STDERR_BLOCKS = [
    (
        "artifact_tool_spreadsheet_warmup_hydrate_crdt",
        re.compile(
            r"Spreadsheet runtime warmup failed during python startup\s*"
            r"Traceback \(most recent call last\):\s*"
            r"(?:\s*File .*\n)+"
            r"artifact_tool\.rpc\.client\.RemoteError: "
            r"hydrateCrdtFromProto requires an empty collaborative document\.?\s*",
            flags=re.IGNORECASE | re.MULTILINE,
        ),
    ),
]


def split_known_benign_stderr(stderr_text: str) -> dict:
    """Separate known CI-host stderr noise from audit-relevant stderr."""
    raw = stderr_text or ""
    remaining = raw
    ignored: list[dict] = []
    for label, pattern in KNOWN_BENIGN_STDERR_BLOCKS:
        matches = list(pattern.finditer(remaining))
        if matches:
            ignored.append({"label": label, "count": len(matches)})
            remaining = pattern.sub("", remaining)

    non_benign = remaining.strip()
    return {
        "raw_stderr_nonblank": bool(raw.strip()),
        "ignored": ignored,
        "known_benign_only": bool(raw.strip() and ignored and not non_benign),
        "non_benign_stderr_nonblank": bool(non_benign),
        "non_benign_stderr": non_benign,
    }


def runtime_root_from(start: Path) -> Path:
    candidates = [
        start.resolve(),
        Path.cwd().resolve(),
        Path(__file__).resolve().parent,
        Path(__file__).resolve().parent.parent,
    ]
    for candidate in candidates:
        if (candidate / "patch" / "run_domestic_audit_v2.py").exists() and (
            candidate / "Data" / "DOMESTIC_with_distances.xlsx"
        ).exists():
            return candidate
    raise FileNotFoundError(
        "Could not locate runtime root. Expected patch/run_domestic_audit_v2.py "
        "and Data/DOMESTIC_with_distances.xlsx."
    )


def _safe_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.resolve().relative_to(parent.resolve())
        return True
    except Exception:
        return False


def _path_has_any_part(path: Path, names: set[str]) -> bool:
    return bool({part.lower() for part in path.parts} & names)


def _is_runtime_or_output_path(path: Path, runtime_root: Path) -> bool:
    lower_parts = {part.lower() for part in path.parts}
    if {"gpts_outputs", "results", "supporting_docs_staging", "input_staging", "__macosx", ".git", "__pycache__"} & lower_parts:
        return True
    if _safe_relative_to(path, runtime_root / "Data"):
        return True
    if _safe_relative_to(path, runtime_root / "SCNT HVDC Domestic Invoice v2.2 Generator"):
        return True
    return False


def _read_text_sample(path: Path, max_chars: int = 60000) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="replace")[:max_chars]
    except Exception:
        try:
            return path.read_text(encoding="latin-1", errors="replace")[:max_chars]
        except Exception:
            return ""


def _concept_score(text: str) -> int:
    low = text.lower()
    score = 0
    for variants in REQUIRED_INVOICE_CONCEPTS.values():
        if any(v in low for v in variants):
            score += 1
    return score


def _keyword_score(text: str) -> int:
    low = text.lower()
    return sum(1 for token in INVOICE_KEYWORDS if token in low)


def _looks_like_invoice_text(text: str) -> bool:
    if not text or len(text.strip()) < 20:
        return False
    return _concept_score(text) >= 4 and _keyword_score(text) >= 5


def _looks_like_invoice_spreadsheet(path: Path) -> bool:
    try:
        import pandas as pd

        if path.suffix.lower() in {".csv", ".tsv"}:
            sep = "\t" if path.suffix.lower() == ".tsv" else None
            if sep:
                df = pd.read_csv(path, sep=sep, nrows=12, dtype=str, encoding_errors="replace")
            else:
                # sep=None lets pandas detect comma/tab/semicolon in pasted exports.
                df = pd.read_csv(path, sep=None, engine="python", nrows=12, dtype=str, encoding_errors="replace")
            preview = " ".join(str(x) for x in list(df.columns) + df.head(3).astype(str).values.ravel().tolist())
            return _looks_like_invoice_text(preview)

        xl = pd.ExcelFile(path)
        for sheet in xl.sheet_names[:3]:
            df = pd.read_excel(path, sheet_name=sheet, header=None, nrows=30, dtype=str)
            preview = " ".join(df.fillna("").astype(str).values.ravel().tolist())
            if _looks_like_invoice_text(preview):
                return True
    except Exception:
        return False
    return False


def _is_non_invoice_by_name(path: Path) -> bool:
    name = path.name.lower()
    if name.startswith("~$"):
        return True
    if name in KNOWN_SUPPORT_FILENAMES:
        return True
    blocked_tokens = [
        "summary",
        "proof",
        "reconciliation",
        "manifest",
        "operation",
        "patch_notes",
        "approvedlanemap",
        "ledger",
        "site_alias",
        "domestic_with_distances",
        "runtime_upload",
        "result_review",
        "ci_runtime",
        "run_log",
    ]
    return any(token in name for token in blocked_tokens)


def _is_invoice_candidate(path: Path, runtime_root: Path) -> bool:
    if not path.is_file() or path.suffix.lower() not in INVOICE_EXTENSIONS:
        return False
    if _is_non_invoice_by_name(path) or _is_runtime_or_output_path(path, runtime_root):
        return False
    suffix = path.suffix.lower()
    if suffix in SPREADSHEET_EXTENSIONS:
        return _looks_like_invoice_spreadsheet(path)
    return _looks_like_invoice_text(_read_text_sample(path))


def iter_invoice_candidates(search_roots: Iterable[Path], runtime_root: Path) -> Iterable[Path]:
    seen: set[Path] = set()
    for root in search_roots:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            try:
                resolved = path.resolve()
            except Exception:
                continue
            if resolved in seen:
                continue
            seen.add(resolved)
            if _is_invoice_candidate(path, runtime_root):
                yield path


def score_invoice(path: Path) -> tuple[int, float, str]:
    name = path.name.lower()
    score = 0
    for token, weight in [
        ("invoice", 25),
        ("domestic", 20),
        ("scnt", 10),
        ("hvdc", 8),
        ("draft", 6),
        ("pasted", 4),
    ]:
        if token in name:
            score += weight
    if path.suffix.lower() in {".xlsx", ".xls"}:
        score += 15
    elif path.suffix.lower() == ".csv":
        score += 10
    elif path.suffix.lower() in TEXT_TABLE_EXTENSIONS:
        score += 5
    try:
        if path.suffix.lower() in SPREADSHEET_EXTENSIONS:
            score += 20 if _looks_like_invoice_spreadsheet(path) else 0
        else:
            text = _read_text_sample(path)
            score += _concept_score(text) * 10 + _keyword_score(text)
    except Exception:
        pass
    return (score, path.stat().st_mtime, path.name)


def _split_markdown_row(line: str) -> list[str]:
    stripped = line.strip()
    if stripped.startswith("|"):
        stripped = stripped[1:]
    if stripped.endswith("|"):
        stripped = stripped[:-1]
    return [cell.strip() for cell in stripped.split("|")]


def _is_markdown_separator(line: str) -> bool:
    cells = _split_markdown_row(line)
    return bool(cells) and all(re.fullmatch(r":?-{3,}:?", c.strip()) or not c.strip() for c in cells)


def _dataframe_from_markdown_table(text: str):
    import pandas as pd

    lines = [line.rstrip() for line in text.splitlines()]
    best_rows = None
    best_score = -1

    i = 0
    while i < len(lines):
        if "|" not in lines[i]:
            i += 1
            continue
        block = []
        while i < len(lines) and "|" in lines[i]:
            if lines[i].strip():
                block.append(lines[i])
            i += 1
        if len(block) < 2:
            continue
        header = _split_markdown_row(block[0])
        data_lines = block[1:]
        if data_lines and _is_markdown_separator(data_lines[0]):
            data_lines = data_lines[1:]
        rows = [_split_markdown_row(line) for line in data_lines if not _is_markdown_separator(line)]
        if not rows:
            continue
        score = _concept_score(" ".join(header)) * 10 + _keyword_score(" ".join(header))
        if score > best_score:
            best_score = score
            best_rows = (header, rows)

    if not best_rows:
        return None
    header, rows = best_rows
    max_len = max(len(header), *(len(r) for r in rows))
    header = header + [f"extra_{idx}" for idx in range(len(header), max_len)]
    rows = [r + [""] * (max_len - len(r)) for r in rows]
    df = pd.DataFrame(rows, columns=header[:max_len])
    return df if _looks_like_invoice_text(" ".join(df.columns.astype(str))) else None


def _dataframe_from_delimited_text(text: str):
    import pandas as pd

    lines = [line for line in text.splitlines() if line.strip()]
    best = None
    best_score = -1
    for sep in ["\t", ",", ";"]:
        for idx, line in enumerate(lines[:80]):
            if sep not in line:
                continue
            header_score = _concept_score(line) * 10 + _keyword_score(line)
            if header_score < 35:
                continue
            block = [line]
            for follow in lines[idx + 1 :]:
                if sep not in follow:
                    break
                block.append(follow)
            if len(block) < 2:
                continue
            try:
                df = pd.read_csv(io.StringIO("\n".join(block)), sep=sep, dtype=str)
            except Exception:
                continue
            if len(df) == 0:
                continue
            if header_score > best_score:
                best_score = header_score
                best = df
    return best


def _convert_text_invoice_to_csv(path: Path, outdir: Path) -> Path:
    text = _read_text_sample(path, max_chars=2_000_000)
    df = _dataframe_from_markdown_table(text)
    if df is None:
        df = _dataframe_from_delimited_text(text)
    if df is None or df.empty:
        raise ValueError(f"Could not extract an invoice table from text file: {path}")
    out = outdir / f"{path.stem}_invoice_table.csv"
    df.to_csv(out, index=False)
    return out


def _normalize_csv_or_tsv(path: Path, outdir: Path) -> Path:
    if path.suffix.lower() != ".tsv":
        return path
    import pandas as pd

    df = pd.read_csv(path, sep="\t", dtype=str)
    out = outdir / f"{path.stem}.csv"
    df.to_csv(out, index=False)
    return out


def _filename_from_url(url: str, default_name: str = "downloaded_invoice") -> str:
    parsed = urllib.parse.urlparse(url)
    name = Path(urllib.parse.unquote(parsed.path)).name
    if not name or "." not in name:
        name = default_name
    return re.sub(r"[^A-Za-z0-9._ -]+", "_", name)


def _download_url_to_path(url: str, outdir: Path) -> Path:
    outdir.mkdir(parents=True, exist_ok=True)
    target = outdir / _filename_from_url(url)
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 ChatGPT-HVDC-Audit"})
    with urllib.request.urlopen(request, timeout=60) as response:
        content_disposition = response.headers.get("Content-Disposition", "")
        match = re.search(r'filename="?([^";]+)"?', content_disposition)
        if match:
            target = outdir / re.sub(r"[^A-Za-z0-9._ -]+", "_", match.group(1))
        with open(target, "wb") as f:
            shutil.copyfileobj(response, f)
    return target


def _extract_zip_input(zip_path: Path, outdir: Path) -> Path:
    extract_dir = outdir / f"{zip_path.stem}_extracted"
    extract_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path) as zf:
        zf.extractall(extract_dir)
    return extract_dir


def _materialize_invoice_path(path: Path, runtime_root: Path, scratch_dir: Path) -> Path:
    suffix = path.suffix.lower()
    if suffix == ".zip":
        extracted = _extract_zip_input(path, scratch_dir)
        candidates = list(iter_invoice_candidates([extracted], runtime_root))
        if not candidates:
            raise FileNotFoundError(f"No invoice file found inside downloaded zip: {path}")
        candidates.sort(key=score_invoice, reverse=True)
        return _materialize_invoice_path(candidates[0], runtime_root, scratch_dir)
    if suffix in TEXT_TABLE_EXTENSIONS:
        return _convert_text_invoice_to_csv(path, scratch_dir)
    if suffix == ".tsv":
        return _normalize_csv_or_tsv(path, scratch_dir)
    return path.resolve()


def find_invoice(
    runtime_root: Path,
    explicit: str | None,
    search_dir: str | None,
    scratch_dir: Path,
    input_url: str | None = None,
) -> Path:
    if explicit:
        invoice = Path(explicit).expanduser().resolve()
        if not invoice.exists():
            raise FileNotFoundError(f"Invoice file not found: {invoice}")
        return _materialize_invoice_path(invoice, runtime_root, scratch_dir)

    if input_url:
        downloaded = _download_url_to_path(input_url, scratch_dir)
        return _materialize_invoice_path(downloaded, runtime_root, scratch_dir)

    search_groups: list[list[Path]] = []
    if search_dir:
        search_groups.append([Path(search_dir).expanduser().resolve()])

    runtime_parent = runtime_root.resolve().parent
    search_groups.append([runtime_parent])
    if Path("/mnt/data").exists():
        search_groups.append([Path("/mnt/data")])

    cwd = Path.cwd().resolve()
    if cwd == runtime_root.resolve() or str(cwd).startswith(str(runtime_parent)):
        search_groups.append([cwd])

    all_candidates: list[Path] = []
    for roots in search_groups:
        unique_roots = []
        seen = set()
        for root in roots:
            try:
                resolved = root.resolve()
            except Exception:
                resolved = root
            if resolved not in seen:
                unique_roots.append(resolved)
                seen.add(resolved)
        candidates = list(iter_invoice_candidates(unique_roots, runtime_root))
        all_candidates.extend(candidates)
        if candidates:
            candidates.sort(key=score_invoice, reverse=True)
            return _materialize_invoice_path(candidates[0], runtime_root, scratch_dir)

    if not all_candidates:
        raise FileNotFoundError(
            "No invoice file found. Upload an .xlsx/.xls/.csv invoice, paste an invoice table as .md/.txt, "
            "or pass --input-url when network access is available."
        )
    all_candidates.sort(key=score_invoice, reverse=True)
    return _materialize_invoice_path(all_candidates[0], runtime_root, scratch_dir)


def _sn_prefix(path: Path) -> str:
    """Return the first invoice S/N parsed from a supporting-doc filename.

    Prefer the post-underscore row marker used by converted POD files, e.g.
    ``02. Domestic ..._12 & 13. HVDC-..._POD.md`` -> ``12``.
    This prevents the section prefix ``02. Domestic`` from being mistaken for
    invoice row 2.
    """
    if _md_extract_row_numbers_from_filename:
        try:
            nums = _md_extract_row_numbers_from_filename(path.name)
            if nums:
                return nums[0]
        except Exception:
            pass
    match = re.match(r"^\s*(\d{1,4})\s*[.)_-]\s*(?!Domestic\b)", path.name, flags=re.IGNORECASE)
    return match.group(1).lstrip("0") if match else ""



def _is_supporting_doc_candidate(path: Path, runtime_root: Path, invoice: Path | None = None) -> bool:
    if not path.is_file() or path.suffix.lower() not in SUPPORTING_DOC_EXTENSIONS:
        return False
    if invoice is not None:
        try:
            if path.resolve() == invoice.resolve():
                return False
        except Exception:
            pass
    if _is_runtime_or_output_path(path, runtime_root):
        return False
    name = path.name.lower()
    if name in KNOWN_SUPPORT_FILENAMES:
        return False
    if any(token in name for token in ["patch_notes", "manifest", "operation", "result_review", "summary"]):
        return False
    if path.suffix.lower() in {".md", ".markdown"}:
        text = _read_text_sample(path, max_chars=12000).lower()
        signals = ["delivery", "note/waybill", "trip no", "consignment", "pod", "rate approval", "confirmed"]
        return ("hvdc-" in name or "hvdc-" in text) and any(sig in name or sig in text for sig in signals)
    return "hvdc-" in name or "pod" in name or "dn" in name or "delivery" in name or "waybill" in name


def stage_supporting_documents(
    runtime_root: Path,
    outdir: Path,
    invoice: Path,
    explicit_docs: str | None,
    search_dir: str | None = None,
) -> Path | None:
    if explicit_docs:
        docs_path = Path(explicit_docs).expanduser().resolve()
        return docs_path if docs_path.exists() else None

    roots: list[Path] = []
    if search_dir:
        roots.append(Path(search_dir).expanduser().resolve())
    if Path("/mnt/data").exists():
        roots.append(Path("/mnt/data"))
    roots.append(runtime_root.resolve().parent)

    staged = outdir / "supporting_docs_staging"
    staged.mkdir(parents=True, exist_ok=True)

    copied = 0
    seen: set[Path] = set()
    for root in roots:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            try:
                resolved = path.resolve()
            except Exception:
                continue
            if resolved in seen:
                continue
            if _safe_relative_to(resolved, outdir):
                continue
            seen.add(resolved)
            if not _is_supporting_doc_candidate(path, runtime_root, invoice):
                continue
            dest = staged / path.name
            if dest.exists():
                stem, suffix = dest.stem, dest.suffix
                dest = staged / f"{stem}_{copied + 1}{suffix}"
            try:
                shutil.copy2(path, dest)
                copied += 1
            except Exception:
                pass

    if copied == 0:
        try:
            staged.rmdir()
        except Exception:
            pass
        return None
    return staged


def read_output_summary(items_csv: Path) -> dict:
    summary: dict = {"items_csv_exists": items_csv.exists()}
    if not items_csv.exists():
        return summary
    try:
        import pandas as pd

        df = pd.read_csv(items_csv)
        summary["row_count"] = int(len(df))
        if "qty" in df.columns:
            summary["qty_sum"] = float(pd.to_numeric(df["qty"], errors="coerce").fillna(0).sum())
        for col in ["verdict", "cg_band"]:
            if col in df.columns:
                summary[f"{col}_counts"] = {
                    str(k): int(v) for k, v in df[col].fillna("").value_counts().to_dict().items()
                }
        for col in ["shipment_ref", "origin_norm", "destination_norm", "ref_lane_id"]:
            if col in df.columns:
                summary[f"{col}_nonblank"] = int(df[col].fillna("").astype(str).str.strip().ne("").sum())
        if "evidence_count" in df.columns:
            summary["evidence_count_sum"] = int(pd.to_numeric(df["evidence_count"], errors="coerce").fillna(0).sum())
            summary["rows_with_evidence"] = int(pd.to_numeric(df["evidence_count"], errors="coerce").fillna(0).gt(0).sum())
    except Exception as exc:  # pragma: no cover - defensive summary
        summary["items_csv_read_error"] = str(exc)
    return summary


def find_outputs(outdir: Path) -> dict:
    xlsx_files = sorted(
        outdir.glob("domestic_audit_report_v2*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    pdf_files = sorted(
        outdir.glob("domestic_audit_report_v2*.pdf"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return {
        "items_csv": str(outdir / "items.csv") if (outdir / "items.csv").exists() else None,
        "audit_workbook": str(xlsx_files[0]) if xlsx_files else None,
        "audit_pdf": str(pdf_files[0]) if pdf_files else None,
        "proof_json": str(outdir / "domestic_audit_proof_v2.json")
        if (outdir / "domestic_audit_proof_v2.json").exists()
        else None,
        "self_check_json": str(outdir / "self_verification_report.json")
        if (outdir / "self_verification_report.json").exists()
        else None,
        "pod_reconciliation_csv": str(outdir / "md_pod_reconciliation.csv")
        if (outdir / "md_pod_reconciliation.csv").exists()
        else None,
        "pod_reconciliation_md": str(outdir / "md_pod_reconciliation.md")
        if (outdir / "md_pod_reconciliation.md").exists()
        else None,
        "md_as_pdf_evidence_inventory_csv": str(outdir / "md_as_pdf_evidence_inventory.csv")
        if (outdir / "md_as_pdf_evidence_inventory.csv").exists()
        else None,
        "md_as_pdf_evidence_inventory_md": str(outdir / "md_as_pdf_evidence_inventory.md")
        if (outdir / "md_as_pdf_evidence_inventory.md").exists()
        else None,
    }


def classify_status(exit_code: int, log_text: str, outputs: dict, self_check: dict | None = None) -> str:
    """Classify the audit result using audit-relevant logs and required outputs."""
    fatal_patterns = [
        r"Traceback \(most recent call last\)",
        r"\[ERROR\]",
        r"Invoice file not found",
        r"Config file not found",
        r"RuntimeError",
        r"FileNotFoundError",
    ]
    fatal = any(re.search(pattern, log_text, flags=re.IGNORECASE) for pattern in fatal_patterns)
    if exit_code != 0 or fatal:
        return "failed"
    if self_check and not self_check.get("hard_pass", False):
        return "failed"
    if outputs.get("audit_workbook") and outputs.get("items_csv"):
        return "succeeded"
    return "inconclusive"


def _check(checks: list[dict], name: str, passed: bool, severity: str, detail: str, **extra: Any) -> None:
    row = {
        "check": name,
        "status": "PASS" if passed else "FAIL",
        "severity": severity,
        "detail": detail,
    }
    row.update(extra)
    checks.append(row)


def _load_invoice_row_count(invoice: Path) -> int | None:
    try:
        import pandas as pd

        if invoice.suffix.lower() == ".csv":
            return int(len(pd.read_csv(invoice)))
        if invoice.suffix.lower() == ".tsv":
            return int(len(pd.read_csv(invoice, sep="\t")))
        if invoice.suffix.lower() in {".xlsx", ".xls"}:
            # Reuse the audit's header-detection helper when available.
            from patch.run_domestic_audit_v2 import find_header_and_load

            return int(len(find_header_and_load(invoice)))
    except Exception:
        return None
    return None


def _extract_trip_numbers(text: str) -> list[str]:
    trips: list[str] = []
    lines = text.splitlines()
    for idx, line in enumerate(lines):
        if re.search(r"Trip\s*No\.?\s*:", line, re.IGNORECASE):
            after = re.sub(r".*Trip\s*No\.?\s*:\s*", "", line, flags=re.IGNORECASE).strip()
            candidates = [after] if after else []
            candidates.extend(lines[idx + 1 : idx + 4])
            for candidate in candidates:
                cleaned = re.sub(r"[^A-Za-z0-9-]", "", candidate.strip())
                if re.search(r"\d", cleaned) and len(cleaned) >= 6:
                    trips.append(cleaned)
                    break
    # Preserve order while de-duplicating.
    return list(dict.fromkeys(trips))


def _extract_waybill_numbers(text: str) -> list[str]:
    nums = re.findall(
        r"(?:Note/Waybill#|Waybill#|Note#)\s*:?\s*([0-9]{4}-[0-9A-Z-]+AUH)",
        text,
        flags=re.IGNORECASE,
    )
    return list(dict.fromkeys(n.strip() for n in nums if n.strip()))


def _extract_rate_approval(text: str) -> str:
    patterns = [
        r"AED\s*[\d,]+(?:\.\d+)?\s*(?:per|/)\s*(?:Trailer|Truck|Trip)",
        r"USD\s*[\d,]+(?:\.\d+)?\s*(?:per|/)\s*(?:Trailer|Truck|Trip)",
        r"Confirmed\.",
        r"kindly approve below charges",
        r"approval",
    ]
    hits = []
    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            snippet = match.group(0).strip()
            if snippet not in hits:
                hits.append(snippet)
    return "; ".join(hits[:5])


def _build_md_pod_reconciliation(items_df, docs_dir: Path | None):
    import pandas as pd

    base_rows = []
    if items_df is None or items_df.empty:
        return pd.DataFrame()

    def _norm_ref(value: Any) -> str:
        return str(value or "").strip().upper().replace("/", "-")

    def _dedupe_infos(infos: list[dict]) -> list[dict]:
        seen = set()
        out = []
        for info in infos:
            key = str(info.get("file_path") or info.get("file_name") or "")
            if key in seen:
                continue
            seen.add(key)
            out.append(info)
        return out

    md_by_sn: dict[str, list[dict]] = {}
    md_by_ref: dict[str, list[dict]] = {}
    if docs_dir and docs_dir.exists():
        for path in list(docs_dir.rglob("*.md")) + list(docs_dir.rglob("*.markdown")):
            text = (_md_read_text(path, max_chars=2_000_000) if _md_read_text else _read_text_sample(path, max_chars=2_000_000))
            row_numbers = []
            if _md_extract_row_numbers_from_filename:
                try:
                    row_numbers = _md_extract_row_numbers_from_filename(path.name)
                except Exception:
                    row_numbers = []
            if not row_numbers:
                sn = _sn_prefix(path)
                row_numbers = [sn] if sn else []

            refs = (_md_extract_hvdc_refs(text + "\n" + path.name) if _md_extract_hvdc_refs else None)
            if refs is None:
                refs = re.findall(r"HVDC-[A-Z0-9]+(?:[-/][A-Z0-9]+)+-\d+", text + "\n" + path.name, flags=re.IGNORECASE)
                refs = [r.upper().replace("/", "-") for r in refs]
            refs = list(dict.fromkeys(refs))

            waybills = (_md_extract_waybill_numbers(text) if _md_extract_waybill_numbers else _extract_waybill_numbers(text))
            trips = (_md_extract_trip_numbers(text) if _md_extract_trip_numbers else _extract_trip_numbers(text))
            approval = (_md_extract_rate_approval(text) if _md_extract_rate_approval else _extract_rate_approval(text))

            info = {
                "file_name": path.name,
                "file_path": str(path),
                "sn_prefix": ",".join(row_numbers),
                "row_numbers": row_numbers,
                "shipment_refs": ",".join(refs),
                "shipment_refs_list": refs,
                "waybill_count": len(waybills),
                "trip_no_count": len(trips),
                "rate_approval": approval,
                "source_format": "MD_AS_PDF_TEXT",
                "md_as_pdf_equivalent": True,
            }
            for sn in row_numbers:
                if sn:
                    md_by_sn.setdefault(str(int(sn)) if str(sn).isdigit() else str(sn), []).append(info)
            for ref in refs:
                if ref:
                    md_by_ref.setdefault(_norm_ref(ref), []).append(info)

    for _, row in items_df.iterrows():
        sn_raw = str(row.get("S/N", row.get("sn", ""))).strip()
        sn = re.sub(r"\.0$", "", sn_raw).lstrip("0") or sn_raw
        shipment_ref = _norm_ref(row.get("shipment_ref", ""))
        qty = row.get("qty", "")
        try:
            qty_num = float(qty)
        except Exception:
            qty_num = None

        docs = []
        match_basis = []
        # Exact shipment reference has highest evidentiary quality.
        if shipment_ref and shipment_ref in md_by_ref:
            docs.extend(md_by_ref[shipment_ref])
            match_basis.append("shipment_ref_exact")
        # Partial shipment reference matching catches small punctuation variants.
        if shipment_ref:
            for ref_key, ref_docs in md_by_ref.items():
                if ref_key == shipment_ref:
                    continue
                if shipment_ref in ref_key or ref_key in shipment_ref:
                    docs.extend(ref_docs)
                    match_basis.append("shipment_ref_partial")
        # Row-number fallback is valid for converted POD bundles named with S/N.
        if sn and sn in md_by_sn:
            docs.extend(md_by_sn[sn])
            match_basis.append("row_number")

        docs = _dedupe_infos(docs)
        waybill_count = sum(int(d.get("waybill_count") or 0) for d in docs)
        trip_no_count = sum(int(d.get("trip_no_count") or 0) for d in docs)

        exact_match_count = 0
        if shipment_ref:
            for d in docs:
                refs = [_norm_ref(x) for x in d.get("shipment_refs_list", [])]
                if shipment_ref in refs:
                    exact_match_count += 1

        # Waybill count is the primary POD quantity signal. Some DSV documents
        # reuse one Trip No. across several waybills/trailers, so unique Trip No.
        # count is recorded but must not override a matching waybill count.
        primary_evidence_count = waybill_count if waybill_count > 0 else trip_no_count
        primary_evidence_type = "waybill" if waybill_count > 0 else ("trip_no" if trip_no_count > 0 else "")
        if docs and qty_num is not None:
            qty_match = "PASS" if abs(primary_evidence_count - qty_num) < 0.001 else "FAIL"
        elif docs:
            qty_match = "CHECK"
        else:
            qty_match = "NO_MD"

        base_rows.append(
            {
                "S/N": sn_raw,
                "shipment_ref": row.get("shipment_ref", ""),
                "qty_invoice": qty,
                "md_file_count": len(docs),
                "md_waybill_count": waybill_count,
                "md_trip_no_count": trip_no_count,
                "md_primary_evidence_type": primary_evidence_type,
                "md_primary_evidence_count": primary_evidence_count,
                "md_qty_match": qty_match,
                "md_match_basis": ",".join(sorted(set(match_basis))),
                "exact_shipment_ref_match_count": exact_match_count,
                "md_as_pdf_equivalent": bool(docs),
                "md_files": "; ".join(d.get("file_name", "") for d in docs),
                "rate_approval_evidence": "; ".join(d.get("rate_approval", "") for d in docs if d.get("rate_approval")),
            }
        )
    return pd.DataFrame(base_rows)


def _write_markdown_table(df, path: Path) -> None:
    if df is None or df.empty:
        path.write_text("(no rows)\n", encoding="utf-8")
        return
    cols = list(df.columns)
    lines = ["| " + " | ".join(cols) + " |", "| " + " | ".join(["---"] * len(cols)) + " |"]
    for _, row in df.iterrows():
        values = [str(row.get(c, "")).replace("\n", " ").replace("|", "\\|") for c in cols]
        lines.append("| " + " | ".join(values) + " |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _append_self_check_to_workbook(workbook_path: Path, checks: list[dict], pod_df, self_check: dict, md_inventory_df=None) -> bool:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = load_workbook(workbook_path)

        def _safe_cell_value(value):
            if value is None:
                return ""
            if isinstance(value, (int, float, bool)):
                return value
            text = str(value)
            # Excel XML cannot store ASCII control characters except tab/newline/CR.
            text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", " ", text)
            return text

        for sheet_name in ["self_check", "pod_reconciliation", "md_as_pdf_evidence", "_format_profile"]:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

        ws = wb.create_sheet("self_check", 0)
        rows = [
            ["HVDC Domestic Audit Self Verification"],
            ["Generated At", self_check.get("generated_at", "")],
            ["Hard Pass", str(self_check.get("hard_pass", ""))],
            ["Warning Count", self_check.get("warning_count", 0)],
            ["MD as PDF Evidence", "MD files are treated as PDF-equivalent text evidence"],
            [],
            ["Check", "Status", "Severity", "Detail"],
        ]
        for row in rows:
            ws.append([_safe_cell_value(v) for v in row])
        for item in checks:
            ws.append([_safe_cell_value(item.get("check", "")), _safe_cell_value(item.get("status", "")), _safe_cell_value(item.get("severity", "")), _safe_cell_value(item.get("detail", ""))])

        def _append_dataframe_sheet(sheet_name: str, df) -> bool:
            if df is None or getattr(df, "empty", True):
                return False
            ws_df = wb.create_sheet(sheet_name)
            ws_df.append([_safe_cell_value(c) for c in list(df.columns)])
            for _, row in df.iterrows():
                ws_df.append([_safe_cell_value(row.get(c, "")) for c in df.columns])
            return True

        pod_added = _append_dataframe_sheet("pod_reconciliation", pod_df)
        inventory_added = _append_dataframe_sheet("md_as_pdf_evidence", md_inventory_df)

        ws_profile = wb.create_sheet("_format_profile")
        profile_rows = [
            ["setting", "value"],
            ["font_family", "Calibri"],
            ["font_size", 10],
            ["horizontal_alignment", "center"],
            ["vertical_alignment", "middle"],
            ["wrap_text", False],
            ["number_format_decimal", "0.00"],
            ["number_format_thousands_decimal", "#,##0.00"],
            ["date_format", "yyyy-mm-dd"],
            ["row_height", "autofit/default 18"],
            ["column_width", "autofit bounded 10-40"],
            ["md_files_are_pdf_equivalent", True],
        ]
        for row in profile_rows:
            ws_profile.append([_safe_cell_value(v) for v in row])
        ws_profile.sheet_state = "hidden"

        # Apply a stable workbook-wide format after self-check sheets are added.
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        body_font = Font(name="Calibri", size=10, color="000000")
        link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        title_font = Font(name="Calibri", size=12, bold=True, color="1F4E78")
        center = Alignment(horizontal="center", vertical="center", wrap_text=False)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        money_like = {
            "Rate (AED)", "rate_usd", "Amount (US$)", "Total (US$)",
            "rate_usd_input", "rate_aed_input", "ref_rate_usd",
            "executed_ref_rate_usd", "Rate_USD"
        }
        pct_like = {"executed_ref_delta_pct", "delta_pct"}
        date_like = {"pdf_issue_date", "issue_date", "Generated At", "printed_date"}

        for sheet in wb.worksheets:
            sheet.sheet_view.showGridLines = True
            sheet.freeze_panes = "A2"
            if sheet.max_row >= 1 and sheet.max_column >= 1:
                sheet.auto_filter.ref = sheet.dimensions

            # Default row height approximates Excel autofit without making rows tall.
            for row_cells in sheet.iter_rows():
                sheet.row_dimensions[row_cells[0].row].height = 18
                for cell in row_cells:
                    cell.font = link_font if cell.hyperlink else body_font
                    cell.alignment = center
                    cell.border = border

            if sheet.max_row >= 1:
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center
                    cell.border = border
                # Self-check title row is a title, not a table header.
                if sheet.title == "self_check" and sheet["A1"].value == "HVDC Domestic Audit Self Verification":
                    sheet["A1"].font = title_font
                    sheet["A1"].fill = PatternFill(fill_type=None)
                    for cell in sheet[7]:
                        cell.font = header_font
                        cell.fill = header_fill

            header_values = [str(sheet.cell(1, c).value or "") for c in range(1, sheet.max_column + 1)]
            for c_idx, header in enumerate(header_values, start=1):
                col_letter = get_column_letter(c_idx)
                max_len = len(header)
                for r_idx in range(1, min(sheet.max_row, 200) + 1):
                    value = sheet.cell(r_idx, c_idx).value
                    if value is None:
                        continue
                    max_len = max(max_len, len(str(value)) if len(str(value)) < 80 else 80)
                width = max(10, min(max_len + 2, 40))
                sheet.column_dimensions[col_letter].width = width

                if header in money_like:
                    for r_idx in range(2, sheet.max_row + 1):
                        sheet.cell(r_idx, c_idx).number_format = "#,##0.00"
                elif header in pct_like:
                    for r_idx in range(2, sheet.max_row + 1):
                        sheet.cell(r_idx, c_idx).number_format = "0.00%"
                elif header in date_like or "date" in header.lower():
                    for r_idx in range(2, sheet.max_row + 1):
                        sheet.cell(r_idx, c_idx).number_format = "yyyy-mm-dd"

        self_check["pod_reconciliation_sheet_added"] = bool(pod_added)
        self_check["md_as_pdf_evidence_sheet_added"] = bool(inventory_added)
        wb.save(workbook_path)
        try:
            from final_excel_link_evidence_repair import finalize_ref_lane_id_clickable_links_ooxml
            finalize_ref_lane_id_clickable_links_ooxml(workbook_path, update_self_check=True)
        except Exception as final_link_exc:
            self_check.setdefault("append_errors", []).append(f"final ref_lane_id clickable link repair failed: {final_link_exc}")
            return False
        return True
    except Exception as exc:
        self_check.setdefault("append_errors", []).append(str(exc))
        return False



def _load_fixed_items_contract() -> list[str]:
    """Load fixed items column contract bundled with runtime."""
    try:
        fmt_path = Path(__file__).resolve().parent / "Data" / "DOMESTIC_SHEET_FORMAT.json"
        payload = json.loads(fmt_path.read_text(encoding="utf-8"))
        for sheet in payload.get("workbook_structure", []):
            if sheet.get("sheet_name") == "items":
                return [c.get("field_name", "") for c in sheet.get("columns", []) if c.get("field_name")]
    except Exception:
        pass
    return []

def _column_indexes_from_header(header: list, names: list[str]) -> dict[str, int]:
    return {name: (header.index(name) + 1) for name in names if name in header}


def run_self_verification(invoice: Path, outdir: Path, outputs: dict, docs_dir: Path | None) -> dict:
    import pandas as pd

    checks: list[dict] = []
    generated_at = datetime.now().isoformat(timespec="seconds")

    items_csv = Path(outputs["items_csv"]) if outputs.get("items_csv") else None
    workbook = Path(outputs["audit_workbook"]) if outputs.get("audit_workbook") else None
    proof_json = Path(outputs["proof_json"]) if outputs.get("proof_json") else None

    _check(checks, "output.items_csv.exists", bool(items_csv and items_csv.exists() and items_csv.stat().st_size > 0), "HARD", str(items_csv or ""))
    _check(checks, "output.audit_workbook.exists", bool(workbook and workbook.exists() and workbook.stat().st_size > 0), "HARD", str(workbook or ""))
    _check(checks, "output.proof_json.exists", bool(proof_json and proof_json.exists() and proof_json.stat().st_size > 0), "HARD", str(proof_json or ""))

    items_df = None
    workbook_items_df = None
    invoice_row_count = _load_invoice_row_count(invoice)
    items_row_count = None
    workbook_row_count = None

    if items_csv and items_csv.exists():
        try:
            items_df = pd.read_csv(items_csv)
            items_row_count = int(len(items_df))
            _check(checks, "items_csv.readable", True, "HARD", f"rows={items_row_count}")
        except Exception as exc:
            _check(checks, "items_csv.readable", False, "HARD", str(exc))

    # Final combined repair is intentionally executed before workbook checks:
    # - convert price/ref_rate HYPERLINK formulas to actual internal hyperlinks,
    # - ensure ref_lane_id points to lane_map_reference,
    # - copy md_as_pdf_evidence semantic fields into the items sheet/CSV.
    early_md_inventory_df = pd.DataFrame()
    if docs_dir and build_md_as_pdf_inventory:
        try:
            early_md_inventory_df = pd.DataFrame(build_md_as_pdf_inventory(docs_dir))
        except Exception as exc:
            _check(checks, "workbook.final_patch.md_inventory_preload", False, "WARN", str(exc))
    if workbook and workbook.exists():
        try:
            from final_excel_link_evidence_repair import repair_workbook_and_items_csv
            repair_report = repair_workbook_and_items_csv(workbook, items_csv, early_md_inventory_df)
            for repair_check in repair_report.get("checks", []):
                _check(
                    checks,
                    str(repair_check.get("check", "")),
                    bool(repair_check.get("passed")),
                    str(repair_check.get("severity", "HARD")),
                    str(repair_check.get("detail", "")),
                )
            if items_csv and items_csv.exists():
                items_df = pd.read_csv(items_csv)
                items_row_count = int(len(items_df))
        except Exception as exc:
            _check(checks, "workbook.final_patch.combined_repair", False, "HARD", str(exc))

    if workbook and workbook.exists():
        try:
            import openpyxl

            fixed_items_columns = _load_fixed_items_contract()
            wb = openpyxl.load_workbook(workbook, read_only=False, data_only=False)
            required_sheets = {"items", "summary_band", "summary_verdict", "lane_map_reference"}
            missing = sorted(required_sheets - set(wb.sheetnames))
            _check(checks, "workbook.required_sheets", not missing, "HARD", f"missing={missing}; sheets={wb.sheetnames}")

            if "items" in wb.sheetnames and fixed_items_columns:
                ws_items_contract = wb["items"]
                observed_header = [
                    ws_items_contract.cell(1, c).value
                    for c in range(1, ws_items_contract.max_column + 1)
                ]
                header_ok = observed_header == fixed_items_columns
                _check(
                    checks,
                    "workbook.items.fixed_column_order",
                    header_ok,
                    "HARD",
                    f"expected={len(fixed_items_columns)} cols; observed={len(observed_header)} cols",
                )

                md_items_columns = [
                    "pdf_dn_number",
                    "pdf_issue_date",
                    "pdf_origin",
                    "pdf_destination",
                    "pdf_content_summary",
                    "pdf_extracted_fields",
                ]
                missing_md_items_columns = [c for c in md_items_columns if c not in observed_header]
                _check(
                    checks,
                    "workbook.items.md_as_pdf_evidence_columns",
                    not missing_md_items_columns,
                    "HARD",
                    f"missing={missing_md_items_columns}; required={md_items_columns}",
                )

                if "md_as_pdf_evidence" in wb.sheetnames and not missing_md_items_columns:
                    try:
                        ws_md = wb["md_as_pdf_evidence"]
                        md_header = [ws_md.cell(1, c).value for c in range(1, ws_md.max_column + 1)]
                        item_indexes = _column_indexes_from_header(observed_header, ["S/N"] + md_items_columns)
                        md_header_map = {str(v).strip().lower(): i + 1 for i, v in enumerate(md_header) if v is not None}

                        def _md_col(*names):
                            for name in names:
                                idx = md_header_map.get(str(name).strip().lower())
                                if idx:
                                    return idx
                            return None

                        def _cell_text(ws, row_idx, col_idx):
                            if not col_idx:
                                return ""
                            value = ws.cell(row_idx, col_idx).value
                            return "" if value is None else str(value).strip()

                        def _parse_md_row_numbers(text):
                            return [int(x) for x in re.findall(r"\d+", str(text or ""))]

                        sn_to_row = {}
                        sn_col = item_indexes.get("S/N")
                        if sn_col:
                            for item_row_idx in range(2, ws_items_contract.max_row + 1):
                                try:
                                    sn_text = str(ws_items_contract.cell(item_row_idx, sn_col).value or "").strip()
                                    if sn_text:
                                        sn_to_row[int(float(sn_text))] = item_row_idx
                                except Exception:
                                    continue

                        semantic_sources = {
                            "pdf_dn_number": _md_col("first_waybill_no", "waybill_no", "dn_number", "dn no", "dn_number"),
                            "pdf_issue_date": _md_col("issue_date", "printed_date", "pdf_issue_date", "date"),
                            "pdf_origin": _md_col("origin", "origin_from_pdf", "loading_address", "loading point"),
                            "pdf_destination": _md_col("destination", "destination_from_pdf", "offloading_address", "destination_hint"),
                            "pdf_content_summary": _md_col("content_summary", "summary", "extracted_content"),
                        }

                        expected_md_item_rows = 0
                        passed_md_item_rows = 0
                        for md_row_idx in range(2, ws_md.max_row + 1):
                            row_nums = _parse_md_row_numbers(_cell_text(ws_md, md_row_idx, _md_col("row_numbers", "row_no", "s/n", "serial")))
                            if not row_nums:
                                continue
                            for sn in row_nums:
                                item_row_idx = sn_to_row.get(sn)
                                if not item_row_idx:
                                    continue
                                expected_md_item_rows += 1
                                row_ok = True
                                for target_col, source_col in semantic_sources.items():
                                    if not source_col:
                                        continue
                                    source_text = _cell_text(ws_md, md_row_idx, source_col)
                                    if not source_text:
                                        continue
                                    target_text = _cell_text(ws_items_contract, item_row_idx, item_indexes.get(target_col))
                                    if not target_text:
                                        row_ok = False
                                        break
                                extracted_text = _cell_text(ws_items_contract, item_row_idx, item_indexes.get("pdf_extracted_fields"))
                                if not extracted_text:
                                    row_ok = False
                                if row_ok:
                                    passed_md_item_rows += 1
                        _check(
                            checks,
                            "workbook.items.md_as_pdf_evidence_mapping",
                            expected_md_item_rows > 0 and passed_md_item_rows == expected_md_item_rows,
                            "HARD",
                            f"mapped_rows={passed_md_item_rows}/{expected_md_item_rows}; source_sheet=md_as_pdf_evidence; semantic_header_match=True",
                        )
                    except Exception as exc:
                        _check(
                            checks,
                            "workbook.items.md_as_pdf_evidence_mapping",
                            False,
                            "HARD",
                            str(exc),
                        )

                price_columns = [
                    "Rate (AED)",
                    "rate_usd",
                    "Amount (US$)",
                    "Total (US$)",
                    "rate_usd_input",
                    "rate_aed_input",
                    "ref_rate_usd",
                    "executed_ref_rate_usd",
                ]
                indexes = _column_indexes_from_header(observed_header, ["ref_lane_id"] + price_columns)
                expected_rows = 0
                price_link_count = 0
                ref_rate_link_count = 0
                ref_lane_id_link_count = 0
                ref_lane_id_lane_map_target_count = 0
                if "ref_lane_id" in indexes:
                    ref_idx = indexes["ref_lane_id"]
                    for row_idx in range(2, ws_items_contract.max_row + 1):
                        lane_cell = ws_items_contract.cell(row_idx, ref_idx)
                        lane_value = lane_cell.value
                        if lane_value is None or str(lane_value).strip() == "":
                            continue
                        expected_rows += 1
                        lane_value_text = str(lane_value).strip()
                        lane_link_target = ""
                        if lane_cell.hyperlink:
                            lane_link_target = str(
                                getattr(lane_cell.hyperlink, "location", None)
                                or getattr(lane_cell.hyperlink, "target", None)
                                or ""
                            )
                        elif str(lane_value).upper().startswith("=HYPERLINK("):
                            lane_link_target = str(lane_value)
                        has_lane_link = bool(lane_cell.hyperlink) or str(lane_value).upper().startswith("=HYPERLINK(")
                        if has_lane_link:
                            ref_lane_id_link_count += 1
                            if "lane_map_reference" in lane_link_target.lower():
                                ref_lane_id_lane_map_target_count += 1
                        for col_name in price_columns:
                            col_idx = indexes.get(col_name)
                            if not col_idx:
                                continue
                            cell = ws_items_contract.cell(row_idx, col_idx)
                            value = cell.value
                            if value is None or str(value).strip() == "":
                                continue
                            has_link = bool(cell.hyperlink) or str(value).upper().startswith("=HYPERLINK(")
                            if has_link:
                                price_link_count += 1
                                if col_name == "ref_rate_usd":
                                    ref_rate_link_count += 1
                _check(
                    checks,
                    "workbook.items.price_links",
                    price_link_count >= expected_rows,
                    "HARD",
                    f"linked_price_cells={price_link_count}; rows_with_ref_lane_id={expected_rows}",
                )
                _check(
                    checks,
                    "workbook.items.ref_rate_usd_links",
                    ref_rate_link_count >= expected_rows,
                    "HARD",
                    f"linked_ref_rate_usd_cells={ref_rate_link_count}; rows_with_ref_lane_id={expected_rows}",
                )
                _check(
                    checks,
                    "workbook.items.ref_lane_id_links",
                    ref_lane_id_link_count >= expected_rows,
                    "HARD",
                    f"linked_ref_lane_id_cells={ref_lane_id_link_count}; rows_with_ref_lane_id={expected_rows}",
                )
                _check(
                    checks,
                    "workbook.items.ref_lane_id_lane_map_targets",
                    ref_lane_id_lane_map_target_count >= expected_rows,
                    "HARD",
                    f"lane_map_reference_targets={ref_lane_id_lane_map_target_count}; rows_with_ref_lane_id={expected_rows}",
                )

                # A second hard check uses data_only mode to catch formulas whose
                # cached values were stripped by later workbook saves.  This is
                # the failure mode that makes ref_rate_usd appear blank in
                # non-calculating previews even when HYPERLINK() formulas exist.
                ref_rate_display_count = 0
                if "ref_lane_id" in indexes and "ref_rate_usd" in indexes:
                    wb_values = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
                    try:
                        ws_values = wb_values["items"]
                        ref_idx = indexes["ref_lane_id"]
                        ref_rate_idx = indexes["ref_rate_usd"]
                        for row_idx in range(2, ws_values.max_row + 1):
                            lane_value = ws_values.cell(row_idx, ref_idx).value
                            if lane_value is None or str(lane_value).strip() == "":
                                continue
                            ref_rate_value = ws_values.cell(row_idx, ref_rate_idx).value
                            if ref_rate_value is not None and str(ref_rate_value).strip() != "":
                                ref_rate_display_count += 1
                    finally:
                        wb_values.close()
                _check(
                    checks,
                    "workbook.items.ref_rate_usd_display_values",
                    ref_rate_display_count >= expected_rows,
                    "HARD",
                    f"display_ref_rate_usd_cells={ref_rate_display_count}; rows_with_ref_lane_id={expected_rows}",
                )
            wb.close()
            workbook_items_df = pd.read_excel(workbook, sheet_name="items")
            workbook_row_count = int(len(workbook_items_df))
            _check(checks, "workbook.items.readable", True, "HARD", f"rows={workbook_row_count}")
        except Exception as exc:
            _check(checks, "workbook.items.readable", False, "HARD", str(exc))

    if invoice_row_count is not None and items_row_count is not None:
        _check(
            checks,
            "row_count.invoice_vs_items",
            invoice_row_count == items_row_count,
            "HARD",
            f"invoice={invoice_row_count}; items={items_row_count}",
        )
    if items_row_count is not None and workbook_row_count is not None:
        _check(
            checks,
            "row_count.items_csv_vs_workbook",
            items_row_count == workbook_row_count,
            "HARD",
            f"items_csv={items_row_count}; workbook_items={workbook_row_count}",
        )

    if items_df is not None:
        for col in ["shipment_ref", "origin_norm", "destination_norm", "ref_lane_id", "verdict", "cg_band"]:
            if col in items_df.columns:
                nonblank = int(items_df[col].fillna("").astype(str).str.strip().ne("").sum())
                _check(
                    checks,
                    f"column.{col}.nonblank",
                    nonblank == len(items_df),
                    "WARN",
                    f"nonblank={nonblank}/{len(items_df)}",
                )

        # Arithmetic consistency check: rate_usd * qty ~= Amount, Amount ~= Total.
        if {"qty", "rate_usd", "Amount (US$)"}.issubset(set(items_df.columns)):
            qty = pd.to_numeric(items_df["qty"], errors="coerce")
            rate = pd.to_numeric(items_df["rate_usd"], errors="coerce")
            amount = pd.to_numeric(items_df["Amount (US$)"], errors="coerce")
            mismatch = ((qty * rate - amount).abs() > 0.05).fillna(False)
            _check(
                checks,
                "arithmetic.qty_x_rate_vs_amount",
                int(mismatch.sum()) == 0,
                "WARN",
                f"mismatches={int(mismatch.sum())}",
            )
        if {"Amount (US$)", "Total (US$)"}.issubset(set(items_df.columns)):
            amount = pd.to_numeric(items_df["Amount (US$)"], errors="coerce")
            total = pd.to_numeric(items_df["Total (US$)"], errors="coerce")
            mismatch = ((amount - total).abs() > 0.05).fillna(False)
            _check(
                checks,
                "arithmetic.amount_vs_total",
                int(mismatch.sum()) == 0,
                "WARN",
                f"mismatches={int(mismatch.sum())}",
            )

    pod_df = _build_md_pod_reconciliation(items_df, docs_dir) if items_df is not None else pd.DataFrame()
    pod_csv = outdir / "md_pod_reconciliation.csv"
    pod_md = outdir / "md_pod_reconciliation.md"
    if not pod_df.empty:
        pod_df.to_csv(pod_csv, index=False)
        _write_markdown_table(pod_df, pod_md)
        if "md_qty_match" in pod_df.columns:
            compared = pod_df[pod_df["md_qty_match"].isin(["PASS", "FAIL"])]
            mismatches = int((compared["md_qty_match"] == "FAIL").sum())
            _check(
                checks,
                "md_pod.qty_reconciliation",
                mismatches == 0,
                "WARN",
                f"compared_rows={len(compared)}; mismatches={mismatches}",
            )
            no_md = int((pod_df["md_qty_match"] == "NO_MD").sum())
            _check(
                checks,
                "md_pod.coverage",
                no_md == 0,
                "WARN",
                f"rows_without_md={no_md}/{len(pod_df)}",
            )
            exact_matches = int(pd.to_numeric(pod_df.get("exact_shipment_ref_match_count", 0), errors="coerce").fillna(0).gt(0).sum()) if "exact_shipment_ref_match_count" in pod_df.columns else 0
            _check(
                checks,
                "md_pod.exact_shipment_ref_match",
                exact_matches == len(pod_df),
                "WARN",
                f"rows_with_exact_ref_match={exact_matches}/{len(pod_df)}; row_number_fallback_is_allowed",
            )
    else:
        _check(checks, "md_pod.reconciliation", True, "INFO", "No MD POD files staged or no items rows.")

    md_inventory_df = pd.DataFrame()
    md_inventory_csv = outdir / "md_as_pdf_evidence_inventory.csv"
    md_inventory_md = outdir / "md_as_pdf_evidence_inventory.md"
    if docs_dir and build_md_as_pdf_inventory:
        try:
            md_inventory_df = pd.DataFrame(build_md_as_pdf_inventory(docs_dir))
            if not md_inventory_df.empty:
                md_inventory_df.to_csv(md_inventory_csv, index=False)
                _write_markdown_table(md_inventory_df, md_inventory_md)
                _check(
                    checks,
                    "md_as_pdf.inventory",
                    True,
                    "INFO",
                    f"md_as_pdf_files={len(md_inventory_df)}; output={md_inventory_csv}",
                )
            else:
                _check(checks, "md_as_pdf.inventory", True, "INFO", "No MD_AS_PDF_TEXT files staged.")
        except Exception as exc:
            _check(checks, "md_as_pdf.inventory", False, "WARN", str(exc))
    else:
        _check(checks, "md_as_pdf.inventory", True, "INFO", "No docs_dir or inventory builder unavailable.")

    pre_hard_failures = [c for c in checks if c["severity"] == "HARD" and c["status"] != "PASS"]
    pre_warning_failures = [c for c in checks if c["severity"] == "WARN" and c["status"] != "PASS"]

    self_check = {
        "generated_at": generated_at,
        "hard_pass": not pre_hard_failures,
        "hard_failure_count": len(pre_hard_failures),
        "warning_count": len(pre_warning_failures),
        "check_count": len(checks),
        "checks": checks,
        "pod_reconciliation_csv": str(pod_csv) if pod_csv.exists() else None,
        "pod_reconciliation_md": str(pod_md) if pod_md.exists() else None,
        "md_as_pdf_evidence_inventory_csv": str(md_inventory_csv) if md_inventory_csv.exists() else None,
        "md_as_pdf_evidence_inventory_md": str(md_inventory_md) if md_inventory_md.exists() else None,
        "workbook_self_check_sheet_added": False,
        "workbook_pod_reconciliation_sheet_added": False,
        "workbook_md_as_pdf_evidence_sheet_added": False,
    }

    if workbook and workbook.exists():
        appended = _append_self_check_to_workbook(workbook, checks, pod_df, self_check, md_inventory_df)
        self_check["workbook_self_check_sheet_added"] = bool(appended)
        self_check["workbook_pod_reconciliation_sheet_added"] = bool(self_check.get("pod_reconciliation_sheet_added"))
        self_check["workbook_md_as_pdf_evidence_sheet_added"] = bool(self_check.get("md_as_pdf_evidence_sheet_added"))
        if not appended:
            _check(checks, "workbook.self_check_sheet_added", False, "HARD", "; ".join(self_check.get("append_errors", [])))
        else:
            _check(checks, "workbook.self_check_sheet_added", True, "HARD", "self_check sheet appended")
            if not md_inventory_df.empty:
                _check(
                    checks,
                    "workbook.md_as_pdf_evidence_sheet_added",
                    bool(self_check.get("md_as_pdf_evidence_sheet_added")),
                    "HARD",
                    "md_as_pdf_evidence sheet appended",
                )

    hard_failures = [c for c in checks if c["severity"] == "HARD" and c["status"] != "PASS"]
    warning_failures = [c for c in checks if c["severity"] == "WARN" and c["status"] != "PASS"]
    self_check["hard_pass"] = not hard_failures
    self_check["hard_failure_count"] = len(hard_failures)
    self_check["warning_count"] = len(warning_failures)
    self_check["check_count"] = len(checks)
    self_check["checks"] = checks

    report_path = outdir / "self_verification_report.json"
    report_path.write_text(json.dumps(self_check, ensure_ascii=False, indent=2), encoding="utf-8")
    return self_check


def run_audit_once(runtime_root: Path, invoice: Path, docs_dir: Path | None, outdir: Path, timeout: int):
    audit_script = runtime_root / "patch" / "run_domestic_audit_v2.py"
    command = [
        sys.executable,
        str(audit_script),
        "--invoice",
        str(invoice),
        "--mapping",
        str(runtime_root / "Data" / "DOMESTIC_with_distances.xlsx"),
        "--ledger",
        str(runtime_root / "Data" / "domestic_rate_ledger.json"),
        "--config",
        str(runtime_root / "patch" / "config_domestic_v2.json"),
        "--outdir",
        str(outdir),
    ]
    if docs_dir:
        command.extend(["--docs", str(docs_dir)])

    env = os.environ.copy()
    pythonpath = [str(runtime_root), str(runtime_root / "utils")]
    if env.get("PYTHONPATH"):
        pythonpath.append(env["PYTHONPATH"])
    env["PYTHONPATH"] = os.pathsep.join(pythonpath)

    started = datetime.now()
    proc = subprocess.run(
        command,
        cwd=str(runtime_root),
        env=env,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        timeout=timeout,
    )
    ended = datetime.now()
    return command, proc, started, ended


def write_markdown_summary(summary_path: Path, payload: dict) -> None:
    lines = [
        "# HVDC Domestic GPT Run Summary",
        "",
        f"- Status: `{payload['status']}`",
        f"- Invoice: `{payload['invoice']}`",
        f"- Exit code: `{payload['exit_code']}`",
        f"- Runtime root: `{payload['runtime_root']}`",
        f"- Output directory: `{payload['outdir']}`",
        f"- Supporting docs: `{payload.get('docs_dir') or ''}`",
        "",
        "## Outputs",
    ]
    for key, value in payload["outputs"].items():
        lines.append(f"- {key}: `{value or ''}`")
    lines.extend(["", "## Item Summary"])
    for key, value in payload.get("item_summary", {}).items():
        lines.append(f"- {key}: `{value}`")
    if payload.get("self_check") is not None:
        sc = payload["self_check"]
        lines.extend(
            [
                "",
                "## Self Verification",
                f"- hard_pass: `{sc.get('hard_pass')}`",
                f"- hard_failure_count: `{sc.get('hard_failure_count')}`",
                f"- warning_count: `{sc.get('warning_count')}`",
                f"- check_count: `{sc.get('check_count')}`",
                f"- workbook_self_check_sheet_added: `{sc.get('workbook_self_check_sheet_added')}`",
            ]
        )
    if payload.get("stderr_review") is not None:
        lines.extend(["", "## Stderr Review"])
        for key, value in payload.get("stderr_review", {}).items():
            lines.append(f"- {key}: `{value}`")
    lines.extend(["", "## Log Tail", "```text", payload.get("log_tail", ""), "```", ""])
    summary_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Run HVDC Domestic audit inside ChatGPT Code Interpreter.")
    parser.add_argument("--invoice", default=None, help="Uploaded invoice .xlsx/.xls/.csv path, or .md/.txt table.")
    parser.add_argument("--input-url", default=None, help="Optional direct URL to an invoice file. Requires network access.")
    parser.add_argument("--search-dir", default=None, help="Directory to search for uploaded invoice.")
    parser.add_argument("--runtime-root", default=None, help="Runtime root containing patch/ and Data/.")
    parser.add_argument("--docs", default=None, help="Optional supporting documents folder. If omitted, uploaded PDF/MD PODs are staged automatically.")
    parser.add_argument("--outdir", default=None, help="Output directory. Defaults to gpts_outputs/run_TIMESTAMP.")
    parser.add_argument("--timeout", type=int, default=900, help="Audit timeout in seconds.")
    args = parser.parse_args()

    runtime_root = runtime_root_from(Path(args.runtime_root).expanduser() if args.runtime_root else Path.cwd())
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    outdir = Path(args.outdir).expanduser().resolve() if args.outdir else runtime_root / "gpts_outputs" / f"run_{timestamp}"
    outdir.mkdir(parents=True, exist_ok=True)

    scratch_dir = outdir / "input_staging"
    scratch_dir.mkdir(parents=True, exist_ok=True)

    invoice = find_invoice(runtime_root, args.invoice, args.search_dir, scratch_dir, args.input_url)
    docs_dir = stage_supporting_documents(runtime_root, outdir, invoice, args.docs, args.search_dir)

    command, proc, started, ended = run_audit_once(runtime_root, invoice, docs_dir, outdir, args.timeout)

    stdout_text = proc.stdout or ""
    stderr_text = proc.stderr or ""
    stderr_review = split_known_benign_stderr(stderr_text)

    log_text = stdout_text + ("\n\n[STDERR]\n" + stderr_text if stderr_text else "")
    log_path = outdir / "domestic_audit_run_log.txt"
    log_path.write_text(log_text, encoding="utf-8")

    outputs = find_outputs(outdir)
    self_check = None
    if outputs.get("audit_workbook") or outputs.get("items_csv"):
        self_check = run_self_verification(invoice, outdir, outputs, docs_dir)
        outputs = find_outputs(outdir)

    item_summary = read_output_summary(outdir / "items.csv")

    classification_log_text = stdout_text
    if stderr_review["non_benign_stderr"]:
        classification_log_text += "\n\n[STDERR]\n" + stderr_review["non_benign_stderr"]
    status = classify_status(proc.returncode, classification_log_text, outputs, self_check)
    payload = {
        "status": status,
        "invoice": str(invoice),
        "docs_dir": str(docs_dir) if docs_dir else None,
        "runtime_root": str(runtime_root),
        "outdir": str(outdir),
        "command": command,
        "exit_code": proc.returncode,
        "started_at": started.isoformat(timespec="seconds"),
        "ended_at": ended.isoformat(timespec="seconds"),
        "duration_seconds": round((ended - started).total_seconds(), 3),
        "outputs": outputs,
        "item_summary": item_summary,
        "self_check": {
            k: v for k, v in (self_check or {}).items()
            if k not in {"checks"}
        } if self_check else None,
        "stderr_review": {
            "raw_stderr_nonblank": stderr_review["raw_stderr_nonblank"],
            "ignored": stderr_review["ignored"],
            "known_benign_only": stderr_review["known_benign_only"],
            "non_benign_stderr_nonblank": stderr_review["non_benign_stderr_nonblank"],
        },
        "log_path": str(log_path),
        "log_tail": "\n".join(log_text.splitlines()[-80:]),
    }

    summary_json = outdir / "gpts_run_summary.json"
    summary_md = outdir / "gpts_run_summary.md"
    summary_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_markdown_summary(summary_md, payload)

    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0 if status == "succeeded" else 2


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except subprocess.TimeoutExpired as exc:
        print(json.dumps({"status": "failed", "error": f"timeout: {exc}"}, ensure_ascii=False, indent=2))
        raise SystemExit(2)
    except Exception as exc:
        print(json.dumps({"status": "failed", "error": str(exc)}, ensure_ascii=False, indent=2))
        raise SystemExit(2)
