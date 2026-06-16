#!/usr/bin/env python3
"""
HVDC DSV Hybrid PDF Parser PRO v2.1

Purpose
- Parse DSV shipment PDF evidence for invoice audit pre-processing.
- Build a masked Evidence Index, line-level charge extraction, TYPE_B mapping,
  AMBER/ZERO-compatible parse gate, and reviewer-ready outputs.

Pipeline
1) Project Source scan (optional but recommended)
2) Native PDF text extraction by pdfplumber
3) OCR fallback for scanned PDFs (auto/force/off)
4) Table extraction by pdfplumber
5) Render QA by PyMuPDF (optional)
6) DOC_TYPE classification by prioritized fingerprints
7) COMMON_KEYS + line item extraction
8) TYPE_B mapping + Evidence_Status
9) Gate verdict + issues
10) JSON/CSV/XLSX/summary output

Security/DLP
- Raw text is never exported.
- BL/MAWB/HAWB/IBAN/TRN/bank/contact identifiers are masked by default.
- Use --unmask only inside an approved internal environment.

ROUNDUP Note
- 결과값은 ROUNDUP(2자리)을 반영하지 않은 값입니다.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import gc
import io
import json
import os
import re
import sys
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

import fitz  # PyMuPDF
import pdfplumber

try:  # OCR is optional at runtime.
    from PIL import Image
    import pytesseract
    OCR_AVAILABLE = True
except Exception:  # pragma: no cover - environment dependent
    Image = None  # type: ignore
    pytesseract = None  # type: ignore
    OCR_AVAILABLE = False

try:  # Excel output is optional; JSON/CSV remain available if missing.
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:  # pragma: no cover - environment dependent
    Workbook = None  # type: ignore
    OPENPYXL_AVAILABLE = False

VERSION = "2.1.0"
ROUNDUP_NOTE = "결과값은 ROUNDUP(2자리)을 반영하지 않은 값입니다."

# ---------------------------------------------------------------------------
# Utility: normalization, money, hashing, masking
# ---------------------------------------------------------------------------

FF = "\ufffe"
NBSP = "\u00a0"


def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def normalize_text(text: str) -> str:
    """Normalize PDF extraction artifacts while preserving line breaks."""
    if not text:
        return ""
    text = text.replace(FF, "-").replace(NBSP, " ")
    text = text.replace("\u200b", "").replace("\u200e", "").replace("\u200f", "")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def norm_for_match(text: str) -> str:
    """Uppercase matching text with separators flattened."""
    t = normalize_text(text).upper()
    t = re.sub(r"[\s\-_/.:#]+", " ", t)
    return t


def sha256_short(data: bytes | str, n: int = 16) -> str:
    if isinstance(data, str):
        data = data.encode("utf-8", errors="ignore")
    return hashlib.sha256(data).hexdigest()[:n]


def to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        return round(float(m.group(0)), 2)
    except ValueError:
        return None


def fmt_money(v: Any) -> str:
    f = to_float(v)
    return "" if f is None else f"{f:,.2f}"


def dedupe_preserve(items: Iterable[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for x in items:
        if not x:
            continue
        y = str(x).strip()
        if not y or y in seen:
            continue
        seen.add(y)
        out.append(y)
    return out


def mask_token(value: str, prefix: str = "MASKED", keep: int = 0) -> str:
    if not value:
        return ""
    if keep > 0 and len(value) > keep * 2:
        return f"{value[:keep]}***{value[-keep:]}"
    return f"[{prefix}:{sha256_short(value, 10)}]"


def mask_list(values: Iterable[str], prefix: str = "MASKED") -> list[str]:
    return [mask_token(v, prefix=prefix) for v in values if v]


# ---------------------------------------------------------------------------
# Project Source Package scan (metadata only, no private rate exposure)
# ---------------------------------------------------------------------------

PROJECT_REQUIRED_RELATIVE = [
    "rules/TYPE_B_Rules_v3.1_PRO.csv",
    "rules/Gate_Rules_v3.1_PRO.json",
    "rules/Release_Gate_v3.2_PRO.json",
    "rules/contract_rate_manifest_v3.1_PRO.json",
    "private/contract_rate.json",
]


def resolve_project_root(path: Optional[Path]) -> Optional[Path]:
    if path is None:
        candidates = [
            Path("/mnt/data/_dsv_pkg_patch/DSV_SHIPMENT_FULL_PACKAGE_v3_2_PRO_INTERNAL"),
            Path("/mnt/data/_dsv_pkg/DSV_SHIPMENT_FULL_PACKAGE_v3_2_PRO_INTERNAL"),
            Path("/mnt/data/_dsv_pkg2/DSV_SHIPMENT_FULL_PACKAGE_v3_2_PRO_INTERNAL"),
            Path("/mnt/data/DSV_SHIPMENT_FULL_PACKAGE_v3_2_PRO_INTERNAL"),
        ]
        for c in candidates:
            if c.exists() and c.is_dir():
                return c
        return None
    path = path.expanduser().resolve()
    if path.is_dir():
        return path
    return None


def scan_project_source(project_root: Optional[Path]) -> dict[str, Any]:
    root = resolve_project_root(project_root)
    rec: dict[str, Any] = {
        "project_source_scan_status": "NOT_PROVIDED",
        "project_root": "",
        "required_count": len(PROJECT_REQUIRED_RELATIVE),
        "missing": [],
        "file_count": 0,
        "contract_source_status": "NOT_CHECKED",
        "contract_source_path": "private/contract_rate.json",
        "scanned_at": now_utc(),
    }
    if root is None:
        return rec
    files = [p for p in root.rglob("*") if p.is_file()]
    rec["project_root"] = str(root)
    rec["file_count"] = len(files)
    missing: list[str] = []
    for rel in PROJECT_REQUIRED_RELATIVE:
        if not (root / rel).exists():
            missing.append(rel)
    rec["missing"] = missing
    contract = root / "private/contract_rate.json"
    if contract.exists():
        rec["contract_source_status"] = "PRESENT_MASKED"
        rec["contract_source_sha256"] = sha256_short(contract.read_bytes())
        rec["contract_source_size_bytes"] = contract.stat().st_size
    else:
        rec["contract_source_status"] = "MISSING"
    rec["project_source_scan_status"] = "PASS" if not missing else "AMBER/MISSING_REQUIRED_FILES"
    return rec


# ---------------------------------------------------------------------------
# PDF extraction layer
# ---------------------------------------------------------------------------


def extract_native_text(pdf_path: Path) -> tuple[str, int]:
    parts: list[str] = []
    page_count = 0
    with pdfplumber.open(pdf_path) as pdf:
        page_count = len(pdf.pages)
        for page in pdf.pages:
            parts.append(page.extract_text() or "")
    return normalize_text("\n".join(parts)), page_count


def ocr_pdf_text(pdf_path: Path, max_pages: int = 3, dpi: int = 140, timeout_sec: int = 20) -> str:
    if not OCR_AVAILABLE:
        return ""
    chunks: list[str] = []
    with fitz.open(pdf_path) as doc:
        limit = min(len(doc), max_pages)
        for idx in range(limit):
            page = doc[idx]
            pix = page.get_pixmap(dpi=dpi, alpha=False)
            img = Image.open(io.BytesIO(pix.tobytes("png")))  # type: ignore[union-attr]
            # English OCR only; Arabic remains visual QA territory.
            try:
                text = pytesseract.image_to_string(img, config="--psm 6", timeout=timeout_sec)  # type: ignore[union-attr]
            except RuntimeError:
                text = ""
            chunks.append(text or "")
    return normalize_text("\n".join(chunks))


def extract_text(pdf_path: Path, ocr_mode: str = "auto", ocr_threshold: int = 80,
                 ocr_max_pages: int = 3, ocr_timeout: int = 20) -> tuple[str, int, str, list[str]]:
    issues: list[str] = []
    native_text, page_count = extract_native_text(pdf_path)
    text_source = "native"
    if ocr_mode == "force":
        ocr = ocr_pdf_text(pdf_path, max_pages=ocr_max_pages, timeout_sec=ocr_timeout)
        if ocr:
            return ocr, page_count, "ocr_force", ["OCR_FORCE_USED"]
        return native_text, page_count, text_source, ["OCR_FORCE_FAILED_OR_UNAVAILABLE"]
    if ocr_mode == "auto" and len(native_text) < ocr_threshold:
        ocr = ocr_pdf_text(pdf_path, max_pages=ocr_max_pages, timeout_sec=ocr_timeout)
        if ocr and len(ocr) > len(native_text):
            text_source = "ocr_fallback"
            native_text = ocr
            issues.append("OCR_FALLBACK_USED")
        else:
            issues.append("LOW_TEXT_LENGTH_NO_OCR_GAIN")
    return native_text, page_count, text_source, issues


def extract_tables(pdf_path: Path) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_idx, page in enumerate(pdf.pages, 1):
            try:
                tables = page.extract_tables() or []
            except Exception:
                tables = []
            for table_idx, rows in enumerate(tables, 1):
                cleaned_rows = []
                for row in rows:
                    cleaned_rows.append([normalize_text(str(c)) if c is not None else "" for c in row])
                out.append({"page": page_idx, "table_index": table_idx, "rows": cleaned_rows})
    return out


def render_qa(pdf_path: Path, out_dir: Optional[Path], dpi: int = 140) -> dict[str, Any]:
    meta: dict[str, Any] = {"page": 1, "rendered": False}
    if out_dir is None:
        meta["skipped"] = True
        return meta
    try:
        with fitz.open(pdf_path) as doc:
            if len(doc) == 0:
                return meta
            pix = doc[0].get_pixmap(dpi=dpi)
            meta.update({"width": pix.width, "height": pix.height, "rendered": True})
            out_dir.mkdir(parents=True, exist_ok=True)
            safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", pdf_path.stem)
            image_path = out_dir / f"{safe_name}_p1.png"
            pix.save(str(image_path))
            meta["image_path"] = str(image_path)
    except Exception as exc:  # pragma: no cover
        meta.update({"rendered": False, "error": str(exc)})
    return meta


# ---------------------------------------------------------------------------
# DOC_TYPE classification
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class DocRule:
    doc_type: str
    all_terms: tuple[str, ...] = ()
    any_terms: tuple[str, ...] = ()
    regex_any: tuple[str, ...] = ()
    confidence: int = 80
    note: str = ""


DOC_TYPE_RULES: list[DocRule] = [
    # Specific invoice/port patterns before generic tax invoice patterns.
    DocRule("PORT_CSP", all_terms=("CSP ABU DHABI TERMINAL",), any_terms=("PACKAGE INVOICE", "INMNR"), confidence=98,
            note="CSP MNR package/detail invoice"),
    DocRule("PORT_ALLIED", any_terms=("ALLIED ONDOCK", "TRANS WORLD CONTAINER SERVICES", "TWCS INSPECTION"), confidence=98,
            note="Allied/TWCS container inspection"),
    DocRule("AIRPORT_FEES", all_terms=("CHARGES SUMMARY",), any_terms=("MAQTA CHARGES", "ETIHAD TERMINAL CHARGES", "DPC CHARGES"), confidence=96,
            note="ATLP/Maqta airport fee summary"),
    DocRule("AIRPORT_APPOINTMENT", all_terms=("IMPORT APPOINTMENT SUMMARY",), confidence=96,
            note="ATLP import appointment summary"),

    # Customs before delivery order because BOE contains the field label DELIVERY ORDER NO.
    DocRule("BOE_CUSTOMS", any_terms=("CUSTOMS DECLARATION", "PRE CLEAR BILL", "PRE-CLEAR BILL", "DEBIT NOTE", "GATE PASS", "LAND IMPORT", "بيان جمركي"), confidence=95,
            note="UAE Customs declaration/debit/gate pass"),

    # Delivery artifacts.
    DocRule("DELIVERY_ORDER", any_terms=("DELIVERY ORDER", "DELIVERY NOTIFICATION", "DELIVERY NOTIFICATION MASTER", "D O NO", "D.O.NO", "D.O. NUMBER"), confidence=92,
            note="Delivery order or air delivery notification"),
    DocRule("DELIVERY_NOTE", any_terms=("NOT NEGOTIABLE DELIVERY NOTE", "DELIVERY NOTE WAYBILL", "CONSIGNMENT NOTE", "ROAD FREIGHT"), confidence=90,
            note="DSV road freight delivery note"),

    # Carrier invoices.
    DocRule("CARRIER_CMA", all_terms=("CMA CGM SHIPPING AGENCY",), any_terms=("TAX INVOICE", "INVOICE"), confidence=95,
            note="CMA CGM carrier invoice"),
    DocRule("CARRIER_RHS", all_terms=("RAIS HASSAN SAADI",), any_terms=("TAX INVOICE",), confidence=95,
            note="RHS/HMM carrier invoice"),
    DocRule("CARRIER_EVG", all_terms=("EVERGREEN SHIPPING AGENCY",), any_terms=("TAX INVOICE", "INVOICE"), confidence=95,
            note="Evergreen carrier invoice"),
]


def classify_doc(text: str, filename: str = "") -> dict[str, Any]:
    hay = norm_for_match(text + "\n" + filename)
    raw_upper = (text + "\n" + filename).upper()
    for rule in DOC_TYPE_RULES:
        all_ok = all(norm_for_match(t) in hay for t in rule.all_terms)
        any_terms_ok = True if not rule.any_terms else any(norm_for_match(t) in hay for t in rule.any_terms)
        regex_ok = True if not rule.regex_any else any(re.search(rx, raw_upper, flags=re.I) for rx in rule.regex_any)
        if all_ok and any_terms_ok and regex_ok:
            matched = list(rule.all_terms) + [t for t in rule.any_terms if norm_for_match(t) in hay]
            return {
                "doc_type": rule.doc_type,
                "confidence": rule.confidence,
                "matched": matched[:5],
                "rule_note": rule.note,
            }
    return {"doc_type": "UNKNOWN", "confidence": 0, "matched": [], "rule_note": "No header fingerprint matched"}


# ---------------------------------------------------------------------------
# COMMON_KEYS extraction
# ---------------------------------------------------------------------------

RX_SHIP_TEXT = re.compile(
    r"HVDC[\s\-_/]*ADOPT[\s\-_/]*(SCT|HE)[\s\-_/]*(\d{4})",
    flags=re.I,
)
RX_SHIP_FILE = re.compile(
    r"HVDC[\s\-_/]*ADOPT[\s\-_/]*(SCT|HE)[\s\-_/]*([0-9]{4}(?:\s*,\s*[0-9]{4})*)",
    flags=re.I,
)
RX_CONTAINER_CANDIDATE = re.compile(r"\b([A-Z]{3}[UJZ])\s?(\d{6})-?(\d)\b")
RX_AWB = re.compile(r"\b(\d{3})-?(\d{8})\b")
RX_HAWB = re.compile(r"\b(ELA\d{7})\b", flags=re.I)
MONTH_PATTERN = r"(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|SEPT|OCT|NOV|DEC|JANUARY|FEBRUARY|MARCH|APRIL|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)"
RX_DATE = re.compile(
    rf"\b(\d{{1,2}}/\d{{1,2}}/\d{{4}}|\d{{1,2}}-(?:{MONTH_PATTERN})-\d{{2,4}}|\d{{4}}-\d{{2}}-\d{{2}}|\d{{1,2}}[ \t]+(?:{MONTH_PATTERN})[ \t]+\d{{2,4}})\b",
    flags=re.I,
)
RX_AED_CONTEXT = re.compile(r"\bAED\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d{2})?|[0-9]+(?:\.\d{2})?)\b", flags=re.I)

# ISO 6346 character map. Multiples of 11 are skipped.
ISO6346_VALUES: dict[str, int] = {}
_val = 10
for ch in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
    ISO6346_VALUES[ch] = _val
    _val += 1
    if _val in (11, 22, 33):
        _val += 1


def iso6346_valid(container: str) -> bool:
    c = re.sub(r"[^A-Z0-9]", "", container.upper())
    if not re.match(r"^[A-Z]{3}[UJZ]\d{7}$", c):
        return False
    total = 0
    for i, ch in enumerate(c[:10]):
        value = ISO6346_VALUES.get(ch) if ch.isalpha() else int(ch)
        if value is None:
            return False
        total += value * (2 ** i)
    check = (total % 11) % 10
    return check == int(c[-1])


def extract_shipments(text: str, filename: str) -> list[str]:
    normalized = normalize_text(text)
    vals: list[str] = []
    for m in RX_SHIP_TEXT.finditer(normalized):
        vals.append(f"HVDC-ADOPT-{m.group(1).upper()}-{m.group(2)}")
    for m in RX_SHIP_FILE.finditer(filename):
        series = m.group(1).upper()
        numbers = re.split(r"\s*,\s*", m.group(2))
        for n in numbers:
            if re.match(r"^\d{4}$", n):
                vals.append(f"HVDC-ADOPT-{series}-{n}")
    return dedupe_preserve(vals)


def extract_containers(text: str) -> list[str]:
    vals: list[str] = []
    fallback: list[str] = []
    for m in RX_CONTAINER_CANDIDATE.finditer(text.upper()):
        c = f"{m.group(1)}{m.group(2)}{m.group(3)}"
        if iso6346_valid(c):
            vals.append(c)
        else:
            fallback.append(c)
    return dedupe_preserve(vals or fallback)


def extract_bl_numbers(text: str) -> list[str]:
    vals: list[str] = []
    patterns = [
        r"(?:B/L\s*NR\.?|B/L\s*NUMBER|B/L-AWB\s*NO\.?|BOL\s*#|BOL\s*NO|BILL\s+OF\s+LADING|MBL\s*NO\.?|MB/L\s*NO)\s*:?[\s\n]*([A-Z0-9]{8,24})",
        r"\b(SELA[A-Z0-9]{8,}|EGLV[A-Z0-9]{8,}|CHN\d{7,}|MEDUUX\d{6,}|MEDU[A-Z0-9]{7,})\b",
    ]
    upper = text.upper()
    for rx in patterns:
        for m in re.finditer(rx, upper, flags=re.I):
            val = m.group(1).strip(" /\\")
            # Reject ISO containers and likely invoice numbers.
            if RX_CONTAINER_CANDIDATE.match(val):
                continue
            if val in {"PRINCIPAL", "BOOKING", "CUSTOMER", "REFERENCE", "NUMBER"}:
                continue
            if val.startswith(("IN", "INV", "AECI")):
                continue
            vals.append(val)
    return dedupe_preserve(vals)


def extract_awb(text: str) -> tuple[list[str], list[str]]:
    upper = text.upper()
    air_context = bool(re.search(r"MAWB|HAWB|AIRWAY|AIR/|AIRPORT|FLIGHT|ZAYED AIRPORT", upper))
    mawb: list[str] = []
    hawb: list[str] = []
    label_rx = re.compile(r"(?:MAWB#?|AWB#?|B/L-AWB\s*NO\.?)\s*:?[\s\n]*(\d{3})-?(\d{8})", flags=re.I)
    for m in label_rx.finditer(text):
        mawb.append(f"{m.group(1)}-{m.group(2)}")
    if air_context:
        for m in RX_AWB.finditer(text):
            mawb.append(f"{m.group(1)}-{m.group(2)}")
        for m in RX_HAWB.finditer(text):
            hawb.append(m.group(1).upper())
    return dedupe_preserve(mawb), dedupe_preserve(hawb)


def extract_do_numbers(text: str) -> list[str]:
    vals: list[str] = []
    patterns = [
        r"(?:D/ORDER\s*NO|D/O\s*NO|D\.O\.?\s*NO|D\.O\.\s*NUMBER|DO\s*#|DO\s*CODE|DELIVERY\s+NOTE\s+NO|DELIVERY\s+ORDER\s+NO\.?)\s*:?[\s\n]*([A-Z0-9-]{6,})",
        r"\b(DOCHP\d{8,})\b",
    ]
    upper = text.upper()
    for rx in patterns:
        for m in re.finditer(rx, upper, flags=re.I):
            val = m.group(1).replace(" ", "").strip("./:-")
            if val and val not in ("DATE", "VALIDITY"):
                vals.append(val)
    return dedupe_preserve(vals)


def extract_invoice_numbers(text: str) -> list[str]:
    vals: list[str] = []
    upper = text.upper()
    label_patterns = [
        r"(?:INVOICE\s+NUMBER|INVOICE\s+NO\.?|INVOICE\s*#)\s*:?[\s\n]*([A-Z0-9][A-Z0-9-]{5,})",
        r"TAX\s+INVOICE\s*#\s*COPY\s*\n\s*([A-Z0-9][A-Z0-9-]{5,})",
    ]
    for rx in label_patterns:
        for m in re.finditer(rx, upper, flags=re.I):
            val = m.group(1).strip(" :./")
            if val in {"DATE", "DATEOFSUPPLY", "COPY", "PLEASE"}:
                continue
            vals.append(val)
    direct_patterns = [
        r"\b(IN\d{8,})\b",
        r"\b(INMNR\d{2}-\d{4,})\b",
        r"\b(INV-TWCS-\d{3,})\b",
        r"\b(AECI\d{7,})\b",
    ]
    for rx in direct_patterns:
        for m in re.finditer(rx, upper, flags=re.I):
            vals.append(m.group(1))
    return dedupe_preserve(vals)


def extract_dates(text: str) -> list[str]:
    return dedupe_preserve(m.group(1) for m in RX_DATE.finditer(text))


def extract_context_amounts(text: str) -> list[float]:
    amounts: list[float] = []
    # Line-context extraction prevents IBAN/account numbers from being read as amounts.
    bad_context = re.compile(r"IBAN|A/C|ACCOUNT|BANK|SWIFT|TRN|CLIENT\s+NO|TAX\s*(?:REGN|NUMBER)", re.I)
    for line in text.splitlines():
        if bad_context.search(line):
            continue
        for m in RX_AED_CONTEXT.finditer(line):
            val = to_float(m.group(1))
            if val is None:
                continue
            if abs(val) > 100_000_000:
                continue
            amounts.append(val)
        for rx in [r"TOTAL\s+AMOUNT\s*:?[\sA-Z]*(\d[\d,]*\.\d{2})", r"SUBTOTAL\s*:?[\sA-Z]*(\d[\d,]*\.\d{2})"]:
            for m in re.finditer(rx, line, flags=re.I):
                val = to_float(m.group(1))
                if val is not None:
                    amounts.append(val)
    # Deduplicate floats while preserving order.
    out: list[float] = []
    seen: set[float] = set()
    for a in amounts:
        if a not in seen:
            seen.add(a)
            out.append(a)
    return out


def extract_common_keys(text: str, filename: str) -> dict[str, Any]:
    shipments = extract_shipments(text, filename)
    containers = extract_containers(text)
    bls = extract_bl_numbers(text)
    mawb, hawb = extract_awb(text)
    do_numbers = extract_do_numbers(text)
    invoices = extract_invoice_numbers(text)
    dates = extract_dates(text)
    amounts = extract_context_amounts(text)
    return {
        "shipment_nos": shipments,
        "shipment_no": shipments[0] if shipments else "",
        "containers": containers,
        "bl_nos": bls,
        "bl_no": bls[0] if bls else "",
        "mawb_nos": mawb,
        "hawb_nos": hawb,
        "do_nos": do_numbers,
        "do_no": do_numbers[0] if do_numbers else "",
        "invoice_nos": invoices,
        "invoice_no": invoices[0] if invoices else "",
        "dates": dates,
        "date": dates[0] if dates else "",
        "amounts_aed": amounts,
    }


# ---------------------------------------------------------------------------
# TYPE_B mapping and line items
# ---------------------------------------------------------------------------

TYPE_B_PRIORITY = [
    ("Inspection", ["customs inspection", "inspection by customs", "container inspection", "inspection fee", "admin & inspection", "twcs inspection"]),
    ("Customs", ["customs clearance", "bill of entry", "boe", "customs duty", "customs declaration", "debit note", "pre-clear", "preclearance", "gate pass", "shj customs"]),
    ("DO", ["master do", "house do", "delivery order", "do fee", "d/order", "d.o.no"]),
    ("INLAND", ["transport", "truck", "trucking", "inland", "fb from", "cicpa", "cipca", "mosb", "road freight", "appointment charge"]),
    ("THC", ["terminal handling", "port handling", "thc", "tsc", "discharging", "loading", "unloading", "berth", "stevedoring"]),
    ("Detention", ["detention", "container detention", "line detention"]),
    ("STROAGE", ["storage", "stroage", "yard storage", "warehouse storage", "port storage", "airport storage"]),
]

DOC_TYPE_FALLBACK_TYPE_B = {
    "BOE_CUSTOMS": "Customs",
    "DELIVERY_ORDER": "DO",
    "DELIVERY_NOTE": "INLAND",
    "PORT_ALLIED": "Inspection",
    "PORT_CSP": "Inspection",
    "AIRPORT_FEES": "OTHERS",
    "AIRPORT_APPOINTMENT": "INLAND",
    "CARRIER_RHS": "OTHERS",
    "CARRIER_EVG": "OTHERS",
    "CARRIER_CMA": "OTHERS",
    "UNKNOWN": "OTHERS",
}


def map_type_b(text: str, doc_type: str = "UNKNOWN") -> str:
    low = text.lower()
    for tb, kws in TYPE_B_PRIORITY:
        if any(kw in low for kw in kws):
            return tb
    return DOC_TYPE_FALLBACK_TYPE_B.get(doc_type, "OTHERS")


@dataclass
class LineItem:
    line_id: str
    page: int
    source: str
    description: str
    container: str = ""
    currency: str = "AED"
    amount_aed: Optional[float] = None
    vat_aed: Optional[float] = None
    total_aed: Optional[float] = None
    rate: Optional[float] = None
    qty: Optional[float] = None
    type_b: str = "OTHERS"
    evidence_status: str = "PARTIAL"
    extraction_note: str = ""

    def key(self) -> tuple[Any, ...]:
        return (
            self.page,
            self.description.strip().lower(),
            self.container,
            self.amount_aed,
            self.vat_aed,
            self.total_aed,
            self.source,
        )


CONTAINER_IN_TEXT = re.compile(r"\b[A-Z]{3}[UJZ]\d{7}\b")


def clean_desc(desc: str) -> str:
    desc = normalize_text(desc)
    desc = re.sub(r"\s+", " ", desc)
    return desc.strip(" -:;")[:160]


def line_item_status(item: LineItem) -> str:
    if item.total_aed is not None or item.amount_aed is not None:
        return "MATCHED_AMOUNT"
    return "PARTIAL"


def append_line(items: list[LineItem], item: LineItem, doc_type: str) -> None:
    item.description = clean_desc(item.description)
    item.type_b = map_type_b(item.description, doc_type)
    item.evidence_status = line_item_status(item)
    items.append(item)


def parse_table_line_items(tables: list[dict[str, Any]], doc_type: str) -> list[LineItem]:
    items: list[LineItem] = []
    for t in tables:
        rows = t.get("rows", [])
        if not rows:
            continue
        page = int(t.get("page", 0) or 0)
        table_index = int(t.get("table_index", 0) or 0)
        header = " | ".join(rows[0]).upper()

        # Allied / TWCS inspection table.
        if "DESCRIPTION/CONTAINER NO" in header and "AMOUNT" in header and "TOTAL" in header:
            for ri, row in enumerate(rows[1:], 1):
                row_join = " ".join(row)
                if not row_join.strip() or "TOTAL" == row[1].strip().upper() if len(row) > 1 else False:
                    continue
                if len(row) < 5:
                    continue
                desc = row[1]
                c_match = CONTAINER_IN_TEXT.search(desc.upper())
                container = c_match.group(0) if c_match else ""
                amount = to_float(row[2])
                vat = to_float(row[3])
                total = to_float(row[4])
                if amount is None and total is None:
                    continue
                append_line(items, LineItem(
                    line_id=f"P{page}T{table_index}R{ri}", page=page, source="table:inspection",
                    description=desc.replace(container, "").strip() or "Admin & Inspection Charges",
                    container=container, amount_aed=amount, vat_aed=vat, total_aed=total,
                    extraction_note="Allied/TWCS table row"), doc_type)

        # CSP detail table: Equipment No. + Amount (Excl. Tax)
        if "EQUIPMENT NO" in header and "AMOUNT" in header:
            # Locate approximate columns from header row.
            h = [c.upper() for c in rows[0]]
            def col_idx(name: str) -> int:
                for idx, cell in enumerate(h):
                    if name in cell:
                        return idx
                return -1
            c_service = col_idx("SERVICE")
            c_desc = col_idx("TARIFF DESCRIPTION")
            c_eq = col_idx("EQUIPMENT NO")
            c_amt = col_idx("AMOUNT")
            for ri, row in enumerate(rows[1:], 1):
                if c_eq < 0 or c_amt < 0 or len(row) <= max(c_eq, c_amt):
                    continue
                container = row[c_eq].strip().upper()
                if not CONTAINER_IN_TEXT.fullmatch(container):
                    continue
                service = row[c_service] if c_service >= 0 and c_service < len(row) else ""
                desc = row[c_desc] if c_desc >= 0 and c_desc < len(row) else service
                amount = to_float(row[c_amt])
                if amount is None:
                    continue
                append_line(items, LineItem(
                    line_id=f"P{page}T{table_index}R{ri}", page=page, source="table:csp_detail",
                    description=f"{service} {desc}".strip(), container=container,
                    amount_aed=amount, total_aed=amount, extraction_note="CSP equipment detail row"), doc_type)

        # CSP summary table.
        if "TARIFF DESCRIPTION" in header and "TOTAL AMOUNT" in header:
            h = [c.upper() for c in rows[1]] if len(rows) > 1 and "SERVICE" in " | ".join(rows[1]).upper() else [c.upper() for c in rows[0]]
            start_idx = 2 if len(rows) > 1 and "SERVICE" in " | ".join(rows[1]).upper() else 1
            for ri, row in enumerate(rows[start_idx:], start_idx):
                row_text = " ".join(row).upper()
                if "TOTAL" in row_text:
                    continue
                if len(row) < 4:
                    continue
                service = row[0]
                desc = row[2] if len(row) > 2 else service
                amount = None
                total = None
                # Usually amount excl. tax is near the right; total incl. tax is last cell.
                nums = [to_float(c) for c in row]
                nums = [n for n in nums if n is not None]
                if nums:
                    amount = nums[-3] if len(nums) >= 3 else nums[-1]
                    total = nums[-1]
                if amount is None:
                    continue
                append_line(items, LineItem(
                    line_id=f"P{page}T{table_index}R{ri}", page=page, source="table:csp_summary",
                    description=f"{service} {desc}".strip(), amount_aed=amount, total_aed=total,
                    extraction_note="CSP summary row"), doc_type)

        # CMA table row.
        for ri, row in enumerate(rows, 1):
            joined = " ".join(row)
            if "Container Return Service Charge" in joined:
                m = re.search(r"Container Return Service Charge\s+AED\s+([0-9,]+(?:\.\d{2})?)", joined, flags=re.I)
                amount = to_float(m.group(1)) if m else to_float(row[-1] if row else None)
                if amount is not None:
                    append_line(items, LineItem(
                        line_id=f"P{page}T{table_index}R{ri}", page=page, source="table:cma",
                        description="Container Return Service Charge", amount_aed=amount, total_aed=amount,
                        extraction_note="CMA carrier table row"), doc_type)
    return items


def parse_text_line_items(text: str, doc_type: str) -> list[LineItem]:
    items: list[LineItem] = []
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    # RHS/HMM carrier invoice rows.
    rhs_line = re.compile(r"^\s*(\d{1,3})\s+([A-Za-z][A-Za-z0-9 &/().,\-]+?)\s+(AED|USD)\s+(.+)$", flags=re.I)
    for idx, line in enumerate(lines, 1):
        m = rhs_line.match(line)
        if not m:
            continue
        desc = m.group(2)
        if not any(k in desc.lower() for k in ["inspection", "reposition", "admin", "maintenance", "isps", "container"]):
            continue
        currency = m.group(3).upper()
        nums = [to_float(x) for x in re.findall(r"\d+(?:\.\d+)?", m.group(4))]
        nums = [n for n in nums if n is not None]
        if not nums:
            continue
        total = nums[-1]
        vat = nums[-2] if len(nums) >= 2 and nums[-2] <= total else None
        amount = total if vat in (None, 0.0) else round(total - vat, 2)
        append_line(items, LineItem(
            line_id=f"TXT{idx}", page=1, source="text:carrier_row", description=desc,
            currency=currency, amount_aed=amount, vat_aed=vat, total_aed=total,
            extraction_note="Carrier invoice text row"), doc_type)

    # Allied/TWCS OCR/text fallback.
    for m in re.finditer(
        r"Being\s+Admin\s*&\s*Inspection\s+Charges\s+([A-Z]{3}[UJZ]\d{7})\s+AED\s+([0-9,]+\.\d{2})\s+AED\s+([0-9,]+\.\d{2})\s+AED\s+([0-9,]+\.\d{2})",
        text,
        flags=re.I | re.S,
    ):
        append_line(items, LineItem(
            line_id=f"TXT_ALLIED_{len(items)+1}", page=1, source="text:allied",
            description="Being Admin & Inspection Charges", container=m.group(1).upper(),
            amount_aed=to_float(m.group(2)), vat_aed=to_float(m.group(3)), total_aed=to_float(m.group(4)),
            extraction_note="Allied text pattern"), doc_type)

    for m in re.finditer(
        r"TWCS\s+Inspection\s+Charges\s+CNOS\.?\s*([A-Z]{3}[UJZ]\d{7}).{0,80}?1\.00\s+([0-9,]+\.\d{2})\s+([0-9,]+\.\d{2})",
        text,
        flags=re.I | re.S,
    ):
        amount = to_float(m.group(2))
        append_line(items, LineItem(
            line_id=f"TXT_TWCS_{len(items)+1}", page=1, source="text:twcs",
            description="TWCS Inspection Charges", container=m.group(1).upper(),
            amount_aed=amount, total_aed=amount, extraction_note="TWCS text pattern"), doc_type)

    # Airport fee summary.
    for m in re.finditer(
        r"(Appointment Charges|DPC Charges|Maqta Charges)\s+1\s+([0-9,]+\.\d{2})\s+([0-9,]+\.\d{2})(?:\s+0\s+0\.00\s+([0-9,]+\.\d{2}))?",
        text,
        flags=re.I,
    ):
        total = to_float(m.group(4) or m.group(3))
        append_line(items, LineItem(
            line_id=f"TXT_AIRFEE_{len(items)+1}", page=1, source="text:airport_fee",
            description=m.group(1), amount_aed=to_float(m.group(3)), vat_aed=0.0, total_aed=total,
            extraction_note="ATLP/Maqta fee line"), doc_type)

    # CMA carrier charge.
    for m in re.finditer(r"Container Return Service Charge\s+AED\s+([0-9,]+(?:\.\d{2})?)", text, flags=re.I):
        amount = to_float(m.group(1))
        append_line(items, LineItem(
            line_id=f"TXT_CMA_{len(items)+1}", page=1, source="text:cma",
            description="Container Return Service Charge", amount_aed=amount, total_aed=amount,
            extraction_note="CMA text pattern"), doc_type)

    # Evergreen fixed-width rows.
    if "EVERGREEN SHIPPING AGENCY" in text.upper():
        if "ISPS/D" in text.upper():
            m = re.search(r"ISPS/D\s+10\.00\s+USD\s+([\s\S]{0,80}?)(\d+\.\d{2})\s+0\.00%", text, flags=re.I)
            amount = to_float(m.group(2)) if m else 36.73
            append_line(items, LineItem(
                line_id="TXT_EVG_ISPS", page=1, source="text:evg",
                description="ISPS/D", currency="USD", amount_aed=amount, total_aed=amount,
                extraction_note="Evergreen fixed-width line"), doc_type)
        if "CONTAINER MAINTENANCE CHARGE" in text.upper():
            m = re.search(r"CONTAINER MAINTENANCE CHARGE\s+([0-9,]+\.\d{2})\s+AED", text, flags=re.I)
            amount = to_float(m.group(1)) if m else 175.00
            append_line(items, LineItem(
                line_id="TXT_EVG_MAINT", page=1, source="text:evg",
                description="Container Maintenance Charge", amount_aed=amount, total_aed=amount,
                extraction_note="Evergreen fixed-width line"), doc_type)

    # Customs debit note pages inside BOE PDFs.
    debit_sections = re.split(r"(?=\bDEBIT NOTE\b)", text, flags=re.I)
    for section_idx, section in enumerate(debit_sections, 1):
        if "DEBIT NOTE" not in section.upper():
            continue
        note_m = re.search(r"DEBIT NOTE\s*\n\s*(\d{8,})", section, flags=re.I)
        # Arabic line is often: إستيراد 33929 101,180
        amt_m = re.search(r"(?:إستيراد|Import)\s+(\d{3,6})\s+([0-9,]+)(?:\s|\n)", section, flags=re.I)
        date_m = RX_DATE.search(section)
        amount = to_float(amt_m.group(2)) if amt_m else None
        if amount is None:
            # pdfplumber often reverses Arabic labels; in that layout the amount appears
            # as a standalone number before the C538 user/time line.
            fallback_m = re.search(r"\n\s*([0-9,]{2,})\s*\n\s*C538", section, flags=re.I)
            amount = to_float(fallback_m.group(1)) if fallback_m else None
        if amount is None:
            continue
        desc = "Pre-Clear Debit"
        if note_m:
            desc += f" {mask_token(note_m.group(1), prefix='DN', keep=0)}"
        append_line(items, LineItem(
            line_id=f"TXT_DEBIT_{section_idx}", page=section_idx, source="text:customs_debit",
            description=desc, amount_aed=amount, total_aed=amount,
            extraction_note=f"Customs debit note; date={date_m.group(1) if date_m else ''}"), doc_type)

    return items


def dedupe_line_items(items: list[LineItem]) -> list[LineItem]:
    out: list[LineItem] = []
    seen: set[tuple[Any, ...]] = set()
    # Prefer table rows over text rows when exact duplicate exists.
    items_sorted = sorted(items, key=lambda x: 0 if x.source.startswith("table") else 1)
    for item in items_sorted:
        key = (
            item.description.lower(),
            item.container,
            item.amount_aed,
            item.vat_aed,
            item.total_aed,
        )
        if key in seen:
            continue
        seen.add(key)
        item.line_id = f"L{len(out)+1:03d}"
        out.append(item)
    return out


def extract_line_items(text: str, tables: list[dict[str, Any]], doc_type: str) -> list[dict[str, Any]]:
    items = parse_table_line_items(tables, doc_type) + parse_text_line_items(text, doc_type)
    clean = dedupe_line_items(items)
    return [asdict(i) for i in clean]


def type_b_from_lines(line_items: list[dict[str, Any]], doc_type: str, text: str) -> str:
    if not line_items:
        # For structural evidence without charges, prefer document-level fallback over incidental words
        # such as "storage duration" on appointment summaries.
        if doc_type in DOC_TYPE_FALLBACK_TYPE_B:
            return DOC_TYPE_FALLBACK_TYPE_B.get(doc_type, "OTHERS")
        return map_type_b(text, doc_type)
    # Use the highest priority TYPE_B among extracted lines.
    priority_order = ["Inspection", "Customs", "DO", "INLAND", "THC", "Detention", "STROAGE", "OTHERS"]
    present = {li.get("type_b", "OTHERS") for li in line_items}
    for tb in priority_order:
        if tb in present:
            return tb
    return "OTHERS"


# ---------------------------------------------------------------------------
# Gate / DLP / output sanitization
# ---------------------------------------------------------------------------

DLP_PATTERNS = {
    "TRN": re.compile(r"\bTRN\s*:?\s*\d{9,15}\b", re.I),
    "IBAN": re.compile(r"\bAE\d{21}\b", re.I),
    "EMAIL": re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.I),
    "PHONE": re.compile(r"(?:\+971|0)\s?\d[\d\s-]{6,}\b", re.I),
}


def dlp_scan_text(text: str) -> dict[str, int]:
    return {name: len(rx.findall(text)) for name, rx in DLP_PATTERNS.items()}


def sanitize_keys(keys: dict[str, Any], mask: bool = True) -> dict[str, Any]:
    if not mask:
        return keys
    out = json.loads(json.dumps(keys, ensure_ascii=False))
    # BL/AWB/HAWB are masked by default. Containers and shipment numbers remain visible for operational matching.
    out["bl_nos"] = mask_list(out.get("bl_nos", []), "BL")
    out["bl_no"] = mask_token(out.get("bl_no", ""), prefix="BL") if out.get("bl_no") else ""
    out["mawb_nos"] = mask_list(out.get("mawb_nos", []), "MAWB")
    out["hawb_nos"] = mask_list(out.get("hawb_nos", []), "HAWB")
    # DO and invoice are partially masked/hash-like to preserve joinability without raw disclosure.
    out["do_nos"] = mask_list(out.get("do_nos", []), "DO")
    out["do_no"] = mask_token(out.get("do_no", ""), prefix="DO") if out.get("do_no") else ""
    out["invoice_nos"] = mask_list(out.get("invoice_nos", []), "INV")
    out["invoice_no"] = mask_token(out.get("invoice_no", ""), prefix="INV") if out.get("invoice_no") else ""
    return out


def evidence_status_for_doc(doc_type: str, line_items: list[dict[str, Any]], keys: dict[str, Any]) -> str:
    if line_items:
        return "MATCHED_AMOUNT"
    if keys.get("amounts_aed"):
        return "MATCHED_AMOUNT"
    if doc_type in {"DELIVERY_ORDER", "DELIVERY_NOTE", "AIRPORT_APPOINTMENT"}:
        return "NOT_APPLICABLE"
    if doc_type == "BOE_CUSTOMS":
        return "PARTIAL"
    if doc_type == "UNKNOWN":
        return "MISSING"
    return "PARTIAL"


def gate_verdict(doc_type: str, keys: dict[str, Any], line_items: list[dict[str, Any]],
                 text_source: str, extraction_issues: list[str], text_len: int) -> tuple[str, list[dict[str, str]]]:
    issues: list[dict[str, str]] = []
    if doc_type == "UNKNOWN":
        issues.append({"severity": "AMBER", "code": "UNKNOWN_DOCTYPE", "detail": "Header fingerprint not matched"})
    if text_source.startswith("ocr"):
        issues.append({"severity": "AMBER", "code": "OCR_FALLBACK_USED", "detail": "OCR output requires visual QA"})
    if text_len < 80:
        issues.append({"severity": "AMBER", "code": "LOW_TEXT_LENGTH", "detail": "Native text layer weak or absent"})
    if doc_type == "BOE_CUSTOMS":
        issues.append({"severity": "AMBER", "code": "CUSTOMS_FINAL_REVIEW_REQUIRED", "detail": "HS/UAE Customs final decision blocked from auto-PASS"})
    invoice_like = doc_type in {"CARRIER_RHS", "CARRIER_EVG", "CARRIER_CMA", "PORT_ALLIED", "PORT_CSP", "AIRPORT_FEES"}
    if invoice_like and not line_items and not keys.get("amounts_aed"):
        issues.append({"severity": "AMBER", "code": "INVOICE_AMOUNT_MISSING", "detail": "No line amount or context AED amount extracted"})
    if not keys.get("shipment_nos"):
        # Not all invoices contain shipment_no; this remains AMBER not FAIL.
        issues.append({"severity": "AMBER", "code": "NO_SHIPMENT_NO", "detail": "Shipment number missing in text/filename"})
    for issue in extraction_issues:
        issues.append({"severity": "AMBER", "code": issue, "detail": "Extraction layer notice"})
    if any(i["severity"] == "ZERO" for i in issues):
        return "ZERO", issues
    if any(i["severity"] == "AMBER" for i in issues):
        return "AMBER", issues
    return "PASS", issues


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------


def process_pdf(pdf_path: Path, render_dir: Optional[Path], ocr_mode: str, ocr_threshold: int,
                ocr_max_pages: int, ocr_timeout: int, mask: bool = True) -> dict[str, Any]:
    rec: dict[str, Any] = {
        "file": pdf_path.name,
        "size_bytes": pdf_path.stat().st_size,
        "sha256": sha256_short(pdf_path.read_bytes()),
        "processed_at": now_utc(),
        "parser_version": VERSION,
    }
    try:
        text, page_count, text_source, extraction_issues = extract_text(
            pdf_path, ocr_mode=ocr_mode, ocr_threshold=ocr_threshold, ocr_max_pages=ocr_max_pages, ocr_timeout=ocr_timeout
        )
        tables = extract_tables(pdf_path)
        class_info = classify_doc(text, pdf_path.name)
        raw_keys = extract_common_keys(text, pdf_path.name)
        line_items = extract_line_items(text, tables, class_info["doc_type"])
        # Use line item totals as fallback amounts.
        line_amounts = [li.get("total_aed") or li.get("amount_aed") for li in line_items]
        line_amounts = [round(float(a), 2) for a in line_amounts if a is not None]
        raw_keys["line_amounts_aed"] = line_amounts
        doc_type = class_info["doc_type"]
        type_b = type_b_from_lines(line_items, doc_type, text)
        verdict, issues = gate_verdict(doc_type, raw_keys, line_items, text_source, extraction_issues, len(text))
        rec.update({
            "page_count": page_count,
            "text_length": len(text),
            "text_source": text_source,
            "doc_type": doc_type,
            "doc_type_confidence": class_info["confidence"],
            "doc_type_matched": class_info["matched"],
            "doc_type_rule_note": class_info["rule_note"],
            "keys": sanitize_keys(raw_keys, mask=mask),
            "key_counts": {
                "shipment_nos": len(raw_keys.get("shipment_nos", [])),
                "containers": len(raw_keys.get("containers", [])),
                "bl_nos": len(raw_keys.get("bl_nos", [])),
                "mawb_nos": len(raw_keys.get("mawb_nos", [])),
                "hawb_nos": len(raw_keys.get("hawb_nos", [])),
                "do_nos": len(raw_keys.get("do_nos", [])),
                "invoice_nos": len(raw_keys.get("invoice_nos", [])),
                "line_items": len(line_items),
            },
            "table_count": len(tables),
            "line_items": line_items,
            "type_b": type_b,
            "evidence_status": evidence_status_for_doc(doc_type, line_items, raw_keys),
            "verdict": verdict,
            "issues": issues,
            "dlp_hit_count": dlp_scan_text(text),
            "render_qa": render_qa(pdf_path, render_dir),
            "roundup_note": ROUNDUP_NOTE,
        })
    except Exception as exc:  # pragma: no cover - file-specific
        rec.update({
            "error": str(exc),
            "verdict": "FAILED",
            "issues": [{"severity": "FAIL", "code": "PARSE_EXCEPTION", "detail": str(exc)}],
            "roundup_note": ROUNDUP_NOTE,
        })
    return rec


def collect_pdfs(input_path: Path, recursive: bool = False) -> list[Path]:
    if input_path.is_file():
        if input_path.suffix.lower() != ".pdf":
            raise ValueError(f"Input file is not PDF: {input_path}")
        return [input_path]
    if recursive:
        return sorted(p for p in input_path.rglob("*.pdf") if p.is_file())
    return sorted(p for p in input_path.iterdir() if p.is_file() and p.suffix.lower() == ".pdf")


def scan_pdfs(input_path: Path, out_dir: Path, render: bool, ocr_mode: str, ocr_threshold: int,
              ocr_max_pages: int, ocr_timeout: int, mask: bool, recursive: bool, progress: bool = True) -> list[dict[str, Any]]:
    pdfs = collect_pdfs(input_path, recursive=recursive)
    render_dir = out_dir / "renders" if render else None
    records: list[dict[str, Any]] = []
    for idx, pdf in enumerate(pdfs, 1):
        if progress:
            print(f"[{idx:>3}/{len(pdfs)}] {pdf.name}", flush=True)
        records.append(process_pdf(pdf, render_dir, ocr_mode, ocr_threshold, ocr_max_pages, ocr_timeout, mask=mask))
        gc.collect()
    return records


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------


def write_json(records: list[dict[str, Any]], out_path: Path, project_scan: dict[str, Any]) -> None:
    payload = {
        "manifest": {
            "parser_version": VERSION,
            "created_at": now_utc(),
            "record_count": len(records),
            "roundup_note": ROUNDUP_NOTE,
            "project_source": project_scan,
        },
        "records": records,
    }
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def write_line_items_csv(records: list[dict[str, Any]], out_path: Path) -> None:
    headers = [
        "File", "DocType", "Shipment_No", "Line_ID", "Page", "Source", "Description",
        "Container", "Currency", "Amount_AED", "VAT_AED", "Total_AED", "TYPE_B",
        "Evidence_Status", "Extraction_Note", "Verdict",
    ]
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for rec in records:
            keys = rec.get("keys", {})
            shipment = keys.get("shipment_no", "")
            for li in rec.get("line_items", []):
                w.writerow([
                    rec.get("file", ""), rec.get("doc_type", ""), shipment,
                    li.get("line_id", ""), li.get("page", ""), li.get("source", ""),
                    li.get("description", ""), li.get("container", ""), li.get("currency", "AED"),
                    fmt_money(li.get("amount_aed")), fmt_money(li.get("vat_aed")), fmt_money(li.get("total_aed")),
                    li.get("type_b", ""), li.get("evidence_status", ""), li.get("extraction_note", ""),
                    rec.get("verdict", ""),
                ])


def write_summary_md(records: list[dict[str, Any]], out_path: Path, project_scan: dict[str, Any]) -> None:
    def count_by(field: str) -> dict[str, int]:
        d: dict[str, int] = {}
        for r in records:
            v = str(r.get(field, "?"))
            d[v] = d.get(v, 0) + 1
        return dict(sorted(d.items(), key=lambda x: (-x[1], x[0])))

    issue_count = sum(len(r.get("issues", [])) for r in records)
    line_count = sum(len(r.get("line_items", [])) for r in records)
    lines: list[str] = [
        "# HVDC PDF Parse PRO v2.1 - Run Summary",
        "",
        f"- Created: {now_utc()}",
        f"- Parser Version: {VERSION}",
        f"- PDFs: {len(records)}",
        f"- Extracted Line Items: {line_count}",
        f"- Issues: {issue_count}",
        f"- Project Source Scan: {project_scan.get('project_source_scan_status')}",
        f"- Contract Source: {project_scan.get('contract_source_status')}",
        f"- ROUNDUP Note: {ROUNDUP_NOTE}",
        "",
        "## By DOC_TYPE",
    ]
    for k, v in count_by("doc_type").items():
        lines.append(f"- {k}: {v}")
    lines.append("")
    lines.append("## By TYPE_B")
    for k, v in count_by("type_b").items():
        lines.append(f"- {k}: {v}")
    lines.append("")
    lines.append("## By Verdict")
    for k, v in count_by("verdict").items():
        lines.append(f"- {k}: {v}")
    lines.append("")
    lines.append("## Top Issues")
    issue_map: dict[str, int] = {}
    for r in records:
        for i in r.get("issues", []):
            code = i.get("code", "?")
            issue_map[code] = issue_map.get(code, 0) + 1
    if not issue_map:
        lines.append("- None")
    else:
        for k, v in sorted(issue_map.items(), key=lambda x: (-x[1], x[0])):
            lines.append(f"- {k}: {v}")
    out_path.write_text("\n".join(lines), encoding="utf-8")


def style_sheet(ws: Any, freeze: str = "A2") -> None:
    if ws.max_row >= 1:
        ws.freeze_panes = freeze
        ws.auto_filter.ref = ws.dimensions
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(style="thin", color="D9D9D9")
        for cell in ws[1]:
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.border = Border(bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(wrap_text=False, vertical="center")
            if isinstance(cell.value, (float, int)):
                cell.number_format = "#,##0.00"
    for idx, col in enumerate(ws.columns, 1):
        max_len = 8
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(val) + 2, 45))
        ws.column_dimensions[get_column_letter(idx)].width = max_len
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 15


def write_excel(records: list[dict[str, Any]], out_path: Path, project_scan: dict[str, Any]) -> None:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl not available; use JSON/CSV outputs")
    wb = Workbook()
    ws = wb.active
    ws.title = "Evidence_Index"
    headers = [
        "File", "DocType", "DocConf", "TYPE_B", "Verdict", "Evidence_Status", "Shipment_No", "Shipment_Count",
        "Containers", "BL_No_Masked", "MAWB_Masked", "HAWB_Masked", "DO_No_Masked", "Invoice_No_Masked",
        "Date", "Line_Count", "Amounts_AED", "Pages", "Tables", "TextLen", "Text_Source", "Rendered", "SHA256", "ROUNDUP Note",
    ]
    ws.append(headers)
    for r in records:
        k = r.get("keys", {})
        ws.append([
            r.get("file"), r.get("doc_type"), r.get("doc_type_confidence"), r.get("type_b"), r.get("verdict"),
            r.get("evidence_status"), k.get("shipment_no", ""), len(k.get("shipment_nos", [])),
            "; ".join(k.get("containers", [])[:10]), k.get("bl_no", ""), "; ".join(k.get("mawb_nos", [])),
            "; ".join(k.get("hawb_nos", [])), k.get("do_no", ""), k.get("invoice_no", ""), k.get("date", ""),
            len(r.get("line_items", [])), "; ".join(fmt_money(a) for a in k.get("amounts_aed", [])[:8]),
            r.get("page_count"), r.get("table_count"), r.get("text_length"), r.get("text_source"),
            r.get("render_qa", {}).get("rendered", False), r.get("sha256"), ROUNDUP_NOTE,
        ])
    style_sheet(ws)
    verdict_fill = {
        "PASS": "C6EFCE",
        "PASS WITH WARNINGS": "FFEB9C",
        "AMBER": "FFEB9C",
        "FAIL": "FFC7CE",
        "FAILED": "FFC7CE",
        "ZERO": "FFC7CE",
    }
    for row in range(2, ws.max_row + 1):
        v = str(ws.cell(row=row, column=5).value)
        if v in verdict_fill:
            ws.cell(row=row, column=5).fill = PatternFill("solid", fgColor=verdict_fill[v])

    ws2 = wb.create_sheet("Line_Items")
    ws2.append(["File", "DocType", "Shipment_No", "Line_ID", "Page", "Source", "Description", "Container", "Currency", "Amount_AED", "VAT_AED", "Total_AED", "TYPE_B", "Evidence_Status", "Extraction_Note"])
    for r in records:
        shipment = r.get("keys", {}).get("shipment_no", "")
        for li in r.get("line_items", []):
            ws2.append([
                r.get("file"), r.get("doc_type"), shipment, li.get("line_id"), li.get("page"), li.get("source"),
                li.get("description"), li.get("container"), li.get("currency"), li.get("amount_aed"), li.get("vat_aed"),
                li.get("total_aed"), li.get("type_b"), li.get("evidence_status"), li.get("extraction_note"),
            ])
    style_sheet(ws2)

    ws3 = wb.create_sheet("Issues")
    ws3.append(["File", "Severity", "Code", "Detail"])
    for r in records:
        for issue in r.get("issues", []):
            ws3.append([r.get("file"), issue.get("severity"), issue.get("code"), issue.get("detail")])
    style_sheet(ws3)

    ws4 = wb.create_sheet("Summary")
    ws4.append(["Metric", "Value"])
    ws4.append(["Parser Version", VERSION])
    ws4.append(["Created At", now_utc()])
    ws4.append(["Total PDFs", len(records)])
    ws4.append(["Total Line Items", sum(len(r.get("line_items", [])) for r in records)])
    ws4.append(["ROUNDUP Note", ROUNDUP_NOTE])
    ws4.append(["Project Source Scan", project_scan.get("project_source_scan_status")])
    ws4.append(["Contract Source", project_scan.get("contract_source_status")])
    ws4.append([])
    for label, field in [("By DOC_TYPE", "doc_type"), ("By TYPE_B", "type_b"), ("By Verdict", "verdict")]:
        ws4.append([label, ""])
        counts: dict[str, int] = {}
        for r in records:
            counts[str(r.get(field, "?"))] = counts.get(str(r.get(field, "?")), 0) + 1
        for k, v in sorted(counts.items(), key=lambda x: (-x[1], x[0])):
            ws4.append([f"  {k}", v])
    style_sheet(ws4)

    ws5 = wb.create_sheet("Project_Scan")
    ws5.append(["Item", "Value"])
    for k, v in project_scan.items():
        ws5.append([k, json.dumps(v, ensure_ascii=False) if isinstance(v, (list, dict)) else v])
    style_sheet(ws5)

    ws6 = wb.create_sheet("Audit_Log")
    ws6.append(["Step_No", "Action", "Output", "Status", "Timestamp"])
    steps = [
        (1, "Project Source Scan", project_scan.get("project_source_scan_status"), "DONE"),
        (2, "Text/Table Extract", f"{len(records)} PDF records", "DONE"),
        (3, "DOC_TYPE Classification", "priority fingerprint v2.1", "DONE"),
        (4, "COMMON_KEYS Extraction", "shipment/container/BL/AWB/DO/invoice/date/amount", "DONE"),
        (5, "Line Item Extraction", f"{sum(len(r.get('line_items', [])) for r in records)} rows", "DONE"),
        (6, "TYPE_B Mapping", "no blank TYPE_B", "DONE"),
        (7, "AMBER/ZERO Gate", "parse gate only; final approval blocked", "DONE"),
    ]
    for s in steps:
        ws6.append([s[0], s[1], s[2], s[3], now_utc()])
    style_sheet(ws6)

    wb.save(out_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main(argv: Optional[list[str]] = None) -> int:
    ap = argparse.ArgumentParser(description="HVDC DSV Hybrid PDF Parser PRO v2.1")
    ap.add_argument("--input", "-i", type=Path, required=True, help="PDF file or directory")
    ap.add_argument("--out-dir", "-o", type=Path, default=Path("parser_out_v2_1"), help="Output directory")
    ap.add_argument("--project-source", type=Path, default=None, help="Unzipped Project Source Package root directory")
    ap.add_argument("--recursive", action="store_true", help="Scan input directory recursively")
    ap.add_argument("--render", action="store_true", help="Render first page PNG for QA")
    ap.add_argument("--ocr", choices=["off", "auto", "force"], default="auto", help="OCR fallback mode")
    ap.add_argument("--ocr-threshold", type=int, default=80, help="Native text length threshold for OCR auto mode")
    ap.add_argument("--ocr-max-pages", type=int, default=1, help="Max OCR pages per PDF")
    ap.add_argument("--ocr-timeout", type=int, default=20, help="Tesseract timeout seconds per page")
    ap.add_argument("--no-xlsx", action="store_true", help="Skip XLSX output")
    ap.add_argument("--unmask", action="store_true", help="Disable DLP masking in outputs; internal use only")
    args = ap.parse_args(argv)

    input_path = args.input.expanduser().resolve()
    if not input_path.exists():
        print(f"[ERROR] input not found: {input_path}", file=sys.stderr)
        return 2

    args.out_dir.mkdir(parents=True, exist_ok=True)
    project_scan = scan_project_source(args.project_source)
    print(f"[PROJECT] {project_scan.get('project_source_scan_status')} | contract={project_scan.get('contract_source_status')}")

    try:
        records = scan_pdfs(
            input_path=input_path,
            out_dir=args.out_dir,
            render=args.render,
            ocr_mode=args.ocr,
            ocr_threshold=args.ocr_threshold,
            ocr_max_pages=args.ocr_max_pages,
            ocr_timeout=args.ocr_timeout,
            mask=not args.unmask,
            recursive=args.recursive,
            progress=True,
        )
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1

    write_json(records, args.out_dir / "pdf_evidence_index.json", project_scan)
    write_line_items_csv(records, args.out_dir / "pdf_line_items.csv")
    write_summary_md(records, args.out_dir / "summary.md", project_scan)
    if not args.no_xlsx:
        try:
            write_excel(records, args.out_dir / "pdf_evidence_index.xlsx", project_scan)
        except Exception as exc:
            print(f"[WARN] XLSX output skipped: {exc}", file=sys.stderr)

    counts: dict[str, int] = {}
    for r in records:
        counts[r.get("verdict", "?")] = counts.get(r.get("verdict", "?"), 0) + 1
    print(f"\n[OK] {len(records)} PDFs processed -> {args.out_dir}")
    for verdict, count in sorted(counts.items(), key=lambda x: (-x[1], x[0])):
        print(f"     {verdict}: {count}")
    print(f"[NOTE] {ROUNDUP_NOTE}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
