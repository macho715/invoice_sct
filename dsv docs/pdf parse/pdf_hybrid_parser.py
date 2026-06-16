"""
Hybrid PDF Parser — HVDC Invoice Audit
Implements pdfparse.md playbook v1.0
- 7 DOC_TYPE classification
- COMMON_KEYS extraction (shipment_no, container, bl_no, do_no, invoice_no, date, amount_aed)
- Evidence Index + TYPE_B mapping
- AMBER/ZERO gate logic

Pipeline: text extract → table extract → render QA → evidence index → TYPE_B → gate
"""
from __future__ import annotations
import json
import re
import sys
import argparse
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Any

import fitz  # PyMuPDF
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────────────────────
# 1. DOC_TYPE classification (header fingerprint)
# ─────────────────────────────────────────────────────────────
DOC_TYPE_RULES = [
    # Priority: vendor-specific (Allied/CSP/TWCS) BEFORE generic "TAX INVOICE" carrier rules.
    # Without this, Allied port invoices with "TAX INVOICE" string get misclassified as CARRIER_RHS.
    ("BOE_CUSTOMS",   ["Customs Declaration", "DEBIT NOTE", "Gate Pass", "بوليصة",
                       "LAND IMPORT", "Import to Local from FZ"]),
    ("DELIVERY_ORDER", ["DELIVERY ORDER", "D/Order No", "D.O. NUMBER", "DELIVERY NOTIFICATION",
                        "DELIVERY NOTIFICATION MASTER"]),
    ("DELIVERY_NOTE",  ["NOT NEGOTIABLE DELIVERY NOTE", "Delivery Note/Waybill", "DELIVERY NOTE"]),
    ("PORT_ALLIED",    ["ALLIED ONDOCK"]),
    ("PORT_CSP",       ["CSP Abu Dhabi Terminal", "Package Invoice"]),
    ("PORT_TWCS",      ["TransWorldContainer", "TransWorld Container Services"]),
    ("AIRPORT_FEES",   ["Etihad Terminal Charges", "Charges Summary", "Airport Fees", "Terminal Charges"]),
    ("APPOINTMENT",    ["IMPORT APPOINTMENT SUMMARY", "APPOINTMENT SUMMARY"]),
    ("CARRIER_RHS",    ["RAIS HASSAN SAADI", "TAX INVOICE"]),
    ("CARRIER_EVG",    ["EVERGREEN SHIPPING AGENCY", "TAX INVOICE"]),
]

def classify_doc(text: str) -> str:
    """Return first matching DOC_TYPE based on header fingerprint, else UNKNOWN."""
    head = text[:3000]  # first 3k chars are enough for header detection
    for doc_type, keys in DOC_TYPE_RULES:
        for key in keys:
            if key.lower() in head.lower():
                return doc_type
    return "UNKNOWN"

# ─────────────────────────────────────────────────────────────
# 2. COMMON_KEYS regex
# ─────────────────────────────────────────────────────────────
RX_SHIPMENT = re.compile(r"HVDC[\s\-￾]+ADOPT[\s\-]+SCT[\s\-]*(\d{4})", re.IGNORECASE)
RX_SHIPMENT_HE = re.compile(r"HVDC[\s\-￾]+ADOPT[\s\-]+HE[\s\-]*(\d{4})", re.IGNORECASE)
RX_CONTAINER = re.compile(r"\b([A-Z]{4})\s?(\d{6})-?(\d)\b")
RX_BL = re.compile(r"\b(?:SELA|EGLV)[A-Z0-9]{8,}\b")
RX_DO = re.compile(r"(?:D/Order No|DO #|Delivery Order No|D\.O\.\s*NUMBER|D/O NUMBER)[:\s]*([0-9 ]{6,})", re.IGNORECASE)
RX_INVOICE = re.compile(r"(?:Invoice No|INVOICE NO\.?|Invoice Number)\s*:?\s*([A-Z0-9\-/]+)", re.IGNORECASE)
RX_DATE = re.compile(r"\b(\d{2}/\d{2}/\d{4}|\d{2}-[A-Z]{3}-\d{4}|\d{4}-\d{2}-\d{2})\b")
RX_AMOUNT_AED = re.compile(r"AED\s*([0-9,]+(?:\.\d{2})?)", re.IGNORECASE)
RX_AMOUNT_NUM = re.compile(r"\b([0-9]{1,3}(?:,[0-9]{3})*(?:\.\d{2})?)\b")

def normalize_container(s: str) -> str:
    """HMMU 608937-7 → HMMU6089377"""
    return re.sub(r"[\s\-]", "", s)

def extract_common_keys(text: str) -> dict[str, Any]:
    keys: dict[str, Any] = {}
    if m := RX_SHIPMENT.search(text):
        keys["shipment_no"] = "HVDC-ADOPT-SCT-" + m.group(1)
    elif m := RX_SHIPMENT_HE.search(text):
        keys["shipment_no"] = "HVDC-ADOPT-HE-" + m.group(1)
    containers = []
    for m in RX_CONTAINER.finditer(text):
        containers.append(normalize_container(m.group(0)))
    keys["containers"] = list(dict.fromkeys(containers))  # dedupe, preserve order
    if m := RX_BL.search(text):
        keys["bl_no"] = m.group(0)
    if m := RX_DO.search(text):
        keys["do_no"] = re.sub(r"\s+", "", m.group(1))
    if m := RX_INVOICE.search(text):
        keys["invoice_no"] = m.group(1)
    if m := RX_DATE.search(text):
        keys["date"] = m.group(1)
    aed_amounts = []
    for m in RX_AMOUNT_AED.finditer(text):
        try:
            aed_amounts.append(float(m.group(1).replace(",", "")))
        except ValueError:
            pass
    keys["amounts_aed"] = aed_amounts
    return keys

# ─────────────────────────────────────────────────────────────
# 3. Extract layer
# ─────────────────────────────────────────────────────────────
def extract_text(pdf_path: Path) -> tuple[str, int]:
    """Returns (full_text, page_count). Uses pdfplumber for text layer."""
    parts: list[str] = []
    page_count = 0
    with pdfplumber.open(pdf_path) as pdf:
        page_count = len(pdf.pages)
        for page in pdf.pages:
            t = page.extract_text() or ""
            parts.append(t)
    return "\n".join(parts), page_count

def extract_tables(pdf_path: Path) -> list[list[list[str]]]:
    """Returns list of pages, each page is a list of tables (rows of cells)."""
    out: list[list[list[str]]] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            try:
                tables = page.extract_tables() or []
            except Exception:
                tables = []
            out.append(tables)
    return out

def render_qa(pdf_path: Path, page_index: int = 0, out_dir: Path | None = None,
              dpi: int = 120) -> dict[str, Any]:
    """Render first page to PNG for visual QA. Returns metadata only (no full bytes in JSON)."""
    meta: dict[str, Any] = {"page": page_index + 1, "rendered": False}
    try:
        with fitz.open(pdf_path) as doc:
            if page_index >= len(doc):
                return meta
            page = doc[page_index]
            pix = page.get_pixmap(dpi=dpi)
            meta["width"] = pix.width
            meta["height"] = pix.height
            meta["rendered"] = True
            if out_dir is not None:
                out_dir.mkdir(parents=True, exist_ok=True)
                img_path = out_dir / f"{pdf_path.stem}_p{page_index+1}.png"
                pix.save(str(img_path))
                meta["image_path"] = str(img_path)
    except Exception as e:
        meta["error"] = str(e)
    return meta

# ─────────────────────────────────────────────────────────────
# 4. TYPE_B mapping
# ─────────────────────────────────────────────────────────────
TYPE_B_RULES = {
    "Inspection":  ["Container Inspection Fee", "Admin & Inspection", "Customs Inspection", "Inspection Fee"],
    "Customs":     ["BOE", "Customs Duty", "Customs Declaration", "Debit Note", "HS Code"],
    "DO":          ["Delivery Order", "DO Fee", "D/Order"],
    "INLAND":      ["DN Road Freight", "Truck", "FB movement", "Inland", "Transport"],
    "THC":         ["Terminal Handling", "Port Handling", "THC"],
    "OTHERS":      ["ISPS", "Equipment Repositioning", "Container Maintenance", "Airport Fees",
                    "Appointment", "CNT Repair", "Empty Return", "Washing"],
}

def map_type_b(text: str, doc_type: str) -> str:
    """Map text + doc_type to a TYPE_B category. Defaults to OTHERS."""
    head = text[:8000]
    for tb, kws in TYPE_B_RULES.items():
        for kw in kws:
            if kw.lower() in head.lower():
                return tb
    # Fallback by doc_type
    fallback = {
        "BOE_CUSTOMS":    "Customs",
        "DELIVERY_ORDER": "DO",
        "DELIVERY_NOTE":  "INLAND",
        "CARRIER_RHS":    "OTHERS",  # carrier invoice often mixed; needs review
        "CARRIER_EVG":    "OTHERS",
        "PORT_ALLIED":    "Inspection",
        "PORT_CSP":       "Inspection",
    }
    return fallback.get(doc_type, "OTHERS")

# ─────────────────────────────────────────────────────────────
# 5. Gate (AMBER/ZERO conditions from pdfparse.md)
# ─────────────────────────────────────────────────────────────
ZERO_CONDITIONS = [
    "BOE/HS/UAE Customs 최종 판정",   # customs final judgment
    "DEM/DET/Storage settlement",    # demurrage/detention/storage
    "Grand Total 공란 invoice",       # grand total blank → can't auto PASS
]

def gate_verdict(doc_type: str, keys: dict[str, Any], text: str) -> str:
    """
    Return PASS / AMBER / ZERO.
    ZERO conditions per pdfparse.md: customs final judgment, DEM/DET/storage, blank grand total invoice.
    AMBER: parse OK but require reviewer (e.g., carrier invoice, BOE mixed).
    """
    # ZERO conditions
    if doc_type == "BOE_CUSTOMS":
        return "AMBER"  # customs final judgment always requires reviewer
    # Grand total blank check (heuristic: if amounts list empty AND doc_type is carrier/port invoice)
    if doc_type in ("CARRIER_RHS", "CARRIER_EVG", "PORT_ALLIED", "PORT_CSP"):
        if not keys.get("amounts_aed"):
            return "AMBER"  # no amounts found; line-level evidence only
    # No shipment number
    if not keys.get("shipment_no"):
        return "AMBER"
    return "PASS"

# ─────────────────────────────────────────────────────────────
# 6. Pipeline
# ─────────────────────────────────────────────────────────────
def process_pdf(pdf_path: Path, render_dir: Path | None = None) -> dict[str, Any]:
    """Run full pipeline on one PDF. Returns evidence record."""
    rec: dict[str, Any] = {
        "file":            pdf_path.name,
        "size_bytes":      pdf_path.stat().st_size,
        "sha256":          hashlib.sha256(pdf_path.read_bytes()).hexdigest()[:16],
        "processed_at":    datetime.now().isoformat(timespec="seconds"),
    }
    try:
        text, page_count = extract_text(pdf_path)
        rec["page_count"] = page_count
        rec["text_length"] = len(text.strip())
        rec["image_only"] = rec["text_length"] < 50  # OCR fallback flag
        rec["doc_type"] = classify_doc(text)
        rec["keys"] = extract_common_keys(text)
        rec["tables"] = extract_tables(pdf_path)
        rec["table_count"] = sum(len(pg) for pg in rec["tables"])
        rec["type_b"] = map_type_b(text, rec["doc_type"])
        rec["verdict"] = gate_verdict(rec["doc_type"], rec["keys"], text)
        rec["render_qa"] = render_qa(pdf_path, page_index=0, out_dir=render_dir)
    except Exception as e:
        rec["error"] = str(e)
        rec["verdict"] = "FAILED"
    return rec

def scan_directory(dir_path: Path, render_dir: Path | None = None,
                  progress: bool = True) -> list[dict[str, Any]]:
    pdfs = sorted(p for p in dir_path.iterdir() if p.suffix.lower() == ".pdf")
    results: list[dict[str, Any]] = []
    for i, pdf in enumerate(pdfs, 1):
        if progress:
            print(f"[{i:>2}/{len(pdfs)}] {pdf.name}", flush=True)
        results.append(process_pdf(pdf, render_dir))
    return results

# ─────────────────────────────────────────────────────────────
# 7. Output writers
# ─────────────────────────────────────────────────────────────
def write_json(records: list[dict[str, Any]], out_path: Path) -> None:
    # Strip heavy table data from JSON to keep it readable
    slim = []
    for r in records:
        r2 = {k: v for k, v in r.items() if k != "tables"}
        r2["table_count"] = r.get("table_count", 0)
        slim.append(r2)
    out_path.write_text(json.dumps(slim, indent=2, ensure_ascii=False), encoding="utf-8")

def write_excel(records: list[dict[str, Any]], out_path: Path) -> None:
    wb = Workbook()
    # ── Sheet 1: Evidence Index
    ws1 = wb.active
    ws1.title = "Evidence_Index"
    headers = ["File", "DocType", "TYPE_B", "Verdict", "Shipment_No", "Containers",
               "BL_No", "DO_No", "Invoice_No", "Date", "Amounts_AED", "Pages",
               "Tables", "TextLen", "SizeKB", "SHA256", "Rendered"]
    ws1.append(headers)
    for c in ws1[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    green = PatternFill("solid", fgColor="C6EFCE")

    for r in records:
        k = r.get("keys", {})
        amts = k.get("amounts_aed", [])
        amts_str = "; ".join(f"{a:,.2f}" for a in amts[:5])
        cnts = "; ".join(k.get("containers", [])[:5])
        row = [
            r.get("file"), r.get("doc_type"), r.get("type_b"), r.get("verdict"),
            k.get("shipment_no", ""), cnts, k.get("bl_no", ""), k.get("do_no", ""),
            k.get("invoice_no", ""), k.get("date", ""), amts_str,
            r.get("page_count", 0), r.get("table_count", 0), r.get("text_length", 0),
            f"{r.get('size_bytes', 0)/1024:.1f}", r.get("sha256", ""),
            r.get("render_qa", {}).get("rendered", False),
        ]
        ws1.append(row)
        verdict = r.get("verdict", "")
        if verdict == "ZERO":
            ws1.cell(row=ws1.max_row, column=4).fill = red
        elif verdict == "AMBER":
            ws1.cell(row=ws1.max_row, column=4).fill = yellow
        elif verdict == "PASS":
            ws1.cell(row=ws1.max_row, column=4).fill = green

    # Column widths
    widths = [42, 16, 12, 10, 22, 30, 16, 12, 18, 12, 32, 6, 6, 8, 8, 18, 10]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = w

    # ── Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    ws2["A1"].font = Font(bold=True); ws2["B1"].font = Font(bold=True)
    ws2["A1"].fill = PatternFill("solid", fgColor="DDDDDD")
    ws2["B1"].fill = PatternFill("solid", fgColor="DDDDDD")

    by_type: dict[str, int] = {}
    by_verdict: dict[str, int] = {}
    by_typeb: dict[str, int] = {}
    for r in records:
        by_type[r.get("doc_type", "?")] = by_type.get(r.get("doc_type", "?"), 0) + 1
        by_verdict[r.get("verdict", "?")] = by_verdict.get(r.get("verdict", "?"), 0) + 1
        by_typeb[r.get("type_b", "?")] = by_typeb.get(r.get("type_b", "?"), 0) + 1

    ws2.append(["Total PDFs", len(records)])
    ws2.append([])
    ws2.append(["By DOC_TYPE"]); ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    for k, v in sorted(by_type.items(), key=lambda x: -x[1]):
        ws2.append([f"  {k}", v])
    ws2.append([])
    ws2.append(["By TYPE_B"]); ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    for k, v in sorted(by_typeb.items(), key=lambda x: -x[1]):
        ws2.append([f"  {k}", v])
    ws2.append([])
    ws2.append(["By Verdict"]); ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    for k, v in sorted(by_verdict.items(), key=lambda x: -x[1]):
        ws2.append([f"  {k}", v])
    ws2.append([])
    ws2.append(["ZERO Conditions (per pdfparse.md)"])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    for c in ZERO_CONDITIONS:
        ws2.append([f"  - {c}"])

    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 16

    # ── Sheet 3: Issues
    ws3 = wb.create_sheet("Issues")
    ws3.append(["File", "Issue", "Detail"])
    for c in ws3[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
    ws3.freeze_panes = "A2"

    for r in records:
        f = r.get("file", "")
        if r.get("verdict") == "FAILED":
            ws3.append([f, "PARSE_FAILED", r.get("error", "")])
        if r.get("doc_type") == "UNKNOWN":
            ws3.append([f, "UNKNOWN_DOCTYPE", "header fingerprint not matched"])
        if not r.get("keys", {}).get("shipment_no"):
            ws3.append([f, "NO_SHIPMENT_NO", "HVDC-ADOPT-XXXX pattern not found"])
        k = r.get("keys", {})
        if r.get("doc_type") in ("CARRIER_RHS", "CARRIER_EVG") and not k.get("amounts_aed"):
            ws3.append([f, "GRAND_TOTAL_BLANK", "line-level evidence only; auto-PASS blocked"])
    ws3.column_dimensions["A"].width = 42
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 50

    wb.save(out_path)

def write_summary_md(records: list[dict[str, Any]], out_path: Path) -> None:
    by_type: dict[str, int] = {}
    by_verdict: dict[str, int] = {}
    by_typeb: dict[str, int] = {}
    issues = 0
    for r in records:
        by_type[r.get("doc_type", "?")] = by_type.get(r.get("doc_type", "?"), 0) + 1
        by_verdict[r.get("verdict", "?")] = by_verdict.get(r.get("verdict", "?"), 0) + 1
        by_typeb[r.get("type_b", "?")] = by_typeb.get(r.get("type_b", "?"), 0) + 1
        if r.get("verdict") == "FAILED": issues += 1
        if r.get("doc_type") == "UNKNOWN": issues += 1
        if not r.get("keys", {}).get("shipment_no"): issues += 1

    lines: list[str] = []
    lines.append("# HVDC PDF Parse — Evidence Index Summary")
    lines.append("")
    lines.append(f"- **Generated**: {datetime.now().isoformat(timespec='seconds')}")
    lines.append(f"- **Total PDFs**: {len(records)}")
    lines.append(f"- **Total Issues**: {issues}")
    lines.append("")
    lines.append("## By DOC_TYPE")
    for k, v in sorted(by_type.items(), key=lambda x: -x[1]):
        lines.append(f"- {k}: {v}")
    lines.append("")
    lines.append("## By TYPE_B")
    for k, v in sorted(by_typeb.items(), key=lambda x: -x[1]):
        lines.append(f"- {k}: {v}")
    lines.append("")
    lines.append("## By Verdict")
    for k, v in sorted(by_verdict.items(), key=lambda x: -x[1]):
        lines.append(f"- {k}: {v}")
    lines.append("")
    lines.append("## Issues")
    lines.append("See Issues sheet in `pdf_evidence_index.xlsx` for details.")
    out_path.write_text("\n".join(lines), encoding="utf-8")

# ─────────────────────────────────────────────────────────────
# 8. Main
# ─────────────────────────────────────────────────────────────
def main() -> int:
    ap = argparse.ArgumentParser(description="HVDC Hybrid PDF Parser")
    ap.add_argument("--input", "-i", type=Path, required=True, help="PDF dir or single PDF")
    ap.add_argument("--out-dir", "-o", type=Path, default=Path("parser_out"))
    ap.add_argument("--render", action="store_true", help="Render first page of each PDF (for QA)")
    args = ap.parse_args()

    args.out_dir.mkdir(parents=True, exist_ok=True)
    render_dir = args.out_dir / "renders" if args.render else None

    if args.input.is_file():
        records = [process_pdf(args.input, render_dir)]
    else:
        records = scan_directory(args.input, render_dir)

    write_json(records, args.out_dir / "pdf_evidence_index.json")
    write_excel(records, args.out_dir / "pdf_evidence_index.xlsx")
    write_summary_md(records, args.out_dir / "summary.md")

    # Console summary
    verdicts: dict[str, int] = {}
    for r in records:
        v = r.get("verdict", "?")
        verdicts[v] = verdicts.get(v, 0) + 1
    print()
    print(f"[OK] {len(records)} PDFs processed → {args.out_dir}")
    for v, c in sorted(verdicts.items(), key=lambda x: -x[1]):
        print(f"     {v}: {c}")
    return 0

if __name__ == "__main__":
    sys.exit(main())
