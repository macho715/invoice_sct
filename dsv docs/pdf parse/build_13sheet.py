"""
HVDC 13-Sheet Audit Workbook Builder
- Source: parser_out_v3/pdf_evidence_index.json
- Contract: 00_Decision ~ 99_Manifest (HVDC workbook spec)
- Rule #0: PASS/AMBER/ZERO/FAILED 모두 Excel 산출
- Number format: 2 decimals, thousand comma. Currency: AED/USD.
- Verdict colors: PASS=green, AMBER=yellow, ZERO/FAILED=red
"""
from __future__ import annotations
import json
from datetime import datetime
from pathlib import Path
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(r"C:\Users\jichu\Downloads\dsv docs")
SRC = ROOT / "parser_out_v3" / "pdf_evidence_index.json"
OUT = ROOT / "20260615_HVDC_13sheet_audit_v1.xlsx"

recs = json.load(open(SRC, encoding="utf-8"))

# ── Styles
BOLD = Font(bold=True, size=10)
TITLE = Font(bold=True, size=14, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

FMT_INT = "#,##0"
FMT_NUM = "#,##0.00"
FMT_AED = '#,##0.00 "AED"'
FMT_USD = '"$"#,##0.00'
FMT_PCT = "0.00%"
FMT_DATE = "yyyy-mm-dd"

def header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = TITLE
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = BORDER
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row+1, column=1)
    ws.auto_filter.ref = ws.dimensions

def style_data_row(ws, row, ncols, verdict_col=None):
    for col in range(1, ncols+1):
        c = ws.cell(row=row, column=col)
        c.alignment = LEFT
        c.border = BORDER
    if verdict_col:
        vc = ws.cell(row=row, column=verdict_col)
        v = (vc.value or "").upper()
        if v == "PASS": vc.fill = PASS_FILL
        elif v == "AMBER": vc.fill = AMBER_FILL
        elif v in ("ZERO","FAILED"): vc.fill = FAIL_FILL

# ─────────────────────────────────────────────────────────────
# 13 Sheets
# ─────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# Aggregate counters
total = len(recs)
verdicts = Counter(r.get("verdict","?") for r in recs)
doc_types = Counter(r.get("doc_type","?") for r in recs)
type_bs = Counter(r.get("type_b","?") for r in recs)
image_only = sum(1 for r in recs if r.get("image_only"))
amt_total_aed = sum(sum(r.get("keys",{}).get("amounts_aed",[])) for r in recs)
amt_per_file = [(r["file"], sum(r.get("keys",{}).get("amounts_aed",[]))) for r in recs]
amt_per_file_sorted = sorted(amt_per_file, key=lambda x: -x[1])

# === 00_Decision
ws = wb.create_sheet("00_Decision")
ws["A1"] = "HVDC Invoice Audit — Top Decision"
ws["A1"].font = Font(bold=True, size=16)
ws.merge_cells("A1:D1")

ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
ws["A2"].font = Font(italic=True, size=9)
ws.merge_cells("A2:D2")

ws["A4"] = "Verdict"; ws["B4"] = "Definition"
ws["A4"].font = BOLD; ws["B4"].font = BOLD
ws["A4"].fill = SUB_FILL; ws["B4"].fill = SUB_FILL

verdict_defs = [
    ("PASS",   "All evidence checks passed. Line-level audit verified.", verdicts.get("PASS", 0)),
    ("AMBER",  "Reviewer approval required. See Issues sheet.", verdicts.get("AMBER", 0)),
    ("ZERO",   "Critical missing data: blank grand total, customs final, DEM/DET. Block auto-PASS.", verdicts.get("ZERO", 0)),
    ("FAILED", "Parser error or unprocessable file. See Issues sheet.", verdicts.get("FAILED", 0)),
]
for i, (v, d, c) in enumerate(verdict_defs, 5):
    ws.cell(row=i, column=1, value=v).font = BOLD
    ws.cell(row=i, column=2, value=d)
    ws.cell(row=i, column=3, value=c).font = BOLD
    ws.cell(row=i, column=4, value=f"=C{i}/{total}").number_format = FMT_PCT
    if v == "PASS": ws.cell(row=i, column=1).fill = PASS_FILL
    elif v == "AMBER": ws.cell(row=i, column=1).fill = AMBER_FILL
    elif v in ("ZERO","FAILED"): ws.cell(row=i, column=1).fill = FAIL_FILL

ws["A10"] = "Headline"; ws["A10"].font = BOLD; ws["A10"].fill = SUB_FILL
ws["B10"] = "Value"; ws["B10"].font = BOLD; ws["B10"].fill = SUB_FILL
metrics = [
    ("Total PDFs",         total),
    ("Total AED amounts",  amt_total_aed),
    ("Image-only (OCR)",   image_only),
    ("Distinct DOC_TYPE",  len(doc_types)),
    ("Distinct TYPE_B",    len(type_bs)),
]
for i, (k, v) in enumerate(metrics, 11):
    ws.cell(row=i, column=1, value=k).font = BOLD
    c = ws.cell(row=i, column=2, value=v)
    if isinstance(v, float):
        c.number_format = FMT_AED if k.startswith("Total") else FMT_NUM
    else:
        c.number_format = FMT_INT
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12

# Top 5 amounts
ws["A18"] = "Top 5 amounts (AED)"; ws["A18"].font = BOLD
ws["A19"] = "File"; ws["B19"] = "Amount AED"; ws["C19"] = "% of total"
for c in ("A19","B19","C19"):
    ws[c].font = BOLD; ws[c].fill = SUB_FILL
for i, (f, a) in enumerate(amt_per_file_sorted[:5], 20):
    ws.cell(row=i, column=1, value=f)
    ws.cell(row=i, column=2, value=a).number_format = FMT_AED
    ws.cell(row=i, column=3, value=f"=B{i}/B12").number_format = FMT_PCT

# === 01_Invoice_Source
ws = wb.create_sheet("01_Invoice_Source")
header_row(ws, 1, ["#", "File", "Designation", "Reason", "Verdict"],
           widths=[5, 48, 16, 60, 10])
for i, r in enumerate(recs, 2):
    k = r.get("keys", {})
    has_amounts = bool(k.get("amounts_aed"))
    has_invoice_no = bool(k.get("invoice_no"))
    dt = r.get("doc_type", "")
    if dt in ("CARRIER_RHS","CARRIER_EVG","PORT_ALLIED","PORT_CSP","PORT_TWCS") and (has_amounts or has_invoice_no):
        desig = "INVOICE"
        reason = f"DOC_TYPE={dt}, has invoice_no/amounts"
    elif dt in ("BOE_CUSTOMS","DELIVERY_ORDER","DELIVERY_NOTE","AIRPORT_FEES","APPOINTMENT","UNKNOWN"):
        desig = "EVIDENCE"
        reason = f"DOC_TYPE={dt}, supporting document"
    else:
        desig = "EVIDENCE"
        reason = "default"
    ws.cell(row=i, column=1, value=i-1)
    ws.cell(row=i, column=2, value=r["file"])
    ws.cell(row=i, column=3, value=desig)
    ws.cell(row=i, column=4, value=reason)
    ws.cell(row=i, column=5, value=r.get("verdict",""))
    style_data_row(ws, i, 5, verdict_col=5)

# === 02_Line_Audit
ws = wb.create_sheet("02_Line_Audit")
header_row(ws, 1, ["File", "DocType", "TypeB", "Table#", "Row#", "Cell content (preview)"],
           widths=[48, 16, 12, 8, 8, 70])
row = 2
for r in recs:
    tables = r.get("tables", [])
    for pg_i, pg_tables in enumerate(tables):
        for t_i, tbl in enumerate(pg_tables):
            for row_i, cells in enumerate(tbl):
                if not any((c or "").strip() for c in cells):
                    continue
                preview = " | ".join((c or "").strip()[:40] for c in cells if (c or "").strip())[:200]
                ws.cell(row=row, column=1, value=r["file"])
                ws.cell(row=row, column=2, value=r.get("doc_type",""))
                ws.cell(row=row, column=3, value=r.get("type_b",""))
                ws.cell(row=row, column=4, value=f"p{pg_i+1}.t{t_i+1}")
                ws.cell(row=row, column=5, value=row_i+1)
                ws.cell(row=row, column=6, value=preview)
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = BORDER
                    ws.cell(row=row, column=col).alignment = LEFT
                row += 1
if row == 2:
    ws.cell(row=2, column=1, value="(no tables extracted — all PDFs are image-only or text-layer without detectable table grid)")

# === 03_TYPE_B_Map
ws = wb.create_sheet("03_TYPE_B_Map")
header_row(ws, 1, ["TYPE_B", "Definition", "Count", "%"],
           widths=[14, 50, 8, 10])
tb_defs = {
    "Customs":   "BOE, Customs Duty, Debit Note, LAND_IMPORT (HE-0499)",
    "Inspection":"Container inspection, washing, admin fees",
    "DO":        "Delivery Order / D.O. fee",
    "INLAND":    "DN Road Freight, FB movement, inland transport",
    "THC":       "Terminal/port handling",
    "OTHERS":    "ISPS, Equipment Repositioning, Airport Fees, Appointment, CNT Repair, etc.",
}
for i, (tb, c) in enumerate(type_bs.most_common(), 2):
    ws.cell(row=i, column=1, value=tb).font = BOLD
    ws.cell(row=i, column=2, value=tb_defs.get(tb, "—"))
    ws.cell(row=i, column=3, value=c).number_format = FMT_INT
    ws.cell(row=i, column=4, value=f"=C{i}/COUNTA(A2:A{1+len(type_bs)})").number_format = FMT_PCT
    for col in range(1, 5):
        ws.cell(row=i, column=col).border = BORDER
        ws.cell(row=i, column=col).alignment = LEFT

# === 04_Key_Extraction
ws = wb.create_sheet("04_Key_Extraction")
header_row(ws, 1, ["File", "DocType", "Verdict", "Shipment_No", "Containers", "BL_No", "DO_No", "Invoice_No", "Date", "Amounts_AED_count"],
           widths=[44, 14, 9, 22, 28, 16, 14, 16, 14, 12])
for i, r in enumerate(recs, 2):
    k = r.get("keys", {})
    ws.cell(row=i, column=1, value=r["file"])
    ws.cell(row=i, column=2, value=r.get("doc_type",""))
    ws.cell(row=i, column=3, value=r.get("verdict",""))
    ws.cell(row=i, column=4, value=k.get("shipment_no",""))
    ws.cell(row=i, column=5, value="; ".join(k.get("containers",[])))
    ws.cell(row=i, column=6, value=k.get("bl_no",""))
    ws.cell(row=i, column=7, value=k.get("do_no",""))
    ws.cell(row=i, column=8, value=k.get("invoice_no",""))
    ws.cell(row=i, column=9, value=k.get("date",""))
    ws.cell(row=i, column=10, value=len(k.get("amounts_aed",[])))
    style_data_row(ws, i, 10, verdict_col=3)

# === 05_Amount_Verification
ws = wb.create_sheet("05_Amount_Verification")
header_row(ws, 1, ["File", "DocType", "TypeB", "Verdict",
                   "AED Line Sum", "AED Max", "AED Count", "FX Note", "Status"],
           widths=[44, 14, 12, 9, 18, 18, 10, 22, 12])
for i, r in enumerate(recs, 2):
    k = r.get("keys", {})
    amts = k.get("amounts_aed", [])
    line_sum = sum(amts)
    amt_max = max(amts) if amts else 0
    # FX: 1 USD ≈ 3.6725 AED (UAE dirham peg). Flag if mixed.
    fx_note = "AED only (HVDC scope)"
    # Status: Δ check (line sum vs grand total — grand total blank → AMBER)
    if r.get("doc_type") in ("CARRIER_RHS","CARRIER_EVG","PORT_ALLIED","PORT_CSP","PORT_TWCS"):
        status = "AMBER" if not amts else "OK (line-level)"
    elif r.get("verdict") == "PASS":
        status = "OK"
    else:
        status = "REVIEW"
    ws.cell(row=i, column=1, value=r["file"])
    ws.cell(row=i, column=2, value=r.get("doc_type",""))
    ws.cell(row=i, column=3, value=r.get("type_b",""))
    ws.cell(row=i, column=4, value=r.get("verdict",""))
    ws.cell(row=i, column=5, value=line_sum).number_format = FMT_AED
    ws.cell(row=i, column=6, value=amt_max).number_format = FMT_AED
    ws.cell(row=i, column=7, value=len(amts))
    ws.cell(row=i, column=8, value=fx_note)
    ws.cell(row=i, column=9, value=status)
    for col in range(1, 10):
        ws.cell(row=i, column=col).border = BORDER
        ws.cell(row=i, column=col).alignment = LEFT if col in (1,2,3,4,8,9) else RIGHT

def _is_aed(fn): return True  # all HVDC PDFs are AED; placeholder

# === 06_Document_Index
ws = wb.create_sheet("06_Document_Index")
header_row(ws, 1, ["File", "Size_KB", "Pages", "SHA256_16", "Processed", "Image_Only", "Tables", "Text_Len"],
           widths=[44, 10, 8, 22, 22, 12, 8, 10])
for i, r in enumerate(recs, 2):
    ws.cell(row=i, column=1, value=r["file"])
    ws.cell(row=i, column=2, value=f"{r.get('size_bytes',0)/1024:.2f}").number_format = FMT_NUM
    ws.cell(row=i, column=3, value=r.get("page_count",0))
    ws.cell(row=i, column=4, value=r.get("sha256",""))
    ws.cell(row=i, column=5, value=r.get("processed_at",""))
    ws.cell(row=i, column=6, value="YES" if r.get("image_only") else "no")
    ws.cell(row=i, column=7, value=r.get("table_count",0))
    ws.cell(row=i, column=8, value=r.get("text_length",0))
    for col in range(1, 9):
        ws.cell(row=i, column=col).border = BORDER
        ws.cell(row=i, column=col).alignment = LEFT

# === 07_Rule_Coverage
ws = wb.create_sheet("07_Rule_Coverage")
header_row(ws, 1, ["Rule", "DocType", "Header keyword", "Hit count", "% of total"],
           widths=[14, 18, 40, 10, 10])
rule_data = [
    ("R01","BOE_CUSTOMS",    "Customs Declaration / DEBIT NOTE / Gate Pass / LAND IMPORT",  doc_types.get("BOE_CUSTOMS",0)),
    ("R02","DELIVERY_ORDER", "DELIVERY ORDER / D/Order No / D.O. NUMBER / DELIVERY NOTIFICATION", doc_types.get("DELIVERY_ORDER",0)),
    ("R03","DELIVERY_NOTE",  "NOT NEGOTIABLE DELIVERY NOTE / Delivery Note/Waybill",     doc_types.get("DELIVERY_NOTE",0)),
    ("R04","PORT_ALLIED",    "ALLIED ONDOCK (priority over TAX INVOICE)",                  doc_types.get("PORT_ALLIED",0)),
    ("R05","PORT_CSP",       "CSP Abu Dhabi Terminal / Package Invoice",                   doc_types.get("PORT_CSP",0)),
    ("R06","PORT_TWCS",      "TransWorldContainer",                                        doc_types.get("PORT_TWCS",0)),
    ("R07","CARRIER_RHS",    "RAIS HASSAN SAADI / TAX INVOICE",                            doc_types.get("CARRIER_RHS",0)),
    ("R08","CARRIER_EVG",    "EVERGREEN SHIPPING AGENCY",                                  doc_types.get("CARRIER_EVG",0)),
    ("R09","AIRPORT_FEES",   "Etihad Terminal / Charges Summary",                          doc_types.get("AIRPORT_FEES",0)),
    ("R10","APPOINTMENT",    "IMPORT APPOINTMENT SUMMARY",                                 doc_types.get("APPOINTMENT",0)),
    ("R11","UNKNOWN",        "no rule matched (text layer 0 → OCR needed)",                doc_types.get("UNKNOWN",0)),
]
for i, (r, dt, kw, c) in enumerate(rule_data, 2):
    ws.cell(row=i, column=1, value=r).font = BOLD
    ws.cell(row=i, column=2, value=dt)
    ws.cell(row=i, column=3, value=kw)
    ws.cell(row=i, column=4, value=c).number_format = FMT_INT
    ws.cell(row=i, column=5, value=f"=D{i}/{total}").number_format = FMT_PCT
    for col in range(1, 6):
        ws.cell(row=i, column=col).border = BORDER
        ws.cell(row=i, column=col).alignment = LEFT

# === 08_Evidence_Issues
ws = wb.create_sheet("08_Evidence_Issues")
header_row(ws, 1, ["File", "Issue", "Detail", "Severity"],
           widths=[44, 22, 60, 12])
row = 2
def add_issue(file, issue, detail, severity):
    global row
    ws.cell(row=row, column=1, value=file)
    ws.cell(row=row, column=2, value=issue).font = BOLD
    ws.cell(row=row, column=3, value=detail)
    sc = ws.cell(row=row, column=4, value=severity)
    sc.font = BOLD
    if severity in ("HIGH","CRITICAL"): sc.fill = FAIL_FILL
    elif severity == "MEDIUM": sc.fill = AMBER_FILL
    else: sc.fill = PASS_FILL
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = BORDER
        ws.cell(row=row, column=col).alignment = LEFT
    row += 1

for r in recs:
    f = r["file"]
    if r.get("verdict") == "FAILED":
        add_issue(f, "PARSE_FAILED", r.get("error",""), "CRITICAL")
    if r.get("doc_type") == "UNKNOWN":
        if r.get("image_only"):
            add_issue(f, "IMAGE_ONLY", "text layer empty → Tesseract OCR fallback", "HIGH")
        else:
            add_issue(f, "UNKNOWN_DOCTYPE", "header fingerprint not matched", "MEDIUM")
    if not r.get("keys",{}).get("shipment_no"):
        if r.get("doc_type") != "UNKNOWN":
            add_issue(f, "NO_SHIPMENT_NO", "HVDC-ADOPT-XXXX pattern not found in body", "MEDIUM")
    if r.get("doc_type") in ("CARRIER_RHS","CARRIER_EVG","PORT_ALLIED","PORT_CSP","PORT_TWCS") and not r.get("keys",{}).get("amounts_aed"):
        add_issue(f, "GRAND_TOTAL_BLANK", "line-level evidence only; auto-PASS blocked (ZERO condition #3)", "HIGH")

# === 09_3-way_Reconcile
ws = wb.create_sheet("09_3-way_Reconcile")
ws["A1"] = "3-way reconciliation: Final Subtotal = Line_Audit = TYPE-B (±0.01)"
ws["A1"].font = BOLD
ws.merge_cells("A1:E1")
ws["A2"] = "Per HVDC contract, 13-sheet workbook requires 3-way integrity. This PDF-only audit run has no Grand Total on invoices (per pdfparse.md §ZERO #3), so 3-way is N/A for carrier/port invoices."
ws["A2"].alignment = LEFT
ws["A2"].font = Font(italic=True, size=9)
ws.merge_cells("A2:E2")
ws.row_dimensions[2].height = 36
header_row(ws, 4, ["File", "DocType", "TypeB", "Line Sum AED", "Reconcile Status"],
           widths=[44, 14, 12, 18, 18])
for i, r in enumerate(recs, 5):
    k = r.get("keys", {})
    line_sum = sum(k.get("amounts_aed",[]))
    if r.get("doc_type") in ("CARRIER_RHS","CARRIER_EVG","PORT_ALLIED","PORT_CSP","PORT_TWCS"):
        status = "N/A (grand total blank — line-level only)"
    else:
        status = "N/A (evidence doc)"
    ws.cell(row=i, column=1, value=r["file"])
    ws.cell(row=i, column=2, value=r.get("doc_type",""))
    ws.cell(row=i, column=3, value=r.get("type_b",""))
    ws.cell(row=i, column=4, value=line_sum).number_format = FMT_AED
    ws.cell(row=i, column=5, value=status)
    for col in range(1, 6):
        ws.cell(row=i, column=col).border = BORDER
        ws.cell(row=i, column=col).alignment = LEFT

# === 10_FX_Policy
ws = wb.create_sheet("10_FX_Policy")
ws["A1"] = "FX Policy — per HVDC rules, only AED/USD allowed (1 USD = 3.6725 AED peg)"
ws["A1"].font = BOLD
ws.merge_cells("A1:D1")
header_row(ws, 3, ["File", "Currency detected", "AED amount", "FX Note"],
           widths=[44, 18, 18, 50])
for i, r in enumerate(recs, 4):
    k = r.get("keys", {})
    amts = k.get("amounts_aed", [])
    ws.cell(row=i, column=1, value=r["file"])
    ws.cell(row=i, column=2, value="AED" if amts else "—")
    ws.cell(row=i, column=3, value=sum(amts)).number_format = FMT_AED
    ws.cell(row=i, column=4, value="OK" if amts else "No AED amount extracted (check OCR or text layer)")
    for col in range(1, 5):
        ws.cell(row=i, column=col).border = BORDER
        ws.cell(row=i, column=col).alignment = LEFT

# === 11_Approval_Gate
ws = wb.create_sheet("11_Approval_Gate")
header_row(ws, 1, ["#", "File", "DocType", "Verdict", "Gate Reason", "Action"],
           widths=[5, 44, 14, 9, 50, 24])
gate_reasons = {
    "PASS":   "All evidence checks passed; line-level verified.",
    "AMBER":  "Reviewer approval required (BOE final / line-level only / image-only).",
    "ZERO":   "Critical: blank grand total, BOE final, or DEM/DET. Block auto-PASS.",
    "FAILED": "Parser error. Inspect Issues sheet.",
}
actions = {
    "PASS":   "Archive",
    "AMBER":  "Reviewer queue",
    "ZERO":   "Escalate to PRIME",
    "FAILED": "Manual triage",
}
for i, r in enumerate(recs, 2):
    v = r.get("verdict","")
    ws.cell(row=i, column=1, value=i-1)
    ws.cell(row=i, column=2, value=r["file"])
    ws.cell(row=i, column=3, value=r.get("doc_type",""))
    ws.cell(row=i, column=4, value=v)
    ws.cell(row=i, column=5, value=gate_reasons.get(v,"—"))
    ws.cell(row=i, column=6, value=actions.get(v,"—"))
    style_data_row(ws, i, 6, verdict_col=4)

# === 99_Manifest
ws = wb.create_sheet("99_Manifest")
ws["A1"] = "HVDC Audit Pack Manifest"
ws["A1"].font = Font(bold=True, size=14)
ws.merge_cells("A1:C1")
header_row(ws, 3, ["#", "File", "Verdict"],
           widths=[5, 56, 10])
for i, r in enumerate(recs, 4):
    ws.cell(row=i, column=1, value=i-3)
    ws.cell(row=i, column=2, value=r["file"])
    ws.cell(row=i, column=3, value=r.get("verdict",""))
    style_data_row(ws, i, 3, verdict_col=3)
last_row = 3 + len(recs)
ws.cell(row=last_row+2, column=1, value="Generated").font = BOLD
ws.cell(row=last_row+2, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
ws.cell(row=last_row+3, column=1, value="Parser").font = BOLD
ws.cell(row=last_row+3, column=2, value="pdf_hybrid_parser.py v3")
ws.cell(row=last_row+4, column=1, value="Source").font = BOLD
ws.cell(row=last_row+4, column=2, value=str(SRC.relative_to(ROOT)))
ws.cell(row=last_row+5, column=1, value="Rule").font = BOLD
ws.cell(row=last_row+5, column=2, value="HVDC Rule #0 — workbook always produced regardless of verdict")

# Reorder sheets to enforce contract
order = ["00_Decision","01_Invoice_Source","02_Line_Audit","03_TYPE_B_Map",
         "04_Key_Extraction","05_Amount_Verification","06_Document_Index",
         "07_Rule_Coverage","08_Evidence_Issues","09_3-way_Reconcile",
         "10_FX_Policy","11_Approval_Gate","99_Manifest"]
wb._sheets = [wb[name] for name in order]

wb.save(OUT)
print(f"[OK] {OUT}")
print(f"     Sheets: {len(wb.sheetnames)} → {wb.sheetnames}")
