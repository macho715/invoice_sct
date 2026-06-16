"""xlsx parser using openpyxl. Supports header alias detection (FR-014)."""
from __future__ import annotations
import io, re
from openpyxl import load_workbook
from app.schemas import InvoiceLine, NormalizedInvoice, InvoiceHeader, normalize_line_id

HEADER_ALIASES: dict[str, list[str]] = {
    'description': [
        'description', 'desc', 'charge', 'charge description', 'item', 'item description',
        'line item', 'service', 'service description', 'particulars', 'details',
        'goods description', 'cargo description', 'description of goods', 'description of services',
        'charge item', 'charge name', 'fee description', 'billing item',
        'operation type', 'transport type',
    ],
    'qty': [
        'qty', 'quantity', 'units', 'count', 'trips', '# trips', 'no of trips',
        'q/ty', 'number', 'no.', 'nos', 'trip count', 'trip', '# of trips',
    ],
    'rate': [
        'rate', 'unit rate', 'unit_price', 'unit price', 'price',
        'rate (usd)', 'rate (aed)', 'rate(usd)', 'rate(aed)',
        'rate usd', 'rate aed', 'usd rate', 'aed rate',
        'invoice rate', 'invoice rate usd', 'invoice rate aed',
        'contract rate', 'unit cost', 'unit amount', 'per unit',
        'applied rate',
    ],
    'amount': [
        'amount', 'line amount', 'line_amount', 'total', 'line total',
        'amount (usd)', 'amount (aed)', 'amount(usd)', 'amount(aed)',
        'amount usd', 'amount aed', 'usd amount', 'aed amount',
        'total (usd)', 'total (aed)', 'total usd', 'total aed',
        'line total usd', 'line total aed', 'sum', 'subtotal',
        'total amount',
    ],
    'currency': [
        'currency', 'ccy', 'curr', 'currency code',
    ],
    'shipment_ref': [
        'shipment no', 'shipment_no', 'shipment ref', 'shipment_ref',
        'shipment id', 'shpt', 'hvdc', 'reference', 'ref',
        'shipment reference', 'shipment', 'shipment#', 'ref no',
        'reference no', 'reference number', 'shipment reference#', 'shpt ref',
    ],
    'job_number': [
        'job no', 'job_no', 'job number', 'job_number', 'job #', 'job#',
        'job id', 'job_id', 'project no', 'project_no', 'wo no', 'wo_no',
        'work order', 'po no', 'po_no', 'purchase order', 'job #',
    ],
    'rate_basis': [
        'rate basis', 'rate_basis', 'unit', 'uom', 'basis',
        'per unit', 'unit of measure', 'measure', 'rate type',
        'billing unit', 'charge unit',
    ],
    'charge_component': [
        'charge component', 'for_charge_component', 'type', 'type b', 'type_b',
        'charge type', 'category', 'cost category', 'cost type',
        'expense type', 'charge code', 'cost code', 'mode',
    ],
    'formula_text': [
        'formula text', 'formula_text', 'formula', 'formula as text',
        'calculation text', 'calc text', 'formula_text_original',
    ],
}

HEADER_FIELD_LABELS: dict[str, list[str]] = {
    'invoice_no':  ['invoice no', 'invoice #', 'inv no', 'invoice number', 'invoice no.',
                    'inv #', 'bill no', 'bill number'],
    'vendor':      ['vendor', 'supplier', 'billed by', 'from', 'vendor name', 'supplier name',
                    'company', 'company name', 'billing company'],
    'issue_date':  ['date', 'issue date', 'invoice date', 'date of issue', 'billing date',
                    'created', 'doc date'],
    'invoice_total': ['invoice total', 'total amount', 'grand total', 'total due',
                       'total invoice', 'amount due'],
}

def _normalize_key(s: str) -> str:
    return re.sub(r'[^a-z0-9]', '', s.strip().lower())

def _detect_headers(header_row: list) -> dict[str, int]:
    """Map canonical field ??column index using fuzzy header matching.
    Tries exact match first, then normalized match (strips all non-alphanum),
    then word-boundary substring match as last resort."""
    norm_exact = [str(c or '').strip().lower() for c in header_row]
    norm_fuzzy = [_normalize_key(str(c or '')) for c in header_row]
    out: dict[str, int] = {}
    used_cols: set[int] = set()
    for canon, aliases in HEADER_ALIASES.items():
        alias_exact = [a.strip().lower() for a in aliases]
        alias_fuzzy = [_normalize_key(a) for a in aliases]
        for i, cell in enumerate(norm_exact):
            if i in used_cols:
                continue
            if cell in alias_exact:
                out[canon] = i
                used_cols.add(i)
                break
            if norm_fuzzy[i] and norm_fuzzy[i] in alias_fuzzy:
                out[canon] = i
                used_cols.add(i)
                break
        if canon not in out:
            for i, cell in enumerate(norm_exact):
                if i in used_cols:
                    continue
                for a in alias_exact:
                    if len(a) >= 3 and re.search(r'\b' + re.escape(a) + r'\b', cell):
                        out[canon] = i
                        used_cols.add(i)
                        break
                if canon in out:
                    break
    return out

def _detect_header_fields(rows: list, max_scan: int = 10) -> dict[str, str | None]:
    result: dict[str, str | None] = {'invoice_no': None, 'vendor': None, 'issue_date': None, 'invoice_total': None}
    found: set[str] = set()
    for r_idx in range(min(max_scan, len(rows))):
        row = rows[r_idx]
        for c_idx, cell in enumerate(row):
            if cell is None:
                continue
            text = str(cell).strip().lower().rstrip(':')
            for field, aliases in HEADER_FIELD_LABELS.items():
                if field in found:
                    continue
                if text in aliases or _normalize_key(text) in [_normalize_key(a) for a in aliases]:
                    value = None
                    if c_idx + 1 < len(row):
                        value = _cell_str(row[c_idx + 1])
                    if value is None and r_idx + 1 < len(rows):
                        below_row = rows[r_idx + 1]
                        if c_idx < len(below_row):
                            value = _cell_str(below_row[c_idx])
                    if value:
                        result[field] = value
                        found.add(field)
                    elif field == 'invoice_total':
                        pass
    return result

def _cell_num(cell) -> float | None:
    if cell is None: return None
    if isinstance(cell, (int, float)): return float(cell)
    try: return float(str(cell).replace(',', ''))
    except (ValueError, TypeError): return None

def _cell_str(cell) -> str | None:
    if cell is None: return None
    s = str(cell).strip()
    return s if s else None

def _currency_from_amount_header(header: object) -> str | None:
    h = _normalize_key(str(header or ''))
    if 'usd' in h:
        return 'USD'
    if 'aed' in h:
        return 'AED'
    return None

def _amount_column_for_row(header_row: list, cmap: dict[str, int]) -> int:
    amount_col = cmap['amount']
    current_currency = _currency_from_amount_header(header_row[amount_col])
    if current_currency != 'AED':
        return amount_col

    for idx, cell in enumerate(header_row):
        if _currency_from_amount_header(cell) != 'USD':
            continue
        candidate_map = _detect_headers([cell])
        if 'amount' in candidate_map:
            return idx
    return amount_col

def _clean_header(s: str) -> str:
    """Collapse embedded newlines / repeated spaces in a column header."""
    return re.sub(r'\s+', ' ', str(s or '').strip())

def _is_charge_col(norm_header: str) -> bool:
    """A DSV summary-matrix charge column (one charge component per column).
    Excludes grand-total / total-amount / remark / BOE meta columns."""
    h = norm_header.lower()
    if not h:
        return False
    if 'grand total' in h or 'total amount' in h or 'remark' in h:
        return False
    if 'bill of entry' in h or 'boe' in h:
        return False
    if any(k in h for k in ('charge', 'surcharge', 'duty', 'at cost')):
        return True
    return 'amount' in h  # e.g. "ADDITIONAL AMOUNT" (totals already excluded above)

def _try_parse_matrix(rows: list, *, ws_title: str, file_id: str, parser_version: str) -> NormalizedInvoice | None:
    """Detect & decompose a DSV summary-matrix layout (charge=column, shipment=row).

    Returns a NormalizedInvoice with one line per non-empty charge cell, or None
    when the sheet is not a matrix (caller then falls back to the flat parser).
    Charge values are USD (the per-charge columns sum to GRAND TOTAL (USD))."""
    for i in range(min(20, len(rows))):
        row = list(rows[i])
        norm = [_clean_header(str(c or '')).lower() for c in row]
        charge_cols = [(j, _clean_header(str(row[j]))) for j, h in enumerate(norm) if _is_charge_col(h)]
        if len(charge_cols) < 3:
            continue
        cmap = _detect_headers(row)
        shp_col = cmap.get('shipment_ref')
        if shp_col is None:
            continue
        job_col = cmap.get('job_number')
        gt_usd = next((j for j, h in enumerate(norm) if 'grand total' in h and 'usd' in h), None)

        lines: list[InvoiceLine] = []
        for r_idx, drow in enumerate(rows[i + 1:], start=i + 2):
            shp = _cell_str(drow[shp_col]) if shp_col < len(drow) else None
            if not shp:
                break  # end of contiguous shipment block (blank / TOTAL row)
            job = _cell_str(drow[job_col]) if job_col is not None and job_col < len(drow) else None
            for cj, cname in charge_cols:
                if cj >= len(drow):
                    continue
                val = _cell_num(drow[cj])
                if val is None or val == 0:
                    continue
                lines.append(InvoiceLine(
                    line_id=normalize_line_id(file_id, ws_title, r_idx, cj),
                    description=cname,
                    normalized_description=cname.upper(),
                    currency='USD',
                    amount=float(val),
                    qty=None,
                    rate=None,
                    source_ref={'sheet': ws_title, 'row': r_idx, 'col': str(cj)},
                    shipment_ref=shp,
                    job_number=job,
                    for_charge_component=cname,
                ))
        if not lines:
            return None

        invoice_total: float | None = None
        for trow in rows[i + 1:]:
            cells = [_clean_header(str(c or '')).lower() for c in trow]
            if any('total amount' in c or c == 'total' or 'grand total' in c for c in cells):
                if gt_usd is not None and gt_usd < len(trow):
                    invoice_total = _cell_num(trow[gt_usd])
                if invoice_total is None:
                    for c in trow:
                        v = _cell_num(c)
                        if v is not None:
                            invoice_total = v
                            break
                if invoice_total is not None:
                    break

        return NormalizedInvoice(
            invoice_id=f"inv_{file_id}",
            invoice_header=InvoiceHeader(
                invoice_no=None, vendor=None, issue_date=None,
                currency='USD', invoice_total=invoice_total
            ),
            invoice_lines=lines,
            evidence_candidates=[],
            parser_confidence=0.85,
            parser_version=parser_version
        )
    return None

def _parse_xlsx_rows(rows: list, *, ws_title: str, file_id: str, parser_version: str, formula_rows: list | None = None) -> NormalizedInvoice:
    if not rows:
        raise ValueError("empty xlsx")

    # DSV summary-matrix layout (charge-per-column) — decompose into charge lines.
    # Falls through to the flat one-charge-per-row parser when not a matrix.
    matrix = _try_parse_matrix(rows, ws_title=ws_title, file_id=file_id, parser_version=parser_version)
    if matrix is not None:
        return matrix

    header_fields = _detect_header_fields(rows)

    header_row_idx: int | None = None
    cmap: dict[str, int] = {}
    for i in range(min(20, len(rows))):
        cmap = _detect_headers(list(rows[i]))
        if 'description' in cmap and 'amount' in cmap:
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError("xlsx missing required columns (description, amount)")

    lines: list[InvoiceLine] = []
    amount_col = _amount_column_for_row(list(rows[header_row_idx]), cmap)
    amount_header_currency = _currency_from_amount_header(rows[header_row_idx][amount_col])
    for r_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
        desc = _cell_str(row[cmap['description']]) if 'description' in cmap else None
        amt  = _cell_num(row[amount_col]) if amount_col < len(row) else None
        if desc is None and amt is None:
            continue
        currency = (_cell_str(row[cmap['currency']]) or amount_header_currency or 'AED') if 'currency' in cmap else (amount_header_currency or 'AED')
        if currency not in ('AED', 'USD'):
            currency = 'AED'
        line = InvoiceLine(
            line_id=normalize_line_id(file_id, ws_title, r_idx, cmap['description']),
            description=desc or '(missing description)',
            currency=currency,
            amount=float(amt or 0.0),
            qty=_cell_num(row[cmap['qty']]) if 'qty' in cmap else None,
            rate=_cell_num(row[cmap['rate']]) if 'rate' in cmap else None,
            source_ref={'sheet': ws_title, 'row': r_idx, 'col': str(amount_col)},
            shipment_ref=_cell_str(row[cmap['shipment_ref']]) if 'shipment_ref' in cmap else None,
            job_number=_cell_str(row[cmap['job_number']]) if 'job_number' in cmap else None,
            rate_basis=_cell_str(row[cmap['rate_basis']]) if 'rate_basis' in cmap else None,
            for_charge_component=_cell_str(row[cmap['charge_component']]) if 'charge_component' in cmap else None
        )
        if 'formula_text' in cmap:
            formula_row = formula_rows[r_idx - 1] if formula_rows and r_idx - 1 < len(formula_rows) else row
            formula_text = _cell_str(formula_row[cmap['formula_text']]) if cmap['formula_text'] < len(formula_row) else None
            if formula_text:
                line.source_ref['formula_text'] = formula_text
        lines.append(line)

    if not lines:
        raise ValueError("xlsx has no usable invoice line")

    header_currency = lines[0].currency
    invoice_total_str = header_fields.get('invoice_total')
    invoice_total = None
    if invoice_total_str:
        try:
            invoice_total = float(str(invoice_total_str).replace(',', '').replace('$', '').replace('\u20a9', ''))
        except (ValueError, TypeError):
            pass
    if invoice_total is None:
        for row in reversed(rows[-10:]):
            for cell in row:
                if cell is None:
                    continue
                text = str(cell).strip().lower().rstrip(':')
                if text in ['total', 'grand total', 'total amount', 'total due', 'invoice total']:
                    for c in row:
                        if c is not None and isinstance(c, (int, float)):
                            invoice_total = float(c)
                            break
                        if c is not None and str(c).replace(',', '').replace('.', '').replace('-', '').isdigit():
                            invoice_total = float(str(c).replace(',', ''))
                            break
                    break
            if invoice_total is not None:
                break

    return NormalizedInvoice(
        invoice_id=f"inv_{file_id}",
        invoice_header=InvoiceHeader(
            invoice_no=header_fields.get('invoice_no'),
            vendor=header_fields.get('vendor'),
            issue_date=header_fields.get('issue_date'),
            currency=header_currency,
            invoice_total=invoice_total
        ),
        invoice_lines=lines,
        evidence_candidates=[],
        parser_confidence=0.9,
        parser_version=parser_version
    )


def parse_xlsx_bytes(raw: bytes, *, file_id: str, file_name: str, parser_version: str) -> NormalizedInvoice:
    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    formula_wb = load_workbook(io.BytesIO(raw), data_only=False, read_only=True)
    if not wb.worksheets:
        raise ValueError("empty xlsx")

    preferred = {
        '04_line_view',
        '90_source_data',
        'invoice',
    }
    worksheets = sorted(
        wb.worksheets,
        key=lambda ws: (0 if ws.title.strip().lower() in preferred else 1, wb.worksheets.index(ws)),
    )

    first_error: ValueError | None = None
    for ws in worksheets:
        rows = list(ws.iter_rows(values_only=True))
        formula_ws = formula_wb[ws.title] if ws.title in formula_wb.sheetnames else None
        formula_rows = list(formula_ws.iter_rows(values_only=True)) if formula_ws is not None else None
        try:
            return _parse_xlsx_rows(rows, ws_title=ws.title, file_id=file_id, parser_version=parser_version, formula_rows=formula_rows)
        except ValueError as exc:
            if first_error is None:
                first_error = exc
            continue

    if first_error is not None:
        raise first_error
    raise ValueError("empty xlsx")
