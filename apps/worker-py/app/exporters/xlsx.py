import hashlib
import openpyxl
from io import BytesIO
from datetime import datetime, timezone
from app.schemas import (
    ExportRequest,
    HeaderCheckRow,
    DuplicateCheckRow,
    RateCheckRow,
    TaxFxCheckRow,
    ShipmentMatchRow,
)


def build_xlsx(req: ExportRequest) -> bytes:
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    ws_dec = wb.create_sheet("00_Decision")
    dec_cols = ["job_id", "verdict", "approved_by", "approved_at", "rule_version", "parser_version", "final_recon_status", "zero_count", "amber_count", "prism_kernel_proof_ref", "costguard_band_summary", "watermark", "generated_at"]
    ws_dec.append(dec_cols)
    sorted_dec = sorted(req.decision_rows, key=lambda x: x.job_id or "")
    for row in sorted_dec:
        ws_dec.append([
            row.job_id,
            row.verdict,
            row.approved_by,
            row.approved_at,
            row.rule_version,
            row.parser_version,
            row.final_recon_status,
            row.zero_count,
            row.amber_count,
            row.prism_kernel_proof_ref,
            row.costguard_band_summary,
            row.watermark,
            row.generated_at or req.generated_at,
        ])

    ws_act = wb.create_sheet("01_Action_Items")
    act_cols = ["action_id", "severity", "shipment_ref", "line_id", "issue_type", "required_action", "owner_role", "status", "prism_kernel_proof_ref"]
    ws_act.append(act_cols)
    sorted_act = sorted(req.action_items_rows, key=lambda x: x.action_id or "")
    for row in sorted_act:
        ws_act.append([
            row.action_id,
            row.severity,
            row.shipment_ref,
            row.line_id,
            row.issue_type,
            row.required_action,
            row.owner_role,
            row.status,
            row.prism_kernel_proof_ref,
        ])

    ws_recon = wb.create_sheet("02_Final_Recon")
    recon_cols = ["currency", "shipment_ref", "invoice_total", "reviewed_total", "variance", "variance_pct", "recon_status", "evidence_ref"]
    ws_recon.append(recon_cols)
    sorted_recon = sorted(req.final_recon_rows, key=lambda x: (x.shipment_ref or "", x.currency or ""))
    for row in sorted_recon:
        ws_recon.append([
            row.currency,
            row.shipment_ref,
            row.invoice_total,
            row.reviewed_total,
            row.variance,
            row.variance_pct,
            row.recon_status,
            row.evidence_ref,
        ])

    ws_hdr = wb.create_sheet("03_Header_Check")
    hdr_cols = ["field_name", "expected_value", "actual_value", "match_status", "severity"]
    ws_hdr.append(hdr_cols)
    for row in sorted(getattr(req, "header_check_rows", []), key=lambda x: x.field_name or ""):
        ws_hdr.append([
            row.field_name,
            row.expected_value,
            row.actual_value,
            row.match_status,
            row.severity,
        ])

    ws_line = wb.create_sheet("04_Line_View")
    line_cols = ["line_id", "shipment_ref", "description", "for_charge_component", "type_b", "amount", "currency", "rate_source", "rate_status", "validity_status", "evidence_status", "gate_status", "band", "delta_pct", "numeric_integrity_status", "difference"]
    ws_line.append(line_cols)
    sorted_lines = sorted(req.line_view_rows, key=lambda x: x.line_id or "")
    for row in sorted_lines:
        ws_line.append([
            row.line_id,
            row.shipment_ref,
            row.description,
            row.for_charge_component,
            row.type_b,
            row.amount,
            row.currency,
            row.rate_source,
            row.rate_status,
            row.validity_status,
            row.evidence_status,
            row.gate_status,
            row.band,
            row.delta_pct,
            row.numeric_integrity_status,
            row.difference,
        ])

    ws_dup = wb.create_sheet("05_Duplicate_Check")
    dup_cols = ["invoice_no_hash", "vendor_hash", "amount_hash", "issue_date_hash", "match_type", "severity", "matched_job_id"]
    ws_dup.append(dup_cols)
    for row in sorted(getattr(req, "duplicate_check_rows", []), key=lambda x: x.invoice_no_hash or ""):
        ws_dup.append([
            row.invoice_no_hash,
            row.vendor_hash,
            row.amount_hash,
            row.issue_date_hash,
            row.match_type,
            row.severity,
            row.matched_job_id,
        ])

    ws_rate = wb.create_sheet("06_Rate_Check")
    rate_cols = ["line_id", "charge_code", "lane", "contract_rate", "invoiced_rate", "rate_basis", "effective_from", "effective_to", "rate_status", "delta_pct", "severity"]
    ws_rate.append(rate_cols)
    for row in sorted(getattr(req, "rate_check_rows", []), key=lambda x: x.line_id or ""):
        ws_rate.append([
            row.line_id,
            row.charge_code,
            row.lane,
            row.contract_rate,
            row.invoiced_rate,
            row.rate_basis,
            row.effective_from,
            row.effective_to,
            row.rate_status,
            row.delta_pct,
            row.severity,
        ])

    ws_tax = wb.create_sheet("07_Tax_FX_Check")
    tax_cols = ["line_id", "currency", "vat_rate", "vat_amount", "fx_rate_applied", "fx_policy_id", "tax_status", "fx_status", "severity"]
    ws_tax.append(tax_cols)
    for row in sorted(getattr(req, "tax_fx_check_rows", []), key=lambda x: x.line_id or ""):
        ws_tax.append([
            row.line_id,
            row.currency,
            row.vat_rate,
            row.vat_amount,
            row.fx_rate_applied,
            row.fx_policy_id,
            row.tax_status,
            row.fx_status,
            row.severity,
        ])

    ws_ship = wb.create_sheet("08_Shipment_Match")
    ship_cols = ["line_id", "shipment_ref", "job_number", "bl_number", "do_number", "po_number", "match_status", "matched_fields", "severity"]
    ws_ship.append(ship_cols)
    for row in sorted(getattr(req, "shipment_match_rows", []), key=lambda x: x.line_id or ""):
        ws_ship.append([
            row.line_id,
            row.shipment_ref,
            row.job_number,
            row.bl_number,
            row.do_number,
            row.po_number,
            row.match_status,
            row.matched_fields,
            row.severity,
        ])

    ws_src = wb.create_sheet("90_Source_Data")
    src_cols = ["file_id", "source_ref", "original_text", "normalized_value", "confidence", "routing_pattern", "pdf_page", "text_span_hash"]
    ws_src.append(src_cols)
    sorted_src = sorted(req.source_data_rows, key=lambda x: (x.file_id or "", x.source_ref or ""))
    for row in sorted_src:
        ws_src.append([
            row.file_id,
            row.source_ref,
            row.original_text,
            row.normalized_value,
            row.confidence,
            row.routing_pattern,
            getattr(row, "pdf_page", None),
            getattr(row, "text_span_hash", None),
        ])

    ws_audit = wb.create_sheet("91_Audit_Detail")
    audit_cols = ["line_id", "rule_id", "reason_code", "sct_trace_id", "cf_mcp_tool", "cf_mcp_latency_ms", "confidence", "decision_input", "fx_override", "fx_policy_id"]
    ws_audit.append(audit_cols)
    sorted_audit = sorted(req.audit_detail_rows, key=lambda x: (x.line_id or "", x.rule_id or ""))
    for row in sorted_audit:
        ws_audit.append([
            row.line_id,
            row.rule_id,
            row.reason_code,
            row.sct_trace_id,
            row.cf_mcp_tool,
            row.cf_mcp_latency_ms,
            row.confidence,
            row.decision_input,
            row.fx_override,
            row.fx_policy_id,
        ])

    ws_ev = wb.create_sheet("92_Evidence_Issues")
    ev_cols = ["line_id", "required_evidence", "matched_evidence", "gap_type", "severity", "action_item_id", "human_gate_trigger_id"]
    ws_ev.append(ev_cols)
    sorted_ev = sorted(req.evidence_issues_rows, key=lambda x: (x.line_id or "", x.gap_type or ""))
    for row in sorted_ev:
        ws_ev.append([
            row.line_id,
            row.required_evidence,
            row.matched_evidence,
            row.gap_type,
            row.severity,
            row.action_item_id,
            row.human_gate_trigger_id,
        ])

    parser_ver = ""
    rule_ver = ""
    if req.decision_rows:
        parser_ver = req.decision_rows[0].parser_version or ""
        rule_ver = req.decision_rows[0].rule_version or ""
    generated_at = req.generated_at or datetime.now(timezone.utc).isoformat()

    ws_manifest = wb.create_sheet("99_Manifest")
    ws_manifest.append(["key", "value"])
    manifest_rows = [
        ("pre_manifest_sha256", "PENDING"),
        ("sheet_count", 13),
        ("generated_at", generated_at),
        ("parser_version", parser_ver),
        ("rule_version", rule_ver),
    ]
    manifest_rows.extend((entry.key, entry.value) for entry in getattr(req, "manifest_entries", []))
    manifest_data_start_row = 2
    for k, v in manifest_rows:
        ws_manifest.append([k, v])

    for ws in wb.worksheets:
        try:
            ws.freeze_panes = "A2"
            if getattr(ws, "max_row", 0) >= 1 and getattr(ws, "max_column", 0) >= 1:
                ws.auto_filter.ref = ws.dimensions
        except Exception:
            pass

    buf_pass1 = BytesIO()
    wb.save(buf_pass1)
    pass1_bytes = buf_pass1.getvalue()
    sha256_hex = hashlib.sha256(pass1_bytes).hexdigest()

    sha256_cell = ws_manifest.cell(row=manifest_data_start_row, column=2)
    sha256_cell.value = sha256_hex

    buf_pass2 = BytesIO()
    wb.save(buf_pass2)
    return buf_pass2.getvalue()
