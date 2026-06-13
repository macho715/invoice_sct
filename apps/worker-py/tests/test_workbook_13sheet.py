"""Tests for 13-Sheet Track 2 Workbook Contract."""
import pytest
from app.schemas import (
    ExportRequest, DecisionRow, ActionItemRow, FinalReconRow,
    HeaderCheckRow, LineViewRow, DuplicateCheckRow, RateCheckRow,
    TaxFxCheckRow, ShipmentMatchRow, SourceDataRow, AuditDetailRow,
    EvidenceIssuesRow
)
from app.exporters.xlsx import build_xlsx
import openpyxl
from io import BytesIO

REQUIRED_SHEETS = [
    "00_Decision", "01_Action_Items", "02_Final_Recon",
    "03_Header_Check", "04_Line_View", "05_Duplicate_Check",
    "06_Rate_Check", "07_Tax_FX_Check", "08_Shipment_Match",
    "90_Source_Data", "91_Audit_Detail", "92_Evidence_Issues",
    "99_Manifest"
]

def _make_minimal_export():
    return ExportRequest(
        job_id="job_test001",
        decision_rows=[DecisionRow(
            job_id="job_test001", verdict="PASS", approved_by=None, approved_at=None,
            rule_version="rule-0.1.0", parser_version="parser-0.1.0",
            final_recon_status="MATCHED", zero_count=0, amber_count=0,
            prism_kernel_proof_ref=None, costguard_band_summary="PASS: 1", watermark="TEST", generated_at="2026-06-13T00:00:00Z"
        )],
        action_items_rows=[],
        final_recon_rows=[FinalReconRow(
            currency="AED", shipment_ref="HVDC-TEST-001",
            invoice_total=1000.0, reviewed_total=1000.0,
            variance=0.0, variance_pct=0.0, recon_status="MATCHED", evidence_ref=None
        )],
        header_check_rows=[HeaderCheckRow(
            field_name="invoice_no", expected_value=None, actual_value="INV-001",
            match_status="PRESENT", severity=None
        )],
        line_view_rows=[LineViewRow(
            line_id="line_001", shipment_ref="HVDC-TEST-001",
            description="Test charge", for_charge_component=None, type_b=None,
            amount=1000.0, currency="AED", rate_source=None, rate_status=None,
            validity_status=None, evidence_status=None, gate_status=None,
            band=None, delta_pct=None, numeric_integrity_status=None, difference=None
        )],
        duplicate_check_rows=[],
        rate_check_rows=[RateCheckRow(
            line_id="line_001", charge_code="TRUCKING", lane="DXB-RUW",
            contract_rate=500.0, invoiced_rate=500.0, rate_basis="PER_TRUCK",
            effective_from="2026-01-01", effective_to="2026-12-31",
            rate_status="MATCHED", delta_pct=0.0, severity="PASS"
        )],
        tax_fx_check_rows=[TaxFxCheckRow(
            line_id="line_001", currency="AED", vat_rate=5.0, vat_amount=50.0,
            fx_rate_applied=None, fx_policy_id=None,
            tax_status="COMPLIANT", fx_status="NOT_APPLICABLE", severity="PASS"
        )],
        shipment_match_rows=[ShipmentMatchRow(
            line_id="line_001", shipment_ref="HVDC-TEST-001", job_number="JOB-001",
            bl_number=None, do_number="DO-001", po_number=None,
            match_status="MATCHED", matched_fields="shipment_ref,job_number", severity="PASS"
        )],
        source_data_rows=[SourceDataRow(
            file_id="file_001", source_ref="sheet1", original_text="Test",
            normalized_value="Test", confidence=0.9, routing_pattern="MANUAL"
        )],
        audit_detail_rows=[AuditDetailRow(
            line_id="line_001", rule_id="CG_RULE_PASS", reason_code="OK",
            sct_trace_id="trace_001", cf_mcp_tool="check_cost_guard",
            cf_mcp_latency_ms=50.0, confidence=0.95, decision_input=None,
            fx_override=None, fx_policy_id=None
        )],
        evidence_issues_rows=[],
        generated_at="2026-06-13T00:00:00Z"
    )

def test_workbook_has_exactly_13_sheets():
    req = _make_minimal_export()
    data = build_xlsx(req)
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True)
    assert len(wb.sheetnames) == 13, f"Expected 13 sheets, got {len(wb.sheetnames)}: {wb.sheetnames}"

def test_workbook_sheet_names_match_contract():
    req = _make_minimal_export()
    data = build_xlsx(req)
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True)
    assert wb.sheetnames == REQUIRED_SHEETS, f"Sheet order mismatch.\nExpected: {REQUIRED_SHEETS}\nActual: {wb.sheetnames}"

def test_workbook_no_track1_type_b_sheet():
    req = _make_minimal_export()
    data = build_xlsx(req)
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True)
    assert "03_Type_B_Summary" not in wb.sheetnames

def test_workbook_freeze_panes():
    req = _make_minimal_export()
    data = build_xlsx(req)
    wb = openpyxl.load_workbook(BytesIO(data), read_only=False)
    for ws in wb.worksheets:
        assert ws.freeze_panes == 'A2', f"Sheet {ws.title} missing freeze_panes"

def test_manifest_sheet_contains_sha256():
    req = _make_minimal_export()
    data = build_xlsx(req)
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True)
    ws = wb["99_Manifest"]
    rows = list(ws.iter_rows(values_only=True))
    keys = [r[0] for r in rows[1:]]
    assert "workbook_sha256" in keys
    assert "sheet_count" in keys
