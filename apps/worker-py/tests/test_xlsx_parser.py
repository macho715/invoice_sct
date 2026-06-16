# apps/worker-py/tests/test_xlsx_parser.py
import io
from openpyxl import Workbook
from app.parsers.xlsx import parse_xlsx_bytes

def _make_xlsx(headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Invoice'
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def test_parses_basic_invoice_lines():
    raw = _make_xlsx(
        ['Description', 'Qty', 'Rate', 'Amount', 'Currency', 'Shipment Ref', 'Job No', 'Unit'],
        [
            ['TRUCKING', 2, 50.0, 100.0, 'AED', 'HV-001', 'J-123', 'PER_MT'],
            ['THC',      1, 75.0,  75.0, 'AED', 'HV-001', 'J-124', 'PER_TEU'],
        ]
    )
    ni = parse_xlsx_bytes(raw, file_id='f1', file_name='inv.xlsx', parser_version='parser-0.1.0')
    assert len(ni.invoice_lines) == 2
    l0 = ni.invoice_lines[0]
    assert l0.description == 'TRUCKING' and l0.qty == 2 and l0.rate == 50.0 and l0.amount == 100.0
    assert l0.currency == 'AED'
    assert l0.shipment_ref == 'HV-001'
    assert l0.job_number == 'J-123'
    assert l0.rate_basis == 'PER_MT'
    assert l0.for_charge_component is None
    assert ni.invoice_header.currency == 'AED'
    assert ni.parser_version == 'parser-0.1.0'

def test_supports_header_aliases_case_insensitive():
    raw = _make_xlsx(
        ['desc', 'quantity', 'unit rate', 'line amount', 'CCY'],
        [['THC', 1, 75.0, 75.0, 'USD']]
    )
    ni = parse_xlsx_bytes(raw, file_id='f1', file_name='inv.xlsx', parser_version='parser-0.1.0')
    assert ni.invoice_lines[0].currency == 'USD'
    assert ni.invoice_lines[0].rate == 75.0
    assert ni.invoice_lines[0].amount == 75.0

def test_skips_empty_rows_and_uses_normalized_line_id():
    raw = _make_xlsx(['Description','Qty','Rate','Amount','Currency'], [[None,None,None,None,None], ['TRUCKING', 1, 10, 10, 'AED']])
    ni = parse_xlsx_bytes(raw, file_id='f1', file_name='inv.xlsx', parser_version='parser-0.1.0')
    assert len(ni.invoice_lines) == 1
    assert ni.invoice_lines[0].line_id.startswith('line_')

def test_raises_when_no_usable_line():
    import pytest
    raw = _make_xlsx(['Description','Qty','Rate','Amount','Currency'], [[None,None,None,None,None]])
    with pytest.raises(ValueError):
        parse_xlsx_bytes(raw, file_id='f1', file_name='inv.xlsx', parser_version='parser-0.1.0')

def test_extracts_header_fields():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Invoice'
    ws.append(['Invoice No:', 'INV-2026-001'])
    ws.append(['Vendor:', 'DSV Solutions'])
    ws.append(['Date:', '02/06/2026'])
    ws.append([])
    ws.append(['Description', 'Qty', 'Rate', 'Amount'])
    ws.append(['THC', 1, 75.0, 75.0])
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()
    ni = parse_xlsx_bytes(raw, file_id='f1', file_name='inv.xlsx', parser_version='parser-0.1.0')
    assert ni.invoice_header.invoice_no == 'INV-2026-001'
    assert ni.invoice_header.vendor == 'DSV Solutions'
    assert ni.invoice_header.issue_date == '02/06/2026'
    assert len(ni.invoice_lines) == 1

def test_extracts_header_fields_alt_labels():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Invoice'
    ws.append(['Inv No', 'INV-999'])
    ws.append(['Supplier', 'Mammoet'])
    ws.append(['Issue Date', '2026-03-15'])
    ws.append([])
    ws.append(['Description', 'Amount'])
    ws.append(['TRANSPORT', 500.0])
    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()
    ni = parse_xlsx_bytes(raw, file_id='f2', file_name='inv2.xlsx', parser_version='parser-0.1.0')
    assert ni.invoice_header.invoice_no == 'INV-999'
    assert ni.invoice_header.vendor == 'Mammoet'
    assert ni.invoice_header.issue_date == '2026-03-15'

def test_header_fields_none_when_missing():
    raw = _make_xlsx(
        ['Description', 'Qty', 'Rate', 'Amount', 'Currency'],
        [['TRUCKING', 2, 50.0, 100.0, 'AED']]
    )
    ni = parse_xlsx_bytes(raw, file_id='f1', file_name='inv.xlsx', parser_version='parser-0.1.0')
    assert ni.invoice_header.invoice_no is None
    assert ni.invoice_header.vendor is None
    assert ni.invoice_header.issue_date is None
    assert len(ni.invoice_lines) == 1

def test_parses_line_view_when_decision_sheet_is_first():
    wb = Workbook()
    ws = wb.active
    ws.title = '00_Decision'
    ws.append(['job_id', 'verdict'])
    ws.append(['job_1', 'AMBER'])
    ws = wb.create_sheet('04_Line_View')
    ws.append([
        'line_id', 'shipment_ref', 'description', 'for_charge_component',
        'type_b', 'amount', 'currency'
    ])
    ws.append([
        'line_1', 'HVDC-001', 'MASTER DO CHARGE', 'MASTER DO CHARGE',
        None, 80.0, 'USD'
    ])
    buf = io.BytesIO()
    wb.save(buf)

    ni = parse_xlsx_bytes(
        buf.getvalue(),
        file_id='audit_pack',
        file_name='audit_pack.xlsx',
        parser_version='parser-0.1.0',
    )

    assert len(ni.invoice_lines) == 1
    assert ni.invoice_lines[0].description == 'MASTER DO CHARGE'
    assert ni.invoice_lines[0].amount == 80.0
    assert ni.invoice_lines[0].currency == 'USD'
    assert ni.invoice_lines[0].source_ref['sheet'] == '04_Line_View'

def test_parses_shpiment_v32_line_view_total_usd_contract():
    wb = Workbook()
    ws = wb.active
    ws.title = '00_Decision'
    ws.append(['Overall Verdict'])
    ws.append(['AMBER'])
    ws = wb.create_sheet('03_Type_B_Summary')
    ws.append(['Shipment_No', 'Customs', 'DO', 'INLAND', 'THC', 'Inspection', 'Detention', 'STROAGE', 'OTHERS', 'Total_AED', 'Total_USD', 'Line_Count'])
    ws = wb.create_sheet('04_Line_View')
    ws.append([
        'Shipment_No', 'Source_Row_ID', 'Rate_Source', 'Description',
        'Total_AED', 'Total_USD', 'TYPE_B', 'Evidence_Status'
    ])
    ws.append([
        'HVDC-001', 'SRC-001', 'CONTRACT', 'Document delivery order fee',
        367.30, 100.0, 'DO', 'AUTO_VERIFIED'
    ])
    buf = io.BytesIO()
    wb.save(buf)

    ni = parse_xlsx_bytes(
        buf.getvalue(),
        file_id='shpiment_pack',
        file_name='DSV_Audit_Pack.xlsx',
        parser_version='parser-0.1.0',
    )

    assert len(ni.invoice_lines) == 1
    assert ni.invoice_lines[0].shipment_ref == 'HVDC-001'
    assert ni.invoice_lines[0].description == 'Document delivery order fee'
    assert ni.invoice_lines[0].amount == 100.0
    assert ni.invoice_lines[0].currency == 'USD'
    assert ni.invoice_lines[0].for_charge_component == 'DO'
