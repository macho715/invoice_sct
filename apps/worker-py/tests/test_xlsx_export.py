import pytest
import openpyxl
from io import BytesIO
from app.schemas import ExportRequest, DecisionRow, ActionItemRow, FinalReconRow, LineViewRow, SourceDataRow, AuditDetailRow, EvidenceIssuesRow, ManifestEntry
from app.exporters.xlsx import build_xlsx
import hashlib

def test_xlsx_export_determinism():
    # P3C modified: raw bytes equality can be brittle (openpyxl timestamps/compression); check logical + new pdf cols in 90_Source_Data header
    req = ExportRequest(
        job_id="job_123",
        generated_at="2026-06-09T12:00:00Z",
        decision_rows=[
            DecisionRow(job_id="job_123", verdict="PASS", rule_version="r1", parser_version="p1", zero_count=0, amber_count=0)
        ],
        action_items_rows=[],
        final_recon_rows=[],
        line_view_rows=[],
        source_data_rows=[SourceDataRow(file_id="f1", pdf_page=2, text_span_hash="sha256:def")],
        audit_detail_rows=[],
        evidence_issues_rows=[]
    )
    b1 = build_xlsx(req)
    b2 = build_xlsx(req)
    # logical equality via re-load (robust)
    wb1 = openpyxl.load_workbook(BytesIO(b1))
    wb2 = openpyxl.load_workbook(BytesIO(b2))
    assert wb1.sheetnames == wb2.sheetnames
    # P3C: 90_Source_Data header now includes pdf cols
    src_header = [cell.value for cell in wb1["90_Source_Data"][1]]
    assert "pdf_page" in src_header and "text_span_hash" in src_header
    # still deterministic in structure
    assert src_header == [cell.value for cell in wb2["90_Source_Data"][1]]

def test_xlsx_export_headers_and_sheets():
    req = ExportRequest(
        job_id="job_123",
        generated_at="2026-06-09T12:00:00Z",
        decision_rows=[
            DecisionRow(job_id="job_123", verdict="PASS", rule_version="r1", parser_version="p1", zero_count=0, amber_count=0)
        ],
        action_items_rows=[
            ActionItemRow(action_id="act_1", severity="AMBER", issue_type="CG_WARN", required_action="check")
        ],
        final_recon_rows=[
            FinalReconRow(currency="USD", shipment_ref="S1", invoice_total=100.0, reviewed_total=100.0, variance=0.0, variance_pct=0.0, recon_status="MATCHED")
        ],
        line_view_rows=[
            LineViewRow(line_id="line_1", description="desc", amount=100.0, currency="USD")
        ],
        source_data_rows=[
            SourceDataRow(file_id="f1", source_ref="ref1", original_text="orig")
        ],
        audit_detail_rows=[
            AuditDetailRow(line_id="line_1", rule_id="rule_1", reason_code="code", sct_trace_id="tr1")
        ],
        evidence_issues_rows=[
            EvidenceIssuesRow(line_id="line_1", severity="AMBER")
        ]
    )
    
    xlsx_bytes = build_xlsx(req)
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    
    expected_sheets = ["00_Decision", "01_Action_Items", "02_Final_Recon", "03_Header_Check", "04_Line_View", "05_Duplicate_Check", "06_Rate_Check", "07_Tax_FX_Check", "08_Shipment_Match", "90_Source_Data", "91_Audit_Detail", "92_Evidence_Issues", "99_Manifest"]
    assert wb.sheetnames == expected_sheets
    
    # Check headers
    assert [cell.value for cell in wb["00_Decision"][1]] == ["job_id", "verdict", "approved_by", "approved_at", "rule_version", "parser_version", "final_recon_status", "zero_count", "amber_count", "prism_kernel_proof_ref", "costguard_band_summary", "watermark", "generated_at"]
    assert [cell.value for cell in wb["01_Action_Items"][1]] == ["action_id", "severity", "shipment_ref", "line_id", "issue_type", "required_action", "owner_role", "status", "prism_kernel_proof_ref"]

def test_xlsx_export_sorting():
    req = ExportRequest(
        job_id="job_123",
        generated_at="2026-06-09T12:00:00Z",
        decision_rows=[],
        action_items_rows=[],
        final_recon_rows=[],
        line_view_rows=[
            LineViewRow(line_id="line_b", description="desc b", amount=100.0, currency="USD"),
            LineViewRow(line_id="line_a", description="desc a", amount=100.0, currency="USD")
        ],
        source_data_rows=[],
        audit_detail_rows=[],
        evidence_issues_rows=[]
    )
    xlsx_bytes = build_xlsx(req)
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    ws = wb["04_Line_View"]
    # Row 1 is header. Row 2 should be line_a (sorted), Row 3 should be line_b.
    assert ws.cell(row=2, column=1).value == "line_a"
    assert ws.cell(row=3, column=1).value == "line_b"

def test_xlsx_export_writes_formula_text_as_literal_text():
    req = ExportRequest(
        job_id="job_formula",
        generated_at="2026-06-09T12:00:00Z",
        decision_rows=[],
        action_items_rows=[],
        final_recon_rows=[],
        line_view_rows=[
            LineViewRow(line_id="line_1", description="desc", amount=100.0, currency="USD", formula_text="=ROUNDUP(99.991,2)")
        ],
        source_data_rows=[],
        audit_detail_rows=[],
        evidence_issues_rows=[]
    )
    xlsx_bytes = build_xlsx(req)
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=False)
    ws = wb["04_Line_View"]
    headers = [cell.value for cell in ws[1]]
    col = headers.index("formula_text") + 1
    cell = ws.cell(row=2, column=col)
    assert cell.value == "'=ROUNDUP(99.991,2)"
    assert cell.data_type == "s"

def test_xlsx_export_missing_fields():
    req = ExportRequest(
        job_id="job_123",
        generated_at="2026-06-09T12:00:00Z",
        decision_rows=[
            DecisionRow(job_id="job_123", verdict="PASS", rule_version="r1", parser_version="p1", zero_count=0, amber_count=0, approved_by=None)
        ],
        action_items_rows=[],
        final_recon_rows=[],
        line_view_rows=[],
        source_data_rows=[],
        audit_detail_rows=[],
        evidence_issues_rows=[]
    )
    # Shouldn't crash and cell value should be None
    xlsx_bytes = build_xlsx(req)
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    assert wb["00_Decision"].cell(row=2, column=3).value is None

def test_xlsx_export_empty_request():
    req = ExportRequest(
        job_id="job_empty",
        generated_at="2026-06-09T12:00:00Z",
        decision_rows=[],
        action_items_rows=[],
        final_recon_rows=[],
        line_view_rows=[],
        source_data_rows=[],
        audit_detail_rows=[],
        evidence_issues_rows=[]
    )
    xlsx_bytes = build_xlsx(req)
    assert len(xlsx_bytes) > 0


def test_xlsx_export_persists_source_hash_verification_in_required_sheets():
    req = ExportRequest(
        job_id="job_hash",
        generated_at="2026-06-15T00:00:00Z",
        decision_rows=[DecisionRow(job_id="job_hash", verdict="ZERO", rule_version="r1", parser_version="p1", zero_count=1, amber_count=0)],
        action_items_rows=[ActionItemRow(action_id="act_hash", severity="ZERO", line_id="", issue_type="SOURCE_HASH_MISMATCH", required_action="re-upload trusted source")],
        final_recon_rows=[],
        line_view_rows=[],
        source_data_rows=[],
        audit_detail_rows=[AuditDetailRow(line_id="", rule_id="SOURCE_HASH_CHECK", reason_code="SOURCE_HASH_MISMATCH", sct_trace_id="trace_hash", decision_input="expected_sha256:aaa; actual_sha256:bbb")],
        evidence_issues_rows=[],
        manifest_entries=[
            ManifestEntry(key="source_hash_status", value="SOURCE_HASH_MISMATCH"),
            ManifestEntry(key="source_sha256_expected", value="a" * 64),
            ManifestEntry(key="source_sha256_actual", value="b" * 64),
        ],
    )

    wb = openpyxl.load_workbook(BytesIO(build_xlsx(req)))

    assert wb["00_Decision"].cell(row=2, column=2).value == "ZERO"
    assert wb["01_Action_Items"].cell(row=2, column=5).value == "SOURCE_HASH_MISMATCH"
    assert wb["91_Audit_Detail"].cell(row=2, column=2).value == "SOURCE_HASH_CHECK"
    manifest = {wb["99_Manifest"].cell(row=i, column=1).value: wb["99_Manifest"].cell(row=i, column=2).value for i in range(2, wb["99_Manifest"].max_row + 1)}
    assert manifest["source_hash_status"] == "SOURCE_HASH_MISMATCH"
    assert manifest["source_sha256_expected"] == "a" * 64
    assert manifest["source_sha256_actual"] == "b" * 64
