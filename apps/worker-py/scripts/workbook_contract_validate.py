#!/usr/bin/env python3
"""Authoritative workbook contract validator for HVDC audit packs.

Resolves Q-006 (authoritative validator command).

Validates the sheet contract of an HVDC audit workbook:
- Track 1 (DSV v3.2 PRO): 8-Sheet contract
- Track 2 (Invoice Audit MVP): 13-Sheet contract (default)

Strict checks:
- sheet count must match expected
- sheet names must match expected exactly (case-sensitive, no extra/missing)
- sheet order must match expected (no reordering)
- hidden sheets are not allowed
- extra sheets are not allowed

Usage:
    python workbook_contract_validate.py <file.xlsx> [--track {1,2}]

Exit codes:
    0 = PASS (contract satisfied)
    1 = FAIL (contract violation or read error)
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        json.dumps(
            {
                "check": "workbook_contract_validate",
                "status": "ERROR",
                "errors": [
                    "openpyxl not installed. Install via `pip install openpyxl`."
                ],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    sys.exit(1)


TRACK1_SHEETS: list[str] = [
    "00_Decision",
    "01_Action_Items",
    "02_Final_Recon",
    "03_Type_B_Summary",
    "04_Line_View",
    "90_Source_Data",
    "91_Audit_Detail",
    "92_Evidence_Issues",
]

TRACK2_SHEETS: list[str] = [
    "00_Decision",
    "01_Action_Items",
    "02_Final_Recon",
    "03_Header_Check",
    "04_Line_View",
    "05_Duplicate_Check",
    "06_Rate_Check",
    "07_Tax_FX_Check",
    "08_Shipment_Match",
    "90_Source_Data",
    "91_Audit_Detail",
    "92_Evidence_Issues",
    "99_Manifest",
]

TRACKS: dict[int, list[str]] = {
    1: TRACK1_SHEETS,
    2: TRACK2_SHEETS,
}


def _diff_sets(expected: Iterable[str], actual: Iterable[str]) -> dict[str, list[str]]:
    exp, act = set(expected), set(actual)
    return {
        "missing": sorted(exp - act),
        "extra": sorted(act - exp),
    }


def _detect_hidden_sheet_names(wb) -> list[str]:
    """Return names of sheets whose state is not 'visible'."""
    hidden = []
    for ws in wb.worksheets:
        # openpyxl exposes sheet_state on each worksheet; default is 'visible'.
        if getattr(ws, "sheet_state", "visible") != "visible":
            hidden.append(ws.title)
    return hidden


def validate(path: Path, track: int) -> dict:
    expected = TRACKS.get(track)
    if expected is None:
        return {
            "check": "workbook_contract_validate",
            "status": "ERROR",
            "workbook": str(path),
            "track": track,
            "errors": [f"unknown track: {track!r}. Use 1 or 2."],
        }

    result: dict = {
        "check": "workbook_contract_validate",
        "workbook": str(path),
        "track": track,
        "expected_sheets": expected,
        "expected_sheet_count": len(expected),
    }

    if not path.exists():
        result["status"] = "FAIL"
        result["errors"] = [f"file_not_found: {path}"]
        return result

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001
        result["status"] = "FAIL"
        result["errors"] = [f"invalid_xlsx: {exc}"]
        return result

    actual_sheets = list(wb.sheetnames)
    hidden_sheets = _detect_hidden_sheet_names(wb)

    result["actual_sheets"] = actual_sheets
    result["actual_sheet_count"] = len(actual_sheets)
    result["hidden_sheets"] = hidden_sheets

    errors: list[str] = []

    # 1. Sheet count
    if len(actual_sheets) != len(expected):
        errors.append(
            f"sheet_count_mismatch: expected {len(expected)}, got {len(actual_sheets)}"
        )

    # 2. Sheet order
    if actual_sheets != expected:
        errors.append(
            f"sheet_order_mismatch: expected {expected!r}, got {actual_sheets!r}"
        )

    # 3. Missing / extra (set diff for granular diagnostics)
    diff = _diff_sets(expected, actual_sheets)
    if diff["missing"]:
        errors.append(f"missing_sheets: {diff['missing']}")
    if diff["extra"]:
        errors.append(f"extra_sheets: {diff['extra']}")

    # 4. Hidden sheets forbidden
    if hidden_sheets:
        errors.append(f"hidden_sheets_detected: {hidden_sheets}")

    result["status"] = "PASS" if not errors else "FAIL"
    result["errors"] = errors
    return result


def _format_summary(result: dict) -> str:
    lines: list[str] = []
    status = result.get("status", "UNKNOWN")
    track = result.get("track")
    expected_count = result.get("expected_sheet_count")
    actual_count = result.get("actual_sheet_count", 0)
    expected = result.get("expected_sheets", [])

    lines.append("=" * 72)
    lines.append(f"Workbook Contract Validator  (Track {track})")
    lines.append("=" * 72)
    lines.append(f"Workbook : {result.get('workbook')}")
    lines.append(f"Status   : {status}")
    lines.append(
        f"Sheets   : {actual_count} found / {expected_count} expected"
    )

    actual_sheets_list = result.get("actual_sheets", []) or []
    if status != "PASS":
        lines.append("-" * 72)
        lines.append("Contract Violations:")
        for err in result.get("errors", []):
            lines.append(f"  - {err}")
        lines.append("-" * 72)
        lines.append("Expected sheet contract (in order):")
        for i, name in enumerate(expected, start=1):
            present = i - 1 < len(actual_sheets_list) and actual_sheets_list[i - 1] == name
            marker = "OK " if present else "  "
            lines.append(f"  {i:2d}. {marker} {name}")
    else:
        lines.append("-" * 72)
        lines.append("All 13 sheets present in correct order. Contract satisfied.")
        lines.append("-" * 72)
        for i, name in enumerate(expected, start=1):
            lines.append(f"  {i:2d}. OK {name}")

    lines.append("=" * 72)
    return "\n".join(lines)


def main() -> int:
    ap = argparse.ArgumentParser(
        description=(
            "Validate HVDC audit workbook sheet contract "
            "(Track 1: 8 sheets, Track 2: 13 sheets)."
        )
    )
    ap.add_argument(
        "workbook",
        help="Path to the .xlsx workbook to validate.",
    )
    ap.add_argument(
        "--track",
        type=int,
        choices=[1, 2],
        default=2,
        help="Contract track: 1 = 8-Sheet (DSV v3.2), 2 = 13-Sheet (MVP). Default: 2.",
    )
    ap.add_argument(
        "--json",
        action="store_true",
        help="Emit machine-readable JSON in addition to the summary.",
    )
    args = ap.parse_args()

    path = Path(args.workbook)
    result = validate(path, track=args.track)

    print(_format_summary(result))
    if args.json:
        print("---JSON---")
        print(json.dumps(result, ensure_ascii=False, indent=2))

    return 0 if result.get("status") == "PASS" else 1


if __name__ == "__main__":
    sys.exit(main())
